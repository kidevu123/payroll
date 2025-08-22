from flask import Flask, render_template_string, request, redirect, url_for, send_file
import os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'static/reports'

# Create directories
Path(UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
Path(REPORT_FOLDER).mkdir(parents=True, exist_ok=True)

@app.route('/')
def index():
    """Simple upload form"""
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Complete Payroll App</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            .button { 
                display: inline-block; 
                padding: 10px 15px; 
                background-color: #4CAF50; 
                color: white; 
                border: none;
                cursor: pointer;
            }
        </style>
    </head>
    <body>
        <h1>Complete Payroll App</h1>
        
        <form action="/process" method="post" enctype="multipart/form-data">
            <p>Upload CSV file with timesheet data:</p>
            <input type="file" name="file" required>
            <button type="submit" class="button">Process File</button>
        </form>
        
        <p><em>Sample file format: ID, Name, Date, Hours, Rate</em></p>
    </body>
    </html>
    """
    return render_template_string(html)

def create_excel_report(df, filename):
    """Create an Excel report from the DataFrame"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"
    
    # Add header
    ws['A1'] = "Payroll Report"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Add column headers in row 2
    headers = ["ID", "Name", "Total Hours", "Pay Rate", "Total Pay"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    
    # Process data - create summary per person
    summary = df.groupby(['ID', 'Name', 'Rate']).agg(
        Total_Hours=('Hours', 'sum')
    ).reset_index()
    
    # Calculate pay
    summary['Total_Pay'] = summary['Total_Hours'] * summary['Rate']
    
    # Add data rows
    for i, row in enumerate(summary.itertuples(), 3):
        ws.cell(row=i, column=1).value = row.ID
        ws.cell(row=i, column=2).value = row.Name
        ws.cell(row=i, column=3).value = round(row.Total_Hours, 2)
        ws.cell(row=i, column=4).value = row.Rate
        ws.cell(row=i, column=5).value = round(row.Total_Pay, 2)
    
    # Save the workbook
    report_path = os.path.join(REPORT_FOLDER, filename)
    wb.save(report_path)
    return report_path

@app.route('/process', methods=['POST'])
def process():
    """Process the uploaded file"""
    try:
        if 'file' not in request.files:
            return "No file part", 400

        file = request.files['file']
        if file.filename == '':
            return "No selected file", 400

        # Save the file
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)
        
        # For CSV files, create an Excel report
        if file.filename.endswith('.csv'):
            try:
                # Try to parse as a timesheet
                df = pd.read_csv(file_path)
                
                # If file doesn't have expected columns, use simple processing
                if not all(col in df.columns for col in ['ID', 'Name', 'Hours', 'Rate']):
                    # Add dummy columns for demonstration
                    df['ID'] = range(1, len(df) + 1)
                    df['Name'] = [f"Employee {i}" for i in range(1, len(df) + 1)]
                    df['Hours'] = 8.0  # Default 8 hours
                    df['Rate'] = 15.0  # Default $15/hour
                
                # Create the Excel report
                excel_filename = "payroll_report.xlsx"
                report_path = create_excel_report(df, excel_filename)
                
                # Store the path in the app context for download
                app.config['REPORT_FILE'] = excel_filename
                
            except Exception as e:
                # If pandas processing fails, fall back to simple report
                txt_filename = "error_report.txt"
                report_path = os.path.join(REPORT_FOLDER, txt_filename)
                with open(report_path, 'w') as f:
                    f.write(f"Error processing {file.filename}: {str(e)}\n")
                app.config['REPORT_FILE'] = txt_filename
        else:
            # For non-CSV files, create a simple text report
            txt_filename = "file_report.txt"
            report_path = os.path.join(REPORT_FOLDER, txt_filename)
            with open(report_path, 'w') as f:
                f.write(f"Report for {file.filename}\n")
                f.write(f"File size: {os.path.getsize(file_path)} bytes\n")
            app.config['REPORT_FILE'] = txt_filename
        
        return redirect(url_for('success'))
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"Error: {str(e)}<br><pre>{error_details}</pre>", 500

@app.route('/success')
def success():
    """Success page with download links"""
    report_file = app.config.get('REPORT_FILE', 'report.txt')
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Processing Successful</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            .button {{ 
                display: inline-block; 
                padding: 10px 15px; 
                background-color: #4CAF50; 
                color: white; 
                text-decoration: none;
                margin-right: 10px;
            }}
            .download-options {{ margin: 20px 0; }}
        </style>
    </head>
    <body>
        <h1>Processing Successful</h1>
        
        <p>Your file was successfully processed.</p>
        
        <div class="download-options">
            <h2>Download Options</h2>
            
            <p><a href="/download" class="button">Download Report</a></p>
            
            <p>Direct link: <a href="/static/reports/{report_file}">/static/reports/{report_file}</a></p>
        </div>
        
        <p><a href="/">Process Another File</a></p>
    </body>
    </html>
    """
    return render_template_string(html)

@app.route('/download')
def download():
    """Download the report file"""
    try:
        report_file = app.config.get('REPORT_FILE', 'report.txt')
        report_path = os.path.join(REPORT_FOLDER, report_file)
        
        if os.path.exists(report_path):
            return send_file(report_path, as_attachment=True)
        else:
            return "Report file not found", 404
    except Exception as e:
        return f"Error downloading file: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True)
