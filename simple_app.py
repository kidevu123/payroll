import os
import pandas as pd
from pathlib import Path
from flask import Flask, request, send_file, render_template_string, redirect, url_for, jsonify, session, flash, get_flashed_messages
from werkzeug.security import generate_password_hash, check_password_hash
from markupsafe import escape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime, timedelta
from functools import wraps
import sqlite3
import hashlib
from collections import defaultdict
import requests
from bs4 import BeautifulSoup
import time
import logging
from logging.handlers import RotatingFileHandler
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO
# Selenium imports removed - not supported on PythonAnywhere

# Import centralized version management
from version import get_version, get_version_display, get_version_info

app = Flask(__name__)
# Use environment variable for secret key, fallback to random for development
app.secret_key = os.getenv('FLASK_SECRET_KEY', os.urandom(24).hex())
if not os.getenv('FLASK_SECRET_KEY'):
    import warnings
    warnings.warn(
        "WARNING: FLASK_SECRET_KEY not set in environment. "
        "Using random key - sessions will not persist across restarts. "
        "Set FLASK_SECRET_KEY environment variable for production.",
        RuntimeWarning
    )
# Use centralized version management
APP_VERSION = get_version()

# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
# Set up application logging for error tracking and debugging

# Create logs directory if it doesn't exist
LOG_FOLDER = 'logs'
Path(LOG_FOLDER).mkdir(parents=True, exist_ok=True)

# Configure logging
log_file = os.path.join(LOG_FOLDER, 'payroll_app.log')
file_handler = RotatingFileHandler(log_file, maxBytes=10485760, backupCount=5)  # 10MB per file, keep 5 backups
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
file_handler.setFormatter(file_formatter)

# Add handler to app logger
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)

app.logger.info(f"Payroll application started - Version {APP_VERSION}")

# Configuration
UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'static/reports'
CONFIG_FILE = 'pay_rates.json'
USERS_FILE = 'users.json'
DATABASE = 'payroll.db'
PAY_RATES_FILE = 'pay_rates.csv'
MISSING_TIMES_FILE = 'missing_times.csv'

# Caching for report data
report_cache = {}
report_cache_expiry = {}

# Caching for pay rates (reduces file I/O)
pay_rates_cache = None
pay_rates_cache_time = None
PAY_RATES_CACHE_TTL = 300  # Cache for 5 minutes

REPORTS_METADATA_FILE = os.path.join(REPORT_FOLDER, 'reports_metadata.json')
REPORTS_LIST_LIMIT = int(os.getenv('REPORTS_LIST_LIMIT', '24'))

def _load_reports_metadata() -> dict:
    try:
        if os.path.exists(REPORTS_METADATA_FILE):
            with open(REPORTS_METADATA_FILE, 'r') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_reports_metadata(meta: dict) -> None:
    try:
        tmp = REPORTS_METADATA_FILE + '.tmp'
        with open(tmp, 'w') as f:
            json.dump(meta, f)
        os.replace(tmp, REPORTS_METADATA_FILE)
    except Exception:
        pass

def _ensure_report_metadata(file_path: str, filename: str, meta: dict) -> dict:
    """Ensure metadata for a report file exists and is up-to-date. Returns the record."""
    try:
        mtime = os.path.getmtime(file_path)
        rec = meta.get(filename)
        if rec and abs(rec.get('mtime', 0) - mtime) < 0.1:
            return rec

        # Extract minimal info once
        from openpyxl import load_workbook
        creator, total_amount, date_range = 'Unknown', None, None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            # Creator
            try:
                # Try AA1 first (where username is stored)
                if ws['AA1'].value:
                    creator = str(ws['AA1'].value)
                # Fallback to A2 if AA1 is empty
                elif ws['A2'].value and 'Processed by:' in str(ws['A2'].value):
                    creator = str(ws['A2'].value).replace('Processed by:', '').strip()
            except Exception:
                pass
            # Date Range: Extract from A1 cell (e.g., "Payroll Summary - 2025-01-04 to 2025-01-10")
            try:
                if ws['A1'].value:
                    import re
                    a1_value = str(ws['A1'].value)
                    date_range_match = re.search(r'(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})', a1_value)
                    if date_range_match:
                        date_range = f"{date_range_match.group(1)} to {date_range_match.group(2)}"
            except Exception:
                pass
            # Amount: search first 30 rows for GRAND TOTAL and pick rightmost numeric
            try:
                max_rows = min(ws.max_row, 40)
                for r in range(3, max_rows + 1):
                    row_text = ''.join([str(ws.cell(row=r, column=c).value or '') for c in range(1, min(ws.max_column, 18))])
                    if 'GRAND TOTAL' in row_text.upper():
                        for c in range(min(ws.max_column, 20), 1, -1):
                            val = ws.cell(row=r, column=c).value
                            if isinstance(val, (int, float)) and val > 0:
                                total_amount = float(val)
                                break
                        break
            except Exception:
                pass
        except Exception:
            pass

        rec = {
            'mtime': mtime,
            'creator': creator,
            'total_amount': total_amount,
            'date_range': date_range,
        }
        meta[filename] = rec
        return rec
    except Exception:
        return meta.get(filename) or {}


# ═══════════════════════════════════════════════════════════════════════════════
# ZOHO BOOKS INTEGRATION
# ═══════════════════════════════════════════════════════════════════════════════
# Configuration and functions for Zoho Books API integration
# Handles expense creation, receipt attachment, and account management
# Configuration is read from environment variables so secrets are not stored in code.
# For each company (haute, boomin) set:
#   ZB_<COMPANY>_ORG_ID, ZB_<COMPANY>_CLIENT_ID, ZB_<COMPANY>_CLIENT_SECRET, ZB_<COMPANY>_REFRESH_TOKEN
# Optional helpers (recommended to avoid extra lookups):
#   ZB_<COMPANY>_EXPENSE_ACCOUNT_NAME  (e.g., "5300 Payroll Expenses")
#   ZB_<COMPANY>_EXPENSE_ACCOUNT_ID    (Chart of Accounts ID for the expense account)
#   ZB_<COMPANY>_PAID_THROUGH_NAME     (Bank/Cash account name shown in Zoho)
#   ZB_<COMPANY>_PAID_THROUGH_ID       (Bank/Cash account ID)
#   ZB_<COMPANY>_VENDOR_ID             (Optional: vendor/contact id to tag)
# Global options:
#   ZB_DOMAIN (default https://www.zohoapis.com)
ZB_DOMAIN = os.getenv('ZB_DOMAIN', 'https://www.zohoapis.com')
ZB_ACCOUNTS_DOMAIN = os.getenv('ZB_ACCOUNTS_DOMAIN', 'https://accounts.zoho.com')

# In-memory token cache to avoid refreshing every call
zoho_token_cache = {  # company -> {access_token, expires_at}
}

def get_zoho_company_key(company_raw):
    company = (company_raw or '').strip().lower()
    if company in ('haute', 'haute-brands', 'hautebrands', 'haute_brands'):
        return 'HAUTE'
    if company in ('boomin', 'boomin-brands', 'boominbrands', 'boomin_brands', 'boominbrand', 'boomin_brand'):
        return 'BOOMIN'
    return None

def get_zoho_company_cfg(company_raw):
    key = get_zoho_company_key(company_raw)
    if not key:
        return None
    prefix = f'ZB_{key}_'
    cfg = {
        'org_id': os.getenv(prefix + 'ORG_ID', '').strip(),
        'client_id': os.getenv(prefix + 'CLIENT_ID', '').strip(),
        'client_secret': os.getenv(prefix + 'CLIENT_SECRET', '').strip(),
        'refresh_token': os.getenv(prefix + 'REFRESH_TOKEN', '').strip(),
        'expense_account_id': os.getenv(prefix + 'EXPENSE_ACCOUNT_ID', '').strip(),
        'expense_account_name': os.getenv(prefix + 'EXPENSE_ACCOUNT_NAME', '').strip(),
        'paid_through_id': os.getenv(prefix + 'PAID_THROUGH_ID', '').strip(),
        'paid_through_name': os.getenv(prefix + 'PAID_THROUGH_NAME', '').strip(),
        'vendor_id': os.getenv(prefix + 'VENDOR_ID', '').strip()
    }
    # Basic validation
    if not (cfg['org_id'] and cfg['client_id'] and cfg['client_secret'] and cfg['refresh_token']):
        return None
    return cfg

def zoho_refresh_access_token(company_raw):
    """Refresh and cache access token for a company using its refresh token with retry logic."""
    cfg = get_zoho_company_cfg(company_raw)
    if not cfg:
        error_msg = f'Zoho Books credentials not configured for company: {company_raw}'
        app.logger.error(error_msg)
        raise ValueError(error_msg)

    # Return cached token if valid for at least 60 seconds
    cached = zoho_token_cache.get(company_raw)
    if cached and cached.get('expires_at', 0) - time.time() > 60:
        app.logger.debug(f"Using cached Zoho token for {company_raw}")
        return cached['access_token']

    token_url = f'{ZB_ACCOUNTS_DOMAIN}/oauth/v2/token'
    params = {
        'refresh_token': cfg['refresh_token'],
        'client_id': cfg['client_id'],
        'client_secret': cfg['client_secret'],
        'grant_type': 'refresh_token'
    }
    
    # Retry logic for network errors
    max_retries = 3
    for attempt in range(max_retries):
        try:
            app.logger.info(f"Refreshing Zoho token for {company_raw} (attempt {attempt + 1}/{max_retries})")
            resp = requests.post(token_url, params=params, timeout=20)
            
            if resp.status_code != 200:
                error_msg = f"Zoho token refresh failed: {resp.status_code} {resp.text}"
                app.logger.error(error_msg)
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                    continue
                raise RuntimeError(error_msg)
            
            data = resp.json()
            access_token = data.get('access_token')
            expires_in = int(data.get('expires_in', 3600))
            
            if not access_token:
                error_msg = 'Zoho token refresh returned no access_token'
                app.logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            zoho_token_cache[company_raw] = {
                'access_token': access_token,
                'expires_at': time.time() + expires_in
            }
            app.logger.info(f"Successfully refreshed Zoho token for {company_raw}")
            return access_token
            
        except requests.exceptions.Timeout as e:
            app.logger.warning(f"Zoho API timeout (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise RuntimeError(f"Zoho API timeout after {max_retries} attempts")
        except requests.exceptions.ConnectionError as e:
            app.logger.warning(f"Zoho API connection error (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise RuntimeError(f"Cannot connect to Zoho API after {max_retries} attempts")
        except Exception as e:
            app.logger.error(f"Unexpected error refreshing Zoho token: {e}")
            raise

def zoho_headers(company_raw):
    access_token = zoho_refresh_access_token(company_raw)
    return {
        'Authorization': f'Zoho-oauthtoken {access_token}',
        'Content-Type': 'application/json'
    }

def zoho_find_account_id_by_name(company_raw, account_name):
    """Find Chart of Accounts account_id by name; returns None if not found."""
    cfg = get_zoho_company_cfg(company_raw)
    url = f"{ZB_DOMAIN}/books/v3/chartofaccounts?organization_id={cfg['org_id']}&filter_by=AccountType.All&search_text={requests.utils.quote(account_name)}"
    resp = requests.get(url, headers=zoho_headers(company_raw), timeout=20)
    if resp.status_code != 200:
        return None
    for acc in resp.json().get('chartofaccounts', []) or []:
        if str(acc.get('account_name', '')).strip().lower() == account_name.strip().lower():
            return str(acc.get('account_id', '') or '')
    return None

def zoho_find_bank_account_id_by_name(company_raw, account_name):
    """Find bank/cash account id (paid through) by name; returns None if not found. Kept for potential future use."""
    cfg = get_zoho_company_cfg(company_raw)
    url = f"{ZB_DOMAIN}/books/v3/bankaccounts?organization_id={cfg['org_id']}&search_text={requests.utils.quote(account_name)}"
    resp = requests.get(url, headers=zoho_headers(company_raw), timeout=20)
    if resp.status_code != 200:
        return None
    for acc in resp.json().get('bankaccounts', []) or []:
        if str(acc.get('account_name', '')).strip().lower() == account_name.strip().lower():
            return str(acc.get('account_id', '') or '')
    return None

def compute_grand_totals_for_expense(df):
    """Recompute totals like our reports do, and return (total_hours, total_pay, total_rounded)."""
    pay_rates = load_pay_rates()
    df = df.copy()
    df['Daily Hours'] = df.apply(compute_daily_hours, axis=1)
    df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)
    df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)
    weekly_totals = df.groupby('Person ID').agg(
        Total_Hours=('Daily Hours', 'sum'),
        Weekly_Total=('Daily Pay', 'sum')
    ).reset_index()
    weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
    weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
    weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)
    total_hours = float(weekly_totals['Total_Hours'].sum().round(2)) if len(weekly_totals) else 0.0
    total_pay = float(weekly_totals['Weekly_Total'].sum().round(2)) if len(weekly_totals) else 0.0
    total_rounded = int(weekly_totals['Rounded_Weekly'].sum()) if len(weekly_totals) else 0
    return total_hours, total_pay, total_rounded

def compute_expense_date_from_data(week_str: str) -> str:
    """Compute expense posting date as end-of-week + 1 day.
    Preference: use the uploaded CSV's max(Date) + 1. Fallback: parse week_str and add 7 days.
    Returns YYYY-MM-DD string.
    """
    try:
        uploaded_file = session.get('uploaded_file')
        if uploaded_file and os.path.exists(uploaded_file):
            df = pd.read_csv(uploaded_file)
            if 'Date' in df.columns:
                try:
                    df['Date'] = pd.to_datetime(df['Date'])
                    post_date = (df['Date'].max() + pd.Timedelta(days=1)).date()
                    return post_date.strftime('%Y-%m-%d')
                except Exception:
                    pass
    except Exception:
        pass

    # Fallback: week_str (usually min date) + 7 days
    try:
        base = pd.to_datetime(week_str).date()
        return (base + timedelta(days=7)).strftime('%Y-%m-%d')
    except Exception:
        # Last resort: today's date
        return datetime.now().strftime('%Y-%m-%d')

def compute_week_range_strings(week_str: str):
    """Return (start_date_str, end_date_str) for the current run.
    Prefers actual min/max from uploaded CSV; falls back to week_str .. week_str+6.
    """
    try:
        uploaded_file = session.get('uploaded_file')
        if uploaded_file and os.path.exists(uploaded_file):
            df = pd.read_csv(uploaded_file)
            if 'Date' in df.columns:
                try:
                    df['Date'] = pd.to_datetime(df['Date'])
                    start = df['Date'].min().date()
                    end = df['Date'].max().date()
                    return start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')
                except Exception:
                    pass
    except Exception:
        pass

    try:
        base = pd.to_datetime(week_str).date()
        return base.strftime('%Y-%m-%d'), (base + timedelta(days=6)).strftime('%Y-%m-%d')
    except Exception:
        today = datetime.now().date()
        return today.strftime('%Y-%m-%d'), (today + timedelta(days=6)).strftime('%Y-%m-%d')

def _get_existing_expense(company: str, week: str):
    """Return previously created expense_id for company+week from session, if any."""
    try:
        key = f"{company}|{week}"
        mapping = session.get('zoho_expenses', {}) or {}
        return mapping.get(key)
    except Exception:
        return None

def _set_existing_expense(company: str, week: str, expense_id: str):
    """Persist expense id in session to avoid duplicate creations for same run."""
    try:
        key = f"{company}|{week}"
        mapping = session.get('zoho_expenses', {}) or {}
        mapping[key] = str(expense_id)
        session['zoho_expenses'] = mapping
    except Exception:
        pass

def _clear_existing_expense(company: str, week: str):
    """Remove stored expense id for company+week from session mapping."""
    try:
        key = f"{company}|{week}"
        mapping = session.get('zoho_expenses', {}) or {}
        if key in mapping:
            del mapping[key]
            session['zoho_expenses'] = mapping
    except Exception:
        pass

def zoho_get_expense(company_raw, expense_id: str):
    """Return expense JSON if it exists, else None."""
    try:
        cfg = get_zoho_company_cfg(company_raw)
        url = f"{ZB_DOMAIN}/books/v3/expenses/{expense_id}?organization_id={cfg['org_id']}"
        resp = requests.get(url, headers=zoho_headers(company_raw), timeout=20)
        if resp.status_code == 200:
            return resp.json()
        return None
    except Exception:
        return None

def zoho_find_expense_by_reference(company_raw, reference_number: str):
    """
    Search Zoho Books for an expense with the given reference number.
    Returns expense_id if found, else None.
    This provides robust duplicate prevention across sessions and users.
    """
    try:
        cfg = get_zoho_company_cfg(company_raw)
        if not cfg:
            app.logger.warning(f"No Zoho config for {company_raw}")
            return None
        
        # Search expenses with reference number filter
        url = f"{ZB_DOMAIN}/books/v3/expenses?organization_id={cfg['org_id']}&reference_number={reference_number}"
        app.logger.info(f"Searching Zoho for expense with reference: {reference_number}")
        
        resp = requests.get(url, headers=zoho_headers(company_raw), timeout=15)
        
        if resp.status_code == 200:
            data = resp.json()
            expenses = data.get('expenses', [])
            
            if expenses:
                # Return the first matching expense ID
                expense_id = expenses[0].get('expense_id')
                app.logger.info(f"Found existing expense: {expense_id} with reference: {reference_number}")
                return expense_id
            else:
                app.logger.debug(f"No existing expense found with reference: {reference_number}")
                return None
        else:
            app.logger.warning(f"Zoho search failed: {resp.status_code} - {resp.text[:200]}")
            return None
            
    except Exception as e:
        app.logger.error(f"Error searching Zoho for expense: {e}")
        return None

def build_admin_summary_text_from_csv(file_path: str, start_str: str, end_str: str) -> str:
    """Create a compact text version of the admin summary (top table) for Notes.
    Uses the uploaded CSV to recompute the same summary to avoid brittle Excel parsing.
    """
    try:
        if not file_path or not os.path.exists(file_path):
            return ''
        df = pd.read_csv(file_path)
        # Ensure required columns
        req_cols = ['Person ID', 'First Name', 'Last Name', 'Date']
        if not all(col in df.columns for col in req_cols):
            return ''
        # Compute like report
        pay_rates = load_pay_rates()
        if 'Daily Hours' not in df.columns:
            # Derive from Total Work Time when present
            if 'Total Work Time(h)' in df.columns:
                df['Daily Hours'] = df['Total Work Time(h)'].apply(parse_work_hours)
            else:
                # Fallback from Clock In/Out
                df['Daily Hours'] = df.apply(compute_daily_hours, axis=1)
        df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)
        df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)
        weekly_totals = df.groupby('Person ID').agg(
            Total_Hours=('Daily Hours', 'sum'),
            Weekly_Total=('Daily Pay', 'sum'),
            First_Name=('First Name', 'first'),
            Last_Name=('Last Name', 'first')
        ).reset_index()
        weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
        weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
        weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)
        # Grand totals
        grand_hours = float(weekly_totals['Total_Hours'].sum().round(2)) if len(weekly_totals) else 0.0
        grand_pay = float(weekly_totals['Weekly_Total'].sum().round(2)) if len(weekly_totals) else 0.0
        grand_rounded = int(weekly_totals['Rounded_Weekly'].sum()) if len(weekly_totals) else 0
        # Build lines
        lines = [
            f"Payroll Summary — {start_str} to {end_str}",
            "Person ID | Employee | Hours | Pay | Rounded"
        ]
        # Sort by employee name for readability
        weekly_totals = weekly_totals.sort_values(['First_Name', 'Last_Name'])
        for _, row in weekly_totals.iterrows():
            full_name = f"{row['First_Name']} {row['Last_Name']}".strip()
            lines.append(
                f"{row['Person ID']} | {full_name} | {row['Total_Hours']:.2f} | ${row['Weekly_Total']:.2f} | ${int(row['Rounded_Weekly'])}"
            )
        lines.append(f"GRAND TOTAL | | {grand_hours:.2f} | ${grand_pay:.2f} | ${grand_rounded}")
        return "\n".join(lines)
    except Exception:
        return ''

def _compose_zoho_description(base_desc: str, auto_notes: str, extra_desc: str, max_len: int = 500) -> str:
    """Build description prioritizing user notes over auto summary within Zoho's 500-char limit.
    Order: base, user extra notes, then as much of auto summary as fits.
    """
    try:
        base = (base_desc or '').strip()
        extra = (extra_desc or '').strip()
        auto = (auto_notes or '').strip()

        # Start with base and extra notes
        header_parts = [p for p in [base, extra] if p]
        header = "\n\n".join(header_parts)
        if len(header) >= max_len:
            # Trim header but keep user note visible
            return (header[: max_len - 1] + '…') if max_len > 1 else header[:max_len]

        # Fit as much auto summary as possible in remaining space
        remaining = max_len - len(header)
        # account for separator if both header and auto present
        if header and auto:
            remaining -= 2  # for "\n\n"
        if remaining <= 0 or not auto:
            return header

        if len(auto) <= remaining:
            return header + ("\n\n" if header else "") + auto

        # Need to trim auto summary; add friendly suffix
        suffix = " … (see attachment)"
        room = max(0, remaining - len(suffix))
        trimmed_auto = (auto[:room].rstrip() + suffix) if room > 0 else suffix.strip()
        return header + ("\n\n" if header else "") + trimmed_auto
    except Exception:
        return (base_desc or '')[:max_len]

def zoho_create_expense(company_raw, *, date_str, amount, account_id=None, account_name=None,
                        description=None, reference_number=None, tax_id=None,
                        is_inclusive_tax=False, paid_through_account_id=None, paid_through_account_name=None):
    """Create an Expense in Zoho Books and return expense_id."""
    cfg = get_zoho_company_cfg(company_raw)
    if not cfg:
        raise ValueError('Missing Zoho configuration for company: ' + str(company_raw))

    # Resolve IDs if not provided but names are configured
    resolved_account_id = account_id or (cfg['expense_account_id'] if cfg['expense_account_id'] else None)
    if not resolved_account_id and (account_name or cfg['expense_account_name']):
        resolved_account_id = zoho_find_account_id_by_name(company_raw, account_name or cfg['expense_account_name'])

    payload = {
        'account_id': resolved_account_id,
        'date': date_str,
        'amount': amount,
        'is_inclusive_tax': bool(is_inclusive_tax),
        'reference_number': reference_number or '',
        'description': description or ''
    }

    # Enforce Zoho's 500-character limit on description defensively
    try:
        if isinstance(payload.get('description'), str) and len(payload['description']) > 490:
            suffix = ' … (see attachment)'
            cutoff = 500 - len(suffix)
            if cutoff > 0:
                payload['description'] = payload['description'][:cutoff].rstrip() + suffix
            else:
                payload['description'] = payload['description'][:500]
    except Exception:
        pass

    # Prefer resolving Paid Through to an account_id if possible
    resolved_paid_id = paid_through_account_id
    if not resolved_paid_id:
        name_to_use = paid_through_account_name or cfg.get('paid_through_name')
        if name_to_use:
            try:
                resolved_paid_id = zoho_find_bank_account_id_by_name(company_raw, name_to_use)
            except Exception:
                resolved_paid_id = None
    if resolved_paid_id:
        payload['paid_through_account_id'] = resolved_paid_id
    elif paid_through_account_name or cfg.get('paid_through_name'):
        # Fallback to name if ID lookup fails
        payload['paid_through_account_name'] = paid_through_account_name or cfg.get('paid_through_name')

    if tax_id:
        payload['tax_id'] = tax_id

    url = f"{ZB_DOMAIN}/books/v3/expenses?organization_id={cfg['org_id']}"
    resp = requests.post(url, headers=zoho_headers(company_raw), data=json.dumps(payload), timeout=30)
    if resp.status_code not in (200, 201):
        # Retry once with a trimmed description if server complains about Description length
        try:
            txt = resp.text or ''
            if 'Description' in txt or '500 characters' in txt or 'code":15' in txt:
                desc = payload.get('description') or ''
                suffix = ' …'
                payload['description'] = (desc[: max(0, 500 - len(suffix))] + suffix)[:500]
                resp = requests.post(url, headers=zoho_headers(company_raw), data=json.dumps(payload), timeout=30)
        except Exception:
            pass
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Failed to create expense: {resp.status_code} {resp.text}")
    data = resp.json()
    expense_id = str((data.get('expense') or {}).get('expense_id') or (data.get('expenses') or {}).get('expense_id') or '')
    if not expense_id:
        # Some responses wrap under 'expense'
        expense_id = str((data.get('expense') or {}).get('expense_id') or '')
    if not expense_id:
        raise RuntimeError('Expense created but no expense_id returned: ' + str(data))
    return expense_id

def zoho_attach_receipt(company_raw, expense_id, file_path):
    """Attach a file to an expense as receipt. Will attempt direct upload; if unsupported type, raise error."""
    cfg = get_zoho_company_cfg(company_raw)
    url = f"{ZB_DOMAIN}/books/v3/expenses/{expense_id}/receipt?organization_id={cfg['org_id']}"
    headers = {
        'Authorization': zoho_headers(company_raw)['Authorization']
    }
    filename = os.path.basename(file_path)
    # Guess mime
    mime = 'application/octet-stream'
    if filename.lower().endswith('.pdf'):
        mime = 'application/pdf'
    elif filename.lower().endswith('.xlsx'):
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif filename.lower().endswith('.xls'):
        mime = 'application/vnd.ms-excel'

    with open(file_path, 'rb') as f:
        files = {
            'receipt': (filename, f, mime)
        }
        resp = requests.post(url, headers=headers, files=files, timeout=60)
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Failed to attach receipt: {resp.status_code} {resp.text}")
    return True

# Make sure required directories exist
Path(UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
Path(REPORT_FOLDER).mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT & AUTHENTICATION
# ═══════════════════════════════════════════════════════════════════════════════
# User authentication, session management, and access control

# Default admin user (will be created if no users exist)
DEFAULT_USERNAME = 'admin'
DEFAULT_PASSWORD = 'password'
def load_users():
    """Load users from JSON file with error handling"""
    try:
        with open(USERS_FILE, 'r') as f:
            users = json.load(f)
            app.logger.info(f"Successfully loaded {len(users)} users")
            return users
    except FileNotFoundError:
        # Create default user if no users file exists
        app.logger.warning(f"Users file not found. Creating default admin user.")
        users = {DEFAULT_USERNAME: DEFAULT_PASSWORD}
        save_users(users)
        return users
    except json.JSONDecodeError as e:
        app.logger.error(f"Invalid JSON in users file: {e}")
        # Create fresh users file
        users = {DEFAULT_USERNAME: DEFAULT_PASSWORD}
        save_users(users)
        return users
    except Exception as e:
        app.logger.error(f"Unexpected error loading users: {e}")
        return {DEFAULT_USERNAME: DEFAULT_PASSWORD}

def save_users(users):
    """Save users to JSON file with error handling"""
    try:
        # Create backup before saving
        if os.path.exists(USERS_FILE):
            backup_file = f"{USERS_FILE}.backup"
            import shutil
            shutil.copy2(USERS_FILE, backup_file)
        
        with open(USERS_FILE, 'w') as f:
            json.dump(users, f, indent=2)
        app.logger.info(f"Successfully saved {len(users)} users")
    except IOError as e:
        app.logger.error(f"Failed to save users: {e}")
        raise RuntimeError(f"Could not save users: {str(e)}")
    except Exception as e:
        app.logger.error(f"Unexpected error saving users: {e}")
        raise

def hash_password(password):
    """Hash a password using werkzeug's secure password hashing"""
    return generate_password_hash(password, method='pbkdf2:sha256')

def verify_password(stored_password, provided_password):
    """
    Verify a password against the stored hash.
    Also supports legacy plaintext passwords for backwards compatibility.
    """
    # Check if password is already hashed (starts with method identifier)
    if stored_password.startswith('pbkdf2:sha256:') or stored_password.startswith('scrypt:'):
        return check_password_hash(stored_password, provided_password)
    else:
        # Legacy plaintext password - compare directly but log warning
        app.logger.warning(f"Plaintext password detected - please update to hashed password")
        return stored_password == provided_password

def migrate_plaintext_passwords():
    """
    Migrate all plaintext passwords to hashed passwords.
    This function should be called on app startup.
    """
    try:
        users = load_users()
        migrated = False
        
        for username, password in users.items():
            # Check if password is already hashed
            if not (password.startswith('pbkdf2:sha256:') or password.startswith('scrypt:')):
                # Migrate to hashed password
                users[username] = hash_password(password)
                migrated = True
                app.logger.info(f"Migrated password for user: {username}")
        
        if migrated:
            save_users(users)
            app.logger.info("Password migration completed successfully")
        else:
            app.logger.debug("No plaintext passwords found - all passwords already hashed")
            
    except Exception as e:
        app.logger.error(f"Error during password migration: {e}")

def validate_username(username):
    """
    Validate username format
    Returns (is_valid, error_message)
    """
    if not username:
        return False, "Username is required"
    
    if len(username) < 3:
        return False, "Username must be at least 3 characters"
    
    if len(username) > 50:
        return False, "Username must be less than 50 characters"
    
    # Allow alphanumeric, underscore, and hyphen only
    if not re.match(r'^[a-zA-Z0-9_-]+$', username):
        return False, "Username can only contain letters, numbers, underscores, and hyphens"
    
    return True, None

def validate_password(password):
    """
    Validate password strength
    Returns (is_valid, error_message)
    """
    if not password:
        return False, "Password is required"
    
    if len(password) < 8:
        return False, "Password must be at least 8 characters"
    
    if len(password) > 100:
        return False, "Password must be less than 100 characters"
    
    # Check for at least one letter and one number
    if not re.search(r'[a-zA-Z]', password):
        return False, "Password must contain at least one letter"
    
    if not re.search(r'[0-9]', password):
        return False, "Password must contain at least one number"
    
    return True, None

def validate_pay_rate(rate_str):
    """
    Validate pay rate value
    Returns (is_valid, error_message, rate_float)
    """
    try:
        rate = float(rate_str)
        
        if rate < 0:
            return False, "Pay rate cannot be negative", None
        
        if rate > 10000:
            return False, "Pay rate seems unreasonably high (max: $10,000/hour)", None
        
        if rate == 0:
            app.logger.warning("Pay rate of $0.00 set for employee")
        
        return True, None, rate
    except (ValueError, TypeError):
        return False, "Pay rate must be a valid number", None

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# Pay rate management functions
def load_pay_rates():
    """Load pay rates from JSON file with caching to reduce file I/O"""
    global pay_rates_cache, pay_rates_cache_time
    
    # Check if cache is valid
    current_time = time.time()
    if (pay_rates_cache is not None and 
        pay_rates_cache_time is not None and 
        current_time - pay_rates_cache_time < PAY_RATES_CACHE_TTL):
        app.logger.debug("Using cached pay rates")
        return pay_rates_cache
    
    # Load from file
    try:
        with open(CONFIG_FILE, 'r') as f:
            rates = json.load(f)
            app.logger.info(f"Successfully loaded {len(rates)} pay rates from disk")
            # Update cache
            pay_rates_cache = rates
            pay_rates_cache_time = current_time
            return rates
    except FileNotFoundError:
        app.logger.warning(f"Pay rates file not found: {CONFIG_FILE}. Creating new file.")
        return {}
    except json.JSONDecodeError as e:
        app.logger.error(f"Invalid JSON in pay rates file: {e}")
        return {}
    except Exception as e:
        app.logger.error(f"Unexpected error loading pay rates: {e}")
        return {}

def save_pay_rates(rates):
    """Save pay rates to JSON file with error handling and cache invalidation"""
    global pay_rates_cache, pay_rates_cache_time
    
    try:
        # Create backup before saving
        if os.path.exists(CONFIG_FILE):
            backup_file = f"{CONFIG_FILE}.backup"
            import shutil
            shutil.copy2(CONFIG_FILE, backup_file)
        
        with open(CONFIG_FILE, 'w') as f:
            json.dump(rates, f, indent=2)
        app.logger.info(f"Successfully saved {len(rates)} pay rates")
        
        # Invalidate cache after saving
        pay_rates_cache = rates
        pay_rates_cache_time = time.time()
    except IOError as e:
        app.logger.error(f"Failed to save pay rates: {e}")
        raise RuntimeError(f"Could not save pay rates: {str(e)}")
    except Exception as e:
        app.logger.error(f"Unexpected error saving pay rates: {e}")
        raise

def get_employee_names():
    """Extract employee names from uploaded CSV files (front-end display only)"""
    employee_names = {}
    try:
        # Get all CSV files from uploads folder
        csv_files = []
        if os.path.exists(UPLOAD_FOLDER):
            for filename in os.listdir(UPLOAD_FOLDER):
                if filename.endswith('.csv'):
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    csv_files.append((filepath, os.path.getmtime(filepath)))
        
        # Sort by modification time (newest first) and limit to recent files
        csv_files.sort(key=lambda x: x[1], reverse=True)
        recent_files = [f for f, _ in csv_files[:10]]  # Check last 10 files
        
        # Extract employee names from CSV files
        for filepath in recent_files:
            try:
                df = pd.read_csv(filepath, dtype=str)
                if all(col in df.columns for col in ['Person ID', 'First Name', 'Last Name']):
                    for _, row in df.iterrows():
                        emp_id = str(row['Person ID']).strip()
                        first_name = str(row['First Name']).strip()
                        last_name = str(row['Last Name']).strip()
                        if emp_id and first_name and last_name and emp_id not in employee_names:
                            employee_names[emp_id] = f"{first_name} {last_name}"
            except Exception:
                continue  # Skip files that can't be read
    except Exception:
        pass
    
    return employee_names


# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY INITIALIZATION
# ═══════════════════════════════════════════════════════════════════════════════
# Migrate plaintext passwords to hashed passwords on startup
try:
    migrate_plaintext_passwords()
except Exception as e:
    app.logger.error(f"Failed to migrate passwords: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# AUTHENTICATION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════
# Login, logout, and password management endpoints

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login"""
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        users = load_users()

        if username in users and verify_password(users[username], password):
            session['logged_in'] = True
            session['username'] = username
            next_page = request.args.get('next')
            if next_page and next_page.startswith('/'):
                return redirect(next_page)
            return redirect(url_for('index'))
        else:
            error = 'Invalid credentials. Please try again.'

    # Enterprise Login Page
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - Payroll Management System</title>
        <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
        <link rel="stylesheet" href="/static/design-system.css">
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
        <style>
            body {{
                display: flex;
                align-items: center;
                justify-content: center;
                min-height: 100vh;
                background: linear-gradient(135deg, #f8fafc 0%, #e0e7ff 100%);
                padding: var(--spacing-3);
            }}
            .login-wrapper {{
                width: 100%;
                max-width: 420px;
            }}
            .login-card {{
                background: white;
                border-radius: var(--radius-xl);
                box-shadow: var(--shadow-xl);
                padding: var(--spacing-4);
                border: 1px solid var(--color-gray-200);
            }}
            .login-header {{
                text-align: center;
                margin-bottom: var(--spacing-4);
            }}
            .login-logo {{
                width: 64px;
                height: 64px;
                margin: 0 auto var(--spacing-4);
                background: linear-gradient(135deg, var(--color-primary), var(--color-primary-light));
                border-radius: var(--radius-xl);
                display: flex;
                align-items: center;
                justify-content: center;
                box-shadow: var(--shadow-md);
            }}
            .login-logo svg {{
                width: 40px;
                height: 40px;
                color: white;
            }}
            .login-title {{
                font-size: var(--font-size-2xl);
                font-weight: var(--font-weight-bold);
                color: var(--color-gray-900);
                margin-bottom: var(--spacing-2);
            }}
            .login-subtitle {{
                font-size: var(--font-size-sm);
                color: var(--color-gray-600);
                margin: 0;
            }}
            .login-form {{
                margin-top: var(--spacing-3);
            }}
            .form-footer {{
                margin-top: var(--spacing-3);
                text-align: center;
                font-size: var(--font-size-sm);
                color: var(--color-gray-600);
            }}
            .version-badge {{
                display: inline-block;
                margin-top: var(--spacing-4);
                padding: var(--spacing-1) var(--spacing-3);
                background: var(--color-gray-100);
                color: var(--color-gray-700);
                border-radius: var(--radius-full);
                font-size: var(--font-size-xs);
                font-weight: var(--font-weight-medium);
            }}
        </style>
    </head>
    <body>
        <div class="login-wrapper">
            <div class="login-card">
                <div class="login-header">
                    <div class="login-logo">
                        <svg fill="currentColor" viewBox="0 0 20 20">
                            <path d="M8.433 7.418c.155-.103.346-.196.567-.267v1.698a2.305 2.305 0 01-.567-.267C8.07 8.34 8 8.114 8 8c0-.114.07-.34.433-.582zM11 12.849v-1.698c.22.071.412.164.567.267.364.243.433.468.433.582 0 .114-.07.34-.433.582a2.305 2.305 0 01-.567.267z"/>
                            <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm1-13a1 1 0 10-2 0v.092a4.535 4.535 0 00-1.676.662C6.602 6.234 6 7.009 6 8c0 .99.602 1.765 1.324 2.246.48.32 1.054.545 1.676.662v1.941c-.391-.127-.68-.317-.843-.504a1 1 0 10-1.51 1.31c.562.649 1.413 1.076 2.353 1.253V15a1 1 0 102 0v-.092a4.535 4.535 0 001.676-.662C13.398 13.766 14 12.991 14 12c0-.99-.602-1.765-1.324-2.246A4.535 4.535 0 0011 9.092V7.151c.391.127.68.317.843.504a1 1 0 101.511-1.31c-.563-.649-1.413-1.076-2.354-1.253V5z" clip-rule="evenodd"/>
                        </svg>
                    </div>
                    <h1 class="login-title">Payroll Management</h1>
                    <p class="login-subtitle">Sign in to access your payroll system</p>
                </div>

                {{% if error %}}
                <div class="alert alert-danger" role="alert">
                    <svg style="width:20px;height:20px;flex-shrink:0" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zM8.707 7.293a1 1 0 00-1.414 1.414L8.586 10l-1.293 1.293a1 1 0 101.414 1.414L10 11.414l1.293 1.293a1 1 0 001.414-1.414L11.414 10l1.293-1.293a1 1 0 00-1.414-1.414L10 8.586 8.707 7.293z" clip-rule="evenodd"/>
                    </svg>
                    <span>{{{{ error }}}}</span>
                </div>
                {{% endif %}}

                <form action="{{{{ url_for('login', next=request.args.get('next', '')) }}}}" method="post" class="login-form">
                    <div class="form-group">
                        <label for="username" class="form-label">Username</label>
                        <input 
                            type="text" 
                            id="username" 
                            name="username" 
                            class="form-input" 
                            placeholder="Enter your username"
                            required 
                            autofocus
                        >
                    </div>

                    <div class="form-group">
                        <label for="password" class="form-label">Password</label>
                        <input 
                            type="password" 
                            id="password" 
                            name="password" 
                            class="form-input" 
                            placeholder="Enter your password"
                            required
                        >
                    </div>

                    <button type="submit" class="btn btn-primary btn-lg" style="width:100%">
                        <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M5 9V7a5 5 0 0110 0v2a2 2 0 012 2v5a2 2 0 01-2 2H5a2 2 0 01-2-2v-5a2 2 0 012-2zm8-2v2H7V7a3 3 0 016 0z" clip-rule="evenodd"/>
                        </svg>
                        Sign In
                    </button>
                </form>

                <div class="form-footer">
                    <p>Secure payroll processing for your business</p>
                    <span class="version-badge">{get_version_display()}</span>
                </div>
            </div>

            <div class="text-center mt-6 text-sm text-gray-600">
                <p>© 2024-2025 Payroll Management System. All rights reserved.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return render_template_string(html, error=error, request=request)

@app.route('/logout')
def logout():
    """Log out user"""
    session.pop('logged_in', None)
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Change user password"""
    error = None
    success = None
    username = session.get('username', 'Unknown')
    menu_html = get_menu_html(username)

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        users = load_users()

        if not verify_password(users.get(username), current_password):
            error = 'Current password is incorrect'
        elif new_password != confirm_password:
            error = 'New passwords do not match'
        else:
            # Validate new password strength
            valid, validation_error = validate_password(new_password)
            if not valid:
                error = validation_error
            else:
                users[username] = hash_password(new_password)
                save_users(users)
                app.logger.info(f"Password changed for user: {username}")
                success = 'Password changed successfully'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Change Password - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .password-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="password-header">
        <div class="container container-narrow">
            <h1 style="color:white;margin-bottom:var(--spacing-2)">Change Password</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-lg);margin:0">Update your account password</p>
        </div>
    </div>
    
    <div class="container container-narrow">
        {('<div class="alert alert-danger">' + escape(error) + '</div>') if error else ''}
        {('<div class="alert alert-success">' + escape(success) + '</div>') if success else ''}
        
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">
                    <svg style="width:24px;height:24px;display:inline;margin-right:8px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M18 8a6 6 0 01-7.743 5.743L10 14l-1 1-1 1H6v2H2v-4l4.257-4.257A6 6 0 1118 8zm-6-4a1 1 0 100 2 2 2 0 012 2 1 1 0 102 0 4 4 0 00-4-4z" clip-rule="evenodd"/>
                    </svg>
                    Update Password
                </h2>
            </div>
            
            <form action="/change_password" method="post">
                <div class="form-group">
                    <label for="current_password" class="form-label form-label-required">Current Password</label>
                    <input type="password" id="current_password" name="current_password" class="form-input" placeholder="Enter your current password" required autofocus>
                </div>

                <div class="form-group">
                    <label for="new_password" class="form-label form-label-required">New Password</label>
                    <input type="password" id="new_password" name="new_password" class="form-input" placeholder="Enter your new password" required>
                    <span class="form-help">Must be at least 8 characters and include both letters and numbers</span>
                </div>

                <div class="form-group">
                    <label for="confirm_password" class="form-label form-label-required">Confirm New Password</label>
                    <input type="password" id="confirm_password" name="confirm_password" class="form-input" placeholder="Confirm your new password" required>
                </div>

                <div style="margin-top:var(--spacing-3);display:flex;gap:var(--spacing-4);justify-content:flex-end">
                    <a href="/" class="btn btn-secondary">Cancel</a>
                    <button type="submit" class="btn btn-primary">
                        <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd"/>
                        </svg>
                        Change Password
                    </button>
                </div>
            </form>
        </div>
    </div>
</body>
</html>
    """
    return html


# ═══════════════════════════════════════════════════════════════════════════════
# PAYROLL CALCULATION FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
# Time parsing, hours calculation, and pay rate management

def process_csv_data(file_path):
    """Process CSV timesheet data"""
    df = pd.read_csv(file_path, parse_dates=['Date'])

    # Clean data
    df = df.dropna(subset=['Person ID', 'First Name', 'Last Name'])

    # Calculate daily hours - simplified approach
    def parse_work_hours(time_str):
        try:
            if pd.isna(time_str) or str(time_str).strip() == '':
                return 0.0
            parts = list(map(float, str(time_str).split(':')))
            return round(parts[0] + parts[1]/60 + parts[2]/3600, 2)
        except:
            return 0.0

    df['Daily Hours'] = df['Total Work Time(h)'].apply(parse_work_hours)

    # Use fixed pay rates for this simplified version
    df['Hourly Rate'] = 15.0  # Fixed hourly rate

    # Calculate daily pay
    df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)

    # Calculate weekly totals
    weekly_totals = df.groupby('Person ID').agg(
        Total_Hours=('Daily Hours', 'sum'),
        Weekly_Total=('Daily Pay', 'sum'),
        First_Name=('First Name', 'first'),
        Last_Name=('Last Name', 'first')
    ).reset_index()

    # Apply rounding
    weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
    weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
    weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)

    # Merge weekly totals back to each employee's daily records
    df = pd.merge(
        df,
        weekly_totals[['Person ID', 'Total_Hours', 'Weekly_Total', 'Rounded_Weekly']],
        on='Person ID'
    )

    return df[df['Daily Hours'] > 0]

def create_report(df, week_str):
    """Create a simple Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    # Simple header
    ws['A1'] = f"Payroll Report - {week_str}"
    ws['A1'].font = Font(bold=True, size=14)

    # Headers
    headers = ["ID", "Employee", "Total Hours", "Pay", "Rounded Pay"]
    ws.append(headers)

    # Data rows
    seen_employees = set()
    for emp_id, group in df.groupby('Person ID'):
        if emp_id in seen_employees:
            continue

        seen_employees.add(emp_id)
        employee = group.iloc[0]

        ws.append([
            emp_id,
            f"{employee['First Name']} {employee['Last Name']}",
            employee['Total_Hours'],
            employee['Weekly_Total'],
            employee['Rounded_Weekly']
        ])

    # Save the report
    report_path = Path(REPORT_FOLDER) / f"Payroll_Report_{week_str}.xlsx"
    wb.save(report_path)

    return report_path


# ═══════════════════════════════════════════════════════════════════════════════
# UI TEMPLATE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
# HTML generation helpers for consistent UI across pages

def get_base_html_head(title="Payroll Management"):
    """Generate consistent HTML head with Bootstrap 5 and custom styles"""
    return f'''
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{title}</title>
        
        <!-- Bootstrap 5.3 CSS -->
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
        
        <!-- Inter Font -->
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
        
        <!-- Bootstrap Icons -->
        <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css">
        
        <style>
            :root {{
                --bs-body-font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
                --bs-body-font-size: 0.95rem;
                --bs-body-line-height: 1.6;
            }}
            
            body {{
                font-family: var(--bs-body-font-family);
                background-color: #f8f9fa;
                min-height: 100vh;
            }}
            
            .navbar-brand {{
                font-weight: 700;
                font-size: 1.25rem;
            }}
            
            .btn {{
                font-weight: 600;
                padding: 0.5rem 1.25rem;
                border-radius: 0.5rem;
                transition: all 0.2s ease;
            }}
            
            .btn:hover {{
                transform: translateY(-1px);
                box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            }}
            
            .card {{
                border: none;
                border-radius: 1rem;
                box-shadow: 0 2px 8px rgba(0,0,0,0.08);
                margin-bottom: 1.5rem;
            }}
            
            .card-header {{
                background: linear-gradient(135deg, #0d6efd 0%, #0a58ca 100%);
                color: white;
                font-weight: 700;
                border-radius: 1rem 1rem 0 0 !important;
                padding: 1rem 1.5rem;
            }}
            
            .form-label {{
                font-weight: 600;
                color: #495057;
                margin-bottom: 0.5rem;
            }}
            
            .form-control, .form-select {{
                border-radius: 0.5rem;
                border: 1px solid #dee2e6;
                padding: 0.625rem 0.875rem;
            }}
            
            .form-control:focus, .form-select:focus {{
                border-color: #0d6efd;
                box-shadow: 0 0 0 0.2rem rgba(13,110,253,0.25);
            }}
            
            .table {{
                font-size: 0.9rem;
            }}
            
            .badge {{
                font-weight: 600;
                padding: 0.375rem 0.75rem;
                border-radius: 0.375rem;
            }}
            
            .alert {{
                border: none;
                border-radius: 0.75rem;
                font-weight: 500;
            }}
            
            footer {{
                background: white;
                border-top: 1px solid #dee2e6;
                padding: 1.5rem 0;
                margin-top: 3rem;
            }}
        </style>
    </head>
    '''

def get_menu_html(username):
    """Generate enterprise navigation bar with design system"""
    is_admin = username == 'admin'
    admin_link = '''
        <a href="/manage_users" class="nav-link">
            <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                <path d="M9 6a3 3 0 11-6 0 3 3 0 016 0zM17 6a3 3 0 11-6 0 3 3 0 016 0zM12.93 17c.046-.327.07-.66.07-1a6.97 6.97 0 00-1.5-4.33A5 5 0 0119 16v1h-6.07zM6 11a5 5 0 015 5v1H1v-1a5 5 0 015-5z"/>
            </svg>
            <span>Manage Users</span>
        </a>
    ''' if is_admin else ''
    
    return f'''
    <nav class="navbar">
        <div class="navbar-container">
            <div class="navbar-brand">
                <img src="/static/favicon.svg" alt="Logo" class="navbar-logo">
                <span class="navbar-title">Payroll Management</span>
                <span class="navbar-version">{get_version_display()}</span>
            </div>
            
            <button class="navbar-toggle" onclick="toggleMobileMenu()" aria-label="Toggle navigation">
                <svg width="24" height="24" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 6h16M4 12h16M4 18h16"/>
                </svg>
            </button>
            
            <div class="navbar-menu" id="navbarMenu">
                <div class="navbar-left">
                    <a href="/" class="nav-link">
                        <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                            <path d="M10.707 2.293a1 1 0 00-1.414 0l-7 7a1 1 0 001.414 1.414L4 10.414V17a1 1 0 001 1h2a1 1 0 001-1v-2a1 1 0 011-1h2a1 1 0 011 1v2a1 1 0 001 1h2a1 1 0 001-1v-6.586l.293.293a1 1 0 001.414-1.414l-7-7z"/>
                        </svg>
                        <span>Home</span>
                    </a>
                    <a href="/fetch_timecard" class="nav-link">
                        <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd"/>
                        </svg>
                        <span>Fetch Timecard</span>
                    </a>
                    <a href="/manage_rates" class="nav-link">
                        <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                            <path d="M8.433 7.418c.155-.103.346-.196.567-.267v1.698a2.305 2.305 0 01-.567-.267C8.07 8.34 8 8.114 8 8c0-.114.07-.34.433-.582zM11 12.849v-1.698c.22.071.412.164.567.267.364.243.433.468.433.582 0 .114-.07.34-.433.582a2.305 2.305 0 01-.567.267z"/>
                            <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm1-13a1 1 0 10-2 0v.092a4.535 4.535 0 00-1.676.662C6.602 6.234 6 7.009 6 8c0 .99.602 1.765 1.324 2.246.48.32 1.054.545 1.676.662v1.941c-.391-.127-.68-.317-.843-.504a1 1 0 10-1.51 1.31c.562.649 1.413 1.076 2.353 1.253V15a1 1 0 102 0v-.092a4.535 4.535 0 001.676-.662C13.398 13.766 14 12.991 14 12c0-.99-.602-1.765-1.324-2.246A4.535 4.535 0 0011 9.092V7.151c.391.127.68.317.843.504a1 1 0 101.511-1.31c-.563-.649-1.413-1.076-2.354-1.253V5z" clip-rule="evenodd"/>
                        </svg>
                        <span>Pay Rates</span>
                    </a>
                    <a href="/reports" class="nav-link">
                        <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M4 4a2 2 0 012-2h4.586A2 2 0 0112 2.586L15.414 6A2 2 0 0116 7.414V16a2 2 0 01-2 2H6a2 2 0 01-2-2V4zm2 6a1 1 0 011-1h6a1 1 0 110 2H7a1 1 0 01-1-1zm1 3a1 1 0 100 2h6a1 1 0 100-2H7z" clip-rule="evenodd"/>
                        </svg>
                        <span>Reports</span>
                    </a>
                    {admin_link}
                </div>
                
                <div class="navbar-right">
                    <div class="user-menu">
                        <button class="user-menu-button" onclick="toggleUserMenu()" aria-label="User menu">
                            <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                                <path fill-rule="evenodd" d="M10 9a3 3 0 100-6 3 3 0 000 6zm-7 9a7 7 0 1114 0H3z" clip-rule="evenodd"/>
                            </svg>
                            <span>{escape(username)}</span>
                            <svg width="16" height="16" fill="currentColor" viewBox="0 0 20 20">
                                <path fill-rule="evenodd" d="M5.293 7.293a1 1 0 011.414 0L10 10.586l3.293-3.293a1 1 0 111.414 1.414l-4 4a1 1 0 01-1.414 0l-4-4a1 1 0 010-1.414z" clip-rule="evenodd"/>
                            </svg>
                        </button>
                        <div class="user-menu-dropdown" id="userMenuDropdown">
                            <a href="/change_password" class="user-menu-item">
                                <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M18 8a6 6 0 01-7.743 5.743L10 14l-1 1-1 1H6v2H2v-4l4.257-4.257A6 6 0 1118 8zm-6-4a1 1 0 100 2 2 2 0 012 2 1 1 0 102 0 4 4 0 00-4-4z" clip-rule="evenodd"/>
                                </svg>
                                Change Password
                            </a>
                            <a href="/logout" class="user-menu-item user-menu-item-danger">
                                <svg width="20" height="20" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M3 3a1 1 0 00-1 1v12a1 1 0 102 0V4a1 1 0 00-1-1zm10.293 9.293a1 1 0 001.414 1.414l3-3a1 1 0 000-1.414l-3-3a1 1 0 10-1.414 1.414L14.586 9H7a1 1 0 100 2h7.586l-1.293 1.293z" clip-rule="evenodd"/>
                                </svg>
                                Logout
                            </a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </nav>
    
    <style>
        .navbar {{
            background: white;
            border-bottom: 1px solid var(--color-gray-200);
            box-shadow: var(--shadow-sm);
            position: sticky;
            top: 0;
            z-index: var(--z-sticky);
        }}
        .navbar-container {{
            max-width: 1600px;
            margin: 0 auto;
            padding: 0 var(--spacing-3);
            display: flex;
            align-items: center;
            justify-content: space-between;
            height: 64px;
        }}
        .navbar-brand {{
            display: flex;
            align-items: center;
            gap: var(--spacing-3);
        }}
        .navbar-logo {{
            width: 32px;
            height: 32px;
        }}
        .navbar-title {{
            font-size: var(--font-size-lg);
            font-weight: var(--font-weight-semibold);
            color: var(--color-gray-900);
        }}
        .navbar-version {{
            font-size: var(--font-size-xs);
            color: var(--color-gray-500);
            padding: var(--spacing-1) var(--spacing-2);
            background: var(--color-gray-100);
            border-radius: var(--radius-full);
        }}
        .navbar-toggle {{
            display: none;
            background: none;
            border: none;
            color: var(--color-gray-700);
            cursor: pointer;
            padding: var(--spacing-2);
        }}
        .navbar-menu {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            flex: 1;
            margin-left: var(--spacing-4);
        }}
        .navbar-left {{
            display: flex;
            gap: var(--spacing-2);
        }}
        .navbar-right {{
            display: flex;
            align-items: center;
            gap: var(--spacing-4);
        }}
        .nav-link {{
            display: flex;
            align-items: center;
            gap: var(--spacing-2);
            padding: var(--spacing-2) var(--spacing-4);
            font-size: var(--font-size-sm);
            font-weight: var(--font-weight-medium);
            color: var(--color-gray-700);
            text-decoration: none;
            border-radius: var(--radius-md);
            transition: all var(--transition-fast);
        }}
        .nav-link:hover {{
            background: var(--color-gray-100);
            color: var(--color-primary);
        }}
        .user-menu {{
            position: relative;
        }}
        .user-menu-button {{
            display: flex;
            align-items: center;
            gap: var(--spacing-2);
            padding: var(--spacing-2) var(--spacing-3);
            background: var(--color-gray-50);
            border: 1px solid var(--color-gray-200);
            border-radius: var(--radius-md);
            font-size: var(--font-size-sm);
            font-weight: var(--font-weight-medium);
            color: var(--color-gray-700);
            cursor: pointer;
            transition: all var(--transition-fast);
        }}
        .user-menu-button:hover {{
            background: var(--color-gray-100);
            border-color: var(--color-gray-300);
        }}
        .user-menu-dropdown {{
            display: none;
            position: absolute;
            right: 0;
            top: calc(100% + var(--spacing-2));
            background: white;
            border: 1px solid var(--color-gray-200);
            border-radius: var(--radius-md);
            box-shadow: var(--shadow-lg);
            min-width: 200px;
            z-index: var(--z-dropdown);
        }}
        .user-menu-dropdown.show {{
            display: block;
        }}
        .user-menu-item {{
            display: flex;
            align-items: center;
            gap: var(--spacing-3);
            padding: var(--spacing-3) var(--spacing-4);
            font-size: var(--font-size-sm);
            color: var(--color-gray-700);
            text-decoration: none;
            transition: background var(--transition-fast);
        }}
        .user-menu-item:hover {{
            background: var(--color-gray-50);
        }}
        .user-menu-item-danger {{
            color: var(--color-danger);
        }}
        .user-menu-item-danger:hover {{
            background: var(--color-danger-light);
        }}
        
        @media (max-width: 768px) {{
            .navbar-toggle {{
                display: block;
            }}
            .navbar-title {{
                font-size: var(--font-size-base);
            }}
            .navbar-version {{
                display: none;
            }}
            .navbar-menu {{
                display: none;
                position: absolute;
                top: 64px;
                left: 0;
                right: 0;
                background: white;
                border-bottom: 1px solid var(--color-gray-200);
                box-shadow: var(--shadow-lg);
                flex-direction: column;
                padding: var(--spacing-4);
                margin-left: 0;
            }}
            .navbar-menu.show {{
                display: flex;
            }}
            .navbar-left,
            .navbar-right {{
                flex-direction: column;
                width: 100%;
                gap: var(--spacing-2);
            }}
            .nav-link {{
                width: 100%;
                justify-content: flex-start;
            }}
            .user-menu {{
                width: 100%;
            }}
            .user-menu-button {{
                width: 100%;
                justify-content: space-between;
            }}
        }}
    </style>
    
    <script>
        function toggleMobileMenu() {{
            document.getElementById('navbarMenu').classList.toggle('show');
        }}
        function toggleUserMenu() {{
            document.getElementById('userMenuDropdown').classList.toggle('show');
        }}
        // Close dropdowns when clicking outside
        document.addEventListener('click', function(event) {{
            if (!event.target.closest('.user-menu')) {{
                document.getElementById('userMenuDropdown').classList.remove('show');
            }}
        }});
    </script>
    '''

def get_enterprise_sidebar(username, active_page="home"):
    """Generate enterprise sidebar navigation HTML"""
    is_admin = username == 'admin'
    
    admin_menu = '''<a href="/manage_users" class="flex items-center space-x-3 px-3 py-2.5 text-sm font-medium rounded-lg text-secondary hover:bg-gray-100 hover:text-textDark transition-colors">
                        <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 4.354a4 4 0 110 5.292M15 21H3v-1a6 6 0 0112 0v1zm0 0h6v-1a6 6 0 00-9-5.197M13 7a4 4 0 11-8 0 4 4 0 018 0z" />
                        </svg>
                        <span>Manage Users</span>
                    </a>''' if is_admin else ''
    
    def nav_class(page):
        if page == active_page:
            return "flex items-center space-x-3 px-3 py-2.5 text-sm font-medium rounded-lg bg-primary/10 text-primary"
        return "flex items-center space-x-3 px-3 py-2.5 text-sm font-medium rounded-lg text-secondary hover:bg-gray-100 hover:text-textDark transition-colors"
    
    return f'''
    <aside class="w-64 bg-white border-r border-gray-200 flex-shrink-0 hidden lg:block">
        <div class="h-full flex flex-col">
            <!-- Logo -->
            <div class="px-6 py-6 border-b border-gray-200">
                <div class="flex items-center space-x-3">
                    <svg class="w-8 h-8 text-primary" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 7h6m0 10v-3m-3 3h.01M9 17h.01M9 14h.01M12 14h.01M15 11h.01M12 11h.01M9 11h.01M7 21h10a2 2 0 002-2V5a2 2 0 00-2-2H7a2 2 0 00-2 2v14a2 2 0 002 2z" />
                    </svg>
                    <div>
                        <h1 class="text-lg font-bold text-textDark">Payroll</h1>
                        <p class="text-xs text-secondary">Management</p>
                    </div>
                </div>
            </div>
            
            <!-- Navigation -->
            <nav class="flex-1 px-4 py-6 space-y-1">
                <a href="/" class="{nav_class('home')}">
                    <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6" />
                    </svg>
                    <span>Home</span>
                </a>
                
                <a href="/fetch_timecard" class="{nav_class('fetch')}">
                    <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M9 19l3 3m0 0l3-3m-3 3V10" />
                    </svg>
                    <span>Fetch from NGTeco</span>
                </a>
                
                <a href="/manage_rates" class="{nav_class('rates')}">
                    <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 8c-1.657 0-3 .895-3 2s1.343 2 3 2 3 .895 3 2-1.343 2-3 2m0-8c1.11 0 2.08.402 2.599 1M12 8V7m0 1v8m0 0v1m0-1c-1.11 0-2.08-.402-2.599-1M21 12a9 9 0 11-18 0 9 9 0 0118 0z" />
                    </svg>
                    <span>Pay Rates</span>
                </a>
                
                <a href="/reports" class="{nav_class('reports')}">
                    <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z" />
                    </svg>
                    <span>Reports</span>
                </a>
                
                {admin_menu}
            </nav>
            
            <!-- Bottom Section -->
            <div class="px-4 py-4 border-t border-gray-200 space-y-1">
                <a href="/change_password" class="{nav_class('password')}">
                    <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15 7a2 2 0 012 2m4 0a6 6 0 01-7.743 5.743L11 17H9v2H7v2H4a1 1 0 01-1-1v-2.586a1 1 0 01.293-.707l5.964-5.964A6 6 0 1121 9z" />
                    </svg>
                    <span>Change Password</span>
                </a>
                
                <a href="/logout" class="flex items-center space-x-3 px-3 py-2.5 text-sm font-medium rounded-lg text-danger hover:bg-danger/10 transition-colors">
                    <svg class="w-5 h-5" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M17 16l4-4m0 0l-4-4m4 4H7m6 4v1a3 3 0 01-3 3H6a3 3 0 01-3-3V7a3 3 0 013-3h4a3 3 0 013 3v1" />
                    </svg>
                    <span>Logout</span>
                </a>
            </div>
        </div>
    </aside>
    '''

def get_footer_html():
    """Generate consistent footer"""
    return '''
    <footer class="mt-auto">
        <div class="container">
            <div class="row">
                <div class="col-md-6 text-center text-md-start">
                    <p class="mb-0 text-muted">
                        <small>&copy; 2024-2025 Payroll Management System | Secure Processing</small>
                    </p>
                </div>
                <div class="col-md-6 text-center text-md-end">
                    <p class="mb-0">
                        <small class="text-muted">Version <span class="badge bg-primary">{get_version()}</span></small>
                    </p>
                </div>
            </div>
        </div>
    </footer>
    
    <!-- Bootstrap 5.3 JS Bundle -->
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
    '''


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════
# Core application endpoints for timesheet processing and payroll generation

@app.route('/')
@login_required
def index():
    """Home page with payroll upload form"""
    username = session.get('username', 'Unknown')
    menu_html = get_menu_html(username)

    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Home - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .home-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
        .step-number {{
            width: 32px;
            height: 32px;
            background: var(--color-primary-pale);
            color: var(--color-primary);
            border-radius: var(--radius-full);
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: var(--font-weight-bold);
            flex-shrink: 0;
        }}
        .step-item {{
            display: flex;
            gap: var(--spacing-4);
            padding: var(--spacing-4);
            border-radius: var(--radius-md);
            transition: background var(--transition-fast);
        }}
        .step-item:hover {{
            background: var(--color-gray-50);
        }}
        .dropzone {{
            border: 2px dashed var(--color-gray-300);
            border-radius: var(--radius-lg);
            padding: var(--spacing-4);
            text-align: center;
            cursor: pointer;
            transition: all var(--transition-base);
            background: var(--color-gray-50);
            min-height: 180px;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }}
        .dropzone:hover {{
            border-color: var(--color-primary);
            background: var(--color-primary-pale);
        }}
        .dropzone.dragover {{
            border-color: var(--color-primary);
            background: var(--color-primary-pale);
            transform: scale(1.01);
        }}
        .upload-icon {{
            width: 48px;
            height: 48px;
            margin: 0 auto var(--spacing-2);
            color: var(--color-gray-400);
        }}
        .feature-icon {{
            width: 40px;
            height: 40px;
            background: var(--color-success-light);
            color: var(--color-success);
            border-radius: var(--radius-md);
            display: flex;
            align-items: center;
            justify-content: center;
            flex-shrink: 0;
        }}
        .home-layout {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: var(--spacing-4);
            align-items: start;
        }}
        @media (max-width: 1024px) {{
            .home-layout {{
                grid-template-columns: 1fr;
            }}
        }}
        .step-item {{
            padding: var(--spacing-2);
        }}
        .step-number {{
            width: 24px;
            height: 24px;
            font-size: var(--font-size-sm);
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="home-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-1);font-size:var(--font-size-2xl)">Process Payroll</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-sm);margin:0">Upload timesheets and generate professional payroll reports</p>
        </div>
    </div>
    
    <div class="container">
        <div class="home-layout">
            
            <!-- Upload Card -->
            <div class="card" style="margin-bottom:0">
                <div class="card-header" style="padding-bottom:var(--spacing-2);margin-bottom:var(--spacing-2)">
                    <h2 class="card-title" style="font-size:var(--font-size-base)">
                        <svg style="width:20px;height:20px;display:inline;margin-right:6px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd"/>
                        </svg>
                        Upload Timesheet CSV
                    </h2>
                </div>
                
                <form id="upload-form" action="/validate" method="post" enctype="multipart/form-data">
                    <div id="dropzone" class="dropzone">
                        <svg class="upload-icon" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12" />
                        </svg>
                        <h3 style="font-size:var(--font-size-base);font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:var(--spacing-1)">
                            Drag & drop CSV file here
                        </h3>
                        <p style="color:var(--color-gray-600);font-size:var(--font-size-sm);margin-bottom:var(--spacing-2)">
                            or <span style="color:var(--color-primary);font-weight:var(--font-weight-medium)">click to browse</span>
                        </p>
                        <div id="file-note" style="font-size:var(--font-size-xs);color:var(--color-gray-600);font-weight:var(--font-weight-medium)">
                            No file selected
                        </div>
                        <input id="file-input" type="file" name="file" accept=".csv" style="display:none" required>
                    </div>
                    
                    <div style="margin-top:var(--spacing-3);text-align:center">
                        <button type="submit" class="btn btn-primary">
                            <svg style="width:18px;height:18px" fill="currentColor" viewBox="0 0 20 20">
                                <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/>
                            </svg>
                            Process Timesheet
                        </button>
                    </div>
                </form>
            </div>
            
            <!-- Instructions Card -->
            <div class="card" style="margin-bottom:0">
                <div class="card-header" style="padding-bottom:var(--spacing-2);margin-bottom:var(--spacing-2)">
                    <h2 class="card-title" style="font-size:var(--font-size-base)">
                        <svg style="width:20px;height:20px;display:inline;margin-right:6px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7-4a1 1 0 11-2 0 1 1 0 012 0zM9 9a1 1 0 000 2v3a1 1 0 001 1h1a1 1 0 100-2v-3a1 1 0 00-1-1H9z" clip-rule="evenodd"/>
                        </svg>
                        How It Works
                    </h2>
                </div>
                
                <div style="display:flex;flex-direction:column;gap:var(--spacing-2)">
                    <div class="step-item">
                        <div class="step-number">1</div>
                        <div style="flex:1">
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:0;font-size:var(--font-size-sm)">Upload CSV</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Drag & drop or select CSV file</p>
                        </div>
                    </div>
                    
                    <div class="step-item">
                        <div class="step-number">2</div>
                        <div style="flex:1">
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:0;font-size:var(--font-size-sm)">Fix Missing Times</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Review and correct if needed</p>
                        </div>
                    </div>
                    
                    <div class="step-item">
                        <div class="step-number">3</div>
                        <div style="flex:1">
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:0;font-size:var(--font-size-sm)">Select Employees</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Choose employees to include</p>
                        </div>
                    </div>
                    
                    <div class="step-item">
                        <div class="step-number">4</div>
                        <div style="flex:1">
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:0;font-size:var(--font-size-sm)">Generate Reports</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Process and download Excel</p>
                        </div>
                    </div>
                    
                    <div class="step-item">
                        <div class="step-number">5</div>
                        <div style="flex:1">
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:0;font-size:var(--font-size-sm)">Push to Zoho</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Sync expense automatically</p>
                        </div>
                    </div>
                </div>
                
                <div class="alert alert-info" style="margin-top:var(--spacing-3);padding:var(--spacing-2)">
                    <svg style="width:16px;height:16px;flex-shrink:0" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7-4a1 1 0 11-2 0 1 1 0 012 0zM9 9a1 1 0 000 2v3a1 1 0 001 1h1a1 1 0 100-2v-3a1 1 0 00-1-1H9z" clip-rule="evenodd"/>
                    </svg>
                    <div style="font-size:var(--font-size-xs)">
                        <strong>CSV Format:</strong> Person ID, First Name, Last Name, Date, Clock In, Clock Out
                    </div>
                </div>
            </div>
            
        </div>
        
        <!-- Quick Links Row -->
        <div class="card" style="margin-top:var(--spacing-4)">
            <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:var(--spacing-3)">
                <a href="/manage_rates" style="text-decoration:none">
                    <div class="flex items-center gap-3" style="padding:var(--spacing-3);background:var(--color-gray-50);border-radius:var(--radius-md);transition:all var(--transition-fast)">
                        <div class="feature-icon">
                            <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                                <path d="M8.433 7.418c.155-.103.346-.196.567-.267v1.698a2.305 2.305 0 01-.567-.267C8.07 8.34 8 8.114 8 8c0-.114.07-.34.433-.582zM11 12.849v-1.698c.22.071.412.164.567.267.364.243.433.468.433.582 0 .114-.07.34-.433.582a2.305 2.305 0 01-.567.267z"/>
                                <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm1-13a1 1 0 10-2 0v.092a4.535 4.535 0 00-1.676.662C6.602 6.234 6 7.009 6 8c0 .99.602 1.765 1.324 2.246.48.32 1.054.545 1.676.662v1.941c-.391-.127-.68-.317-.843-.504a1 1 0 10-1.51 1.31c.562.649 1.413 1.076 2.353 1.253V15a1 1 0 102 0v-.092a4.535 4.535 0 001.676-.662C13.398 13.766 14 12.991 14 12c0-.99-.602-1.765-1.324-2.246A4.535 4.535 0 0011 9.092V7.151c.391.127.68.317.843.504a1 1 0 101.511-1.31c-.563-.649-1.413-1.076-2.354-1.253V5z" clip-rule="evenodd"/>
                            </svg>
                        </div>
                        <div>
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin:0;font-size:var(--font-size-sm)">Pay Rates</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Manage employees</p>
                        </div>
                    </div>
                </a>
                
                <a href="/reports" style="text-decoration:none">
                    <div class="flex items-center gap-3" style="padding:var(--spacing-3);background:var(--color-gray-50);border-radius:var(--radius-md);transition:all var(--transition-fast)">
                        <div class="feature-icon">
                            <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                                <path fill-rule="evenodd" d="M4 4a2 2 0 012-2h4.586A2 2 0 0112 2.586L15.414 6A2 2 0 0116 7.414V16a2 2 0 01-2 2H6a2 2 0 01-2-2V4zm2 6a1 1 0 011-1h6a1 1 0 110 2H7a1 1 0 01-1-1zm1 3a1 1 0 100 2h6a1 1 0 100-2H7z" clip-rule="evenodd"/>
                            </svg>
                        </div>
                        <div>
                            <h4 style="font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin:0;font-size:var(--font-size-sm)">Reports</h4>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">View & download</p>
                        </div>
                    </div>
                </a>
            </div>
        </div>
    </div>
    
    <script>
        // Drag & Drop File Upload
        (function() {{
            const dz = document.getElementById('dropzone');
            const input = document.getElementById('file-input');
            const note = document.getElementById('file-note');
            
            const updateNote = (file) => {{
                if (!file) {{
                    note.textContent = 'No file selected';
                    note.style.color = 'var(--color-gray-600)';
                    return;
                }}
                note.textContent = '✓ Selected: ' + file.name;
                note.style.color = 'var(--color-success)';
            }};
            
            dz.addEventListener('click', () => input.click());
            input.addEventListener('change', () => updateNote(input.files && input.files[0]));
            
            ['dragenter', 'dragover'].forEach(evt => {{
                dz.addEventListener(evt, (e) => {{
                    e.preventDefault();
                    e.stopPropagation();
                    dz.classList.add('dragover');
                }});
            }});
            
            ['dragleave', 'drop'].forEach(evt => {{
                dz.addEventListener(evt, (e) => {{
                    e.preventDefault();
                    e.stopPropagation();
                    dz.classList.remove('dragover');
                }});
            }});
            
            dz.addEventListener('drop', (e) => {{
                const files = e.dataTransfer && e.dataTransfer.files;
                if (!files || !files.length) return;
                try {{
                    const dt = new DataTransfer();
                    dt.items.add(files[0]);
                    input.files = dt.files;
                }} catch(err) {{
                    // Fallback for older browsers
                }}
                updateNote(files[0]);
            }});
        }})();
    </script>
</body>
</html>
    """
    return html


# ═══════════════════════════════════════════════════════════════════════════════
# PAY RATES MANAGEMENT ROUTES  
# ═══════════════════════════════════════════════════════════════════════════════
# Employee pay rate management - view, add, edit, delete rates

@app.route('/manage_rates')
@login_required
def manage_rates():
    """Manage employee pay rates"""
    username = session.get('username', 'Unknown')
    menu_html = get_menu_html(username)
    
    pay_rates = load_pay_rates()
    employee_names = get_employee_names()  # Get employee names for display
    employees = [{'id': emp_id, 'rate': rate, 'name': employee_names.get(emp_id, 'Unknown')} for emp_id, rate in pay_rates.items()]
    employees.sort(key=lambda x: x['id'])

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate, max-age=0">
    <meta http-equiv="Pragma" content="no-cache">
    <meta http-equiv="Expires" content="0">
    <!-- Cache Buster: 1765823691 -->
    <title>Pay Rates - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .rates-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
        .rate-display.hidden, .rate-edit.hidden, 
        .edit-btn.hidden, .save-btn.hidden, .cancel-btn.hidden {{
            display: none !important;
        }}
        .rate-edit {{
            width: 120px;
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="rates-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-2)">Employee Pay Rates</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-lg);margin:0">Manage hourly rates for all employees</p>
        </div>
    </div>
    
    <div class="container">
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">
                    <svg style="width:24px;height:24px;display:inline;margin-right:8px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                        <path d="M8.433 7.418c.155-.103.346-.196.567-.267v1.698a2.305 2.305 0 01-.567-.267C8.07 8.34 8 8.114 8 8c0-.114.07-.34.433-.582zM11 12.849v-1.698c.22.071.412.164.567.267.364.243.433.468.433.582 0 .114-.07.34-.433.582a2.305 2.305 0 01-.567.267z"/>
                        <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm1-13a1 1 0 10-2 0v.092a4.535 4.535 0 00-1.676.662C6.602 6.234 6 7.009 6 8c0 .99.602 1.765 1.324 2.246.48.32 1.054.545 1.676.662v1.941c-.391-.127-.68-.317-.843-.504a1 1 0 10-1.51 1.31c.562.649 1.413 1.076 2.353 1.253V15a1 1 0 102 0v-.092a4.535 4.535 0 001.676-.662C13.398 13.766 14 12.991 14 12c0-.99-.602-1.765-1.324-2.246A4.535 4.535 0 0011 9.092V7.151c.391.127.68.317.843.504a1 1 0 101.511-1.31c-.563-.649-1.413-1.076-2.354-1.253V5z" clip-rule="evenodd"/>
                    </svg>
                    Current Pay Rates
                </h2>
            </div>
            
            <div class="table-wrapper">
                <table class="table">
                    <thead>
                        <tr>
                            <th>Employee ID</th>
                            <th>Employee Name</th>
                            <th class="text-right">Pay Rate ($/hour)</th>
                            <th class="text-right">Actions</th>
                        </tr>
                    </thead>
                    <tbody>
"""
    
    for emp in employees:
        html += f"""
                        <tr id="row-{escape(emp['id'])}">
                            <td><span class="badge badge-primary">{escape(emp['id'])}</span></td>
                            <td><strong>{escape(emp['name'])}</strong></td>
                            <td class="text-right">
                                <span class="rate-display" style="color:var(--color-success);font-weight:var(--font-weight-semibold);font-size:var(--font-size-lg)">${emp['rate']}</span>
                                <input type="number" class="rate-edit hidden form-input" step="0.01" value="{emp['rate']}" data-original-value="{emp['rate']}">
                            </td>
                            <td class="text-right">
                                <button onclick="editRate('{escape(emp['id'])}')" class="edit-btn btn btn-primary btn-sm">
                                    <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                        <path d="M13.586 3.586a2 2 0 112.828 2.828l-.793.793-2.828-2.828.793-.793zM11.379 5.793L3 14.172V17h2.828l8.38-8.379-2.83-2.828z"/>
                                    </svg>
                                    Edit
                                </button>
                                <button onclick="saveRate('{escape(emp['id'])}')" class="save-btn hidden btn btn-success btn-sm">
                                    <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                        <path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd"/>
                                    </svg>
                                    Save
                                </button>
                                <button onclick="cancelEdit('{escape(emp['id'])}')" class="cancel-btn hidden btn btn-secondary btn-sm">
                                    Cancel
                                </button>
                                <form method="post" action="/delete_rate/{escape(emp['id'])}" style="display:inline;" onsubmit="return confirm('Delete rate for {escape(emp['name'])} ({escape(emp['id'])})?');">
                                    <button type="submit" class="btn btn-danger btn-sm">
                                        <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                            <path fill-rule="evenodd" d="M9 2a1 1 0 00-.894.553L7.382 4H4a1 1 0 000 2v10a2 2 0 002 2h8a2 2 0 002-2V6a1 1 0 100-2h-3.382l-.724-1.447A1 1 0 0011 2H9zM7 8a1 1 0 012 0v6a1 1 0 11-2 0V8zm5-1a1 1 0 00-1 1v6a1 1 0 102 0V8a1 1 0 00-1-1z" clip-rule="evenodd"/>
                                        </svg>
                                        Delete
                                    </button>
                                </form>
                            </td>
                        </tr>
"""
    
    html += """
                    </tbody>
                </table>
            </div>
        </div>
        
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">
                    <svg style="width:24px;height:24px;display:inline;margin-right:8px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M10 3a1 1 0 011 1v5h5a1 1 0 110 2h-5v5a1 1 0 11-2 0v-5H4a1 1 0 110-2h5V4a1 1 0 011-1z" clip-rule="evenodd"/>
                    </svg>
                    Add New Pay Rate
                </h2>
            </div>
            
            <form method="post" action="/add_rate" class="grid grid-cols-2" style="gap:var(--spacing-4);align-items:end">
                <div class="form-group">
                    <label for="employee_id" class="form-label form-label-required">Employee ID</label>
                    <input type="text" id="employee_id" name="employee_id" class="form-input" placeholder="e.g., EMP001" required>
                    <span class="form-help">Alphanumeric ID (letters, numbers, dash, underscore)</span>
                </div>
                
                <div class="form-group">
                    <label for="rate" class="form-label form-label-required">Pay Rate ($/hour)</label>
                    <input type="number" id="rate" name="rate" step="0.01" min="0" max="10000" class="form-input" placeholder="e.g., 25.00" required>
                    <span class="form-help">Between $0.00 and $10,000.00</span>
                </div>
                
                <div style="grid-column:1/-1;text-align:right">
                    <button type="submit" class="btn btn-success">
                        <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M10 3a1 1 0 011 1v5h5a1 1 0 110 2h-5v5a1 1 0 11-2 0v-5H4a1 1 0 110-2h5V4a1 1 0 011-1z" clip-rule="evenodd"/>
                        </svg>
                        Add Pay Rate
                    </button>
                </div>
            </form>
        </div>
    </div>
    
    <script>
        console.log('=== PAY RATES PAGE SCRIPT LOADED - v8.9.6 - 1765823691 ===');
        // Direct event listeners for edit/save/cancel buttons
        function editRate(employeeId) {{
            console.log('editRate called with ID:', employeeId);
            const row = document.getElementById('row-' + employeeId);
            if (!row) {{
                console.error('Row not found for employee ID:', employeeId);
                alert('Error: Could not find row for employee ' + employeeId);
                return;
            }}
            
            const rateDisplay = row.querySelector('.rate-display');
            const rateEdit = row.querySelector('.rate-edit');
            const editBtn = row.querySelector('.edit-btn');
            const saveBtn = row.querySelector('.save-btn');
            const cancelBtn = row.querySelector('.cancel-btn');
            
            if (!rateDisplay || !rateEdit || !editBtn || !saveBtn || !cancelBtn) {{
                console.error('Required elements not found in row');
                return;
            }}
            
            rateDisplay.classList.add('hidden');
            rateEdit.classList.remove('hidden');
            editBtn.classList.add('hidden');
            saveBtn.classList.remove('hidden');
            cancelBtn.classList.remove('hidden');
            rateEdit.focus();
            console.log('Edit mode activated for employee:', employeeId);
        }}
        
        function cancelEdit(employeeId) {{
            console.log('cancelEdit called with ID:', employeeId);
            const row = document.getElementById('row-' + employeeId);
            if (!row) return;
            
            const input = row.querySelector('.rate-edit');
            const originalRate = input.getAttribute('data-original-value') || input.value;
            input.value = originalRate;
            
            row.querySelector('.rate-display').classList.remove('hidden');
            row.querySelector('.rate-edit').classList.add('hidden');
            row.querySelector('.edit-btn').classList.remove('hidden');
            row.querySelector('.save-btn').classList.add('hidden');
            row.querySelector('.cancel-btn').classList.add('hidden');
        }}
        
        function saveRate(employeeId) {{
            console.log('saveRate called with ID:', employeeId);
            const row = document.getElementById('row-' + employeeId);
            if (!row) return;
            
            const newRate = row.querySelector('.rate-edit').value;
            if (!newRate || isNaN(newRate) || parseFloat(newRate) < 0) {{
                alert('Please enter a valid pay rate');
                return;
            }}
            
            const saveBtn = row.querySelector('.save-btn');
            saveBtn.disabled = true;
            saveBtn.innerHTML = '<svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/></svg> Saving...';
            
            fetch('/update_rate/' + employeeId, {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{rate: parseFloat(newRate)}})
            }}).then(response => {{
                if (response.ok) {{
                    location.reload();
                }} else {{
                    saveBtn.disabled = false;
                    saveBtn.innerHTML = '<svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd"/></svg> Save';
                    alert('Error updating rate. Please try again.');
                }}
            }}).catch(error => {{
                saveBtn.disabled = false;
                saveBtn.innerHTML = '<svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd"/></svg> Save';
                alert('Network error. Please check your connection.');
            }});
        }}

    </script>
</body>
</html>"""
    
    return html

@app.route('/add_rate', methods=['POST'])
@login_required
def add_rate():
    """Add a new pay rate with validation"""
    try:
        emp_id = request.form.get('employee_id', '').strip()
        rate_str = request.form.get('rate', '').strip()

        # Validate employee ID
        if not emp_id:
            return "Employee ID is required", 400
        
        if not re.match(r'^[a-zA-Z0-9_-]+$', emp_id):
            return "Invalid employee ID format", 400

        # Validate pay rate
        valid, error, pay_rate = validate_pay_rate(rate_str)
        if not valid:
            return error, 400

        # Load existing rates
        pay_rates = load_pay_rates()

        # Add new rate
        pay_rates[emp_id] = pay_rate

        # Save updated rates
        save_pay_rates(pay_rates)
        app.logger.info(f"Pay rate added for employee {emp_id}: ${pay_rate}")

        return redirect(url_for('manage_rates'))
    except Exception as e:
        app.logger.error(f"Error adding pay rate: {e}")
        return f"Error adding pay rate: {str(e)}", 400


@app.route('/update_rate/<employee_id>', methods=['POST'])
@login_required
def update_rate(employee_id):
    """Update employee pay rate with validation"""
    try:
        data = request.get_json()
        rate_value = data.get('rate', '')
        
        # Validate pay rate
        valid, error, new_rate = validate_pay_rate(rate_value)
        if not valid:
            return jsonify({'error': error}), 400
        
        pay_rates = load_pay_rates()
        pay_rates[employee_id] = new_rate
        save_pay_rates(pay_rates)
        app.logger.info(f"Pay rate updated for employee {employee_id}: ${new_rate}")
        
        return jsonify({'status': 'ok', 'rate': new_rate}), 200
    except Exception as e:
        app.logger.error(f"Error updating pay rate for {employee_id}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/delete_rate/<employee_id>', methods=['POST'])
@login_required
def delete_rate(employee_id):
    """Delete a pay rate"""
    try:
        # Load existing rates
        pay_rates = load_pay_rates()

        # Delete rate if exists
        if employee_id in pay_rates:
            del pay_rates[employee_id]

        # Save updated rates
        save_pay_rates(pay_rates)

        return redirect(url_for('manage_rates'))
    except Exception as e:
        return f"Error deleting pay rate: {str(e)}", 400

@app.route('/import_rates', methods=['POST'])
@login_required
def import_rates():
    """Import pay rates from CSV"""
    try:
        if 'rates_file' not in request.files:
            return "No file uploaded", 400

        file = request.files['rates_file']
        if file.filename == '':
            return "No file selected", 400

        # Check if it's a CSV
        if not file.filename.endswith('.csv'):
            return "Only CSV files allowed", 400

        # Save and process file
        file_path = os.path.join(UPLOAD_FOLDER, 'rates_' + file.filename)
        file.save(file_path)

        # Read CSV
        df = pd.read_csv(file_path)

        # Check for required columns
        required_cols = ['Person ID', 'Rate']
        if not all(col in df.columns for col in required_cols):
            return "CSV must have 'Person ID' and 'Rate' columns", 400

        # Load existing rates
        pay_rates = load_pay_rates()

        # Update rates
        for _, row in df.iterrows():
            emp_id = str(row['Person ID'])
            rate = float(row['Rate'])
            if rate > 0:
                pay_rates[emp_id] = rate

        # Save updated rates
        save_pay_rates(pay_rates)

        return redirect(url_for('manage_rates'))
    except Exception as e:
        return f"Error importing pay rates: {str(e)}", 400

def parse_work_hours(time_str):
    """Parse timesheet hours from string format"""
    try:
        if pd.isna(time_str) or str(time_str).strip() == '':
            return 0.0
        parts = list(map(float, str(time_str).split(':')))
        return round(parts[0] + parts[1]/60 + parts[2]/3600, 2)
    except:
        return 0.0

def compute_daily_hours(row):
    twh = row['Total Work Time(h)']
    # If Total Work Time is missing but Clock In/Out are present, derive it
    if pd.isna(twh) or str(twh).strip() == '':
        # Check if Clock In and Clock Out columns exist and have values
        if 'Clock In' in row and 'Clock Out' in row:
            clock_in = row['Clock In']
            clock_out = row['Clock Out']

            # Verify both values are not empty/null
            if (pd.notna(clock_in) and pd.notna(clock_out) and
                str(clock_in).strip() != '' and str(clock_out).strip() != ''):
                try:
                    # Try to parse the time values
                    start = datetime.strptime(str(clock_in).strip(), '%H:%M:%S')
                    end = datetime.strptime(str(clock_out).strip(), '%H:%M:%S')
                    diff = end - start

                    # Handle overnight shift (e.g., Clock In at 22:00:00, Clock Out at 06:00:00)
                    if diff.total_seconds() < 0:
                        diff += timedelta(days=1)

                    # Return hours rounded to 2 decimal places
                    hours = diff.total_seconds() / 3600
                    return round(hours, 2)
                except ValueError as e:
                    # If time parsing fails, log and return 0
                    app.logger.warning(f"Failed to parse time for Person ID {row.get('Person ID', 'Unknown')}: {e}")
                    return 0.0


        return 0.0
    # otherwise parse the provided Total Work Time
    return parse_work_hours(twh)


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
# Common styling and formatting functions for consistent Excel report generation

def excel_set_header(cell, text, size=14, bold=True):
    """
    Apply consistent header styling to a cell
    
    Args:
        cell: Excel cell object
        text: Header text
        size: Font size (default 14)
        bold: Bold font (default True)
    """
    cell.value = text
    cell.font = Font(bold=bold, size=size)


def excel_set_column_headers(ws, headers, row=1, start_col=1):
    """
    Create styled column headers with gray background
    
    Args:
        ws: Worksheet object
        headers: List of header strings
        row: Row number for headers (default 1)
        start_col: Starting column number (default 1)
    """
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + col_offset)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')


def excel_apply_borders(ws, start_row, end_row, start_col, end_col):
    """
    Apply thin borders to a range of cells
    
    Args:
        ws: Worksheet object
        start_row: Starting row number
        end_row: Ending row number
        start_col: Starting column number
        end_col: Ending column number
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def excel_set_column_widths(ws, widths):
    """
    Set column widths for better readability
    
    Args:
        ws: Worksheet object
        widths: Dict mapping column letters to widths, e.g. {'A': 12, 'B': 25}
    """
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def excel_format_currency(cell, value):
    """
    Format a cell as currency
    
    Args:
        cell: Excel cell object
        value: Numeric value
    """
    cell.value = value
    cell.number_format = '$#,##0.00'


def excel_add_creator_info(ws, creator, row=2):
    """
    Add creator/processor information to worksheet
    
    Args:
        ws: Worksheet object
        creator: Username/creator string
        row: Row number for creator info (default 2)
    """
    ws.cell(row=row, column=1).value = f"Processed by: {creator}"
    ws.cell(row=row, column=1).font = Font(size=10, italic=True)
    # Also store in hidden cell for reports page
    ws['AA1'] = creator


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT GENERATION FUNCTIONS  
# ═══════════════════════════════════════════════════════════════════════════════
# Functions to generate various Excel reports from payroll data

def create_excel_report(df, filename, creator=None):
    """Create an Excel report from the DataFrame with proper timesheet formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    # Get the creator (username) - use parameter or default to Unknown
    if not creator:
        creator = "Unknown"

    # Check if we have the original timesheet format
    is_timesheet_format = all(col in df.columns for col in
                             ['Person ID', 'First Name', 'Last Name', 'Date', 'Total Work Time(h)'])

    if is_timesheet_format:
        # This is the original timesheet format - process properly
        # Load pay rates
        pay_rates = load_pay_rates()

        # Calculate daily hours
        df['Daily Hours'] = df.apply(compute_daily_hours, axis=1)

        # Assign pay rates - use stored rates or default
        df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)

        # Calculate daily pay
        df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)

        # Calculate weekly totals per employee
        weekly_totals = df.groupby('Person ID').agg(
            Total_Hours=('Daily Hours', 'sum'),
            Weekly_Total=('Daily Pay', 'sum'),
            First_Name=('First Name', 'first'),
            Last_Name=('Last Name', 'first'),
            Rate=('Hourly Rate', 'first')
        ).reset_index()

        # Apply rounding
        weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
        weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
        weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)

        # Add header using helper function
        excel_set_header(ws['A1'], "Payroll Report", size=14)

        # Add processor information using helper function
        excel_add_creator_info(ws, creator, row=2)

        # Add column headers using helper function
        headers = ["Person ID", "Employee Name", "Total Hours", "Total Pay", "Rounded Pay"]
        excel_set_column_headers(ws, headers, row=4, start_col=1)

        # Add data rows - using iterrows instead of itertuples
        for i, (_, row) in enumerate(weekly_totals.iterrows(), 5):
            ws.cell(row=i, column=1).value = row['Person ID']
            ws.cell(row=i, column=2).value = f"{row['First_Name']} {row['Last_Name']}"
            ws.cell(row=i, column=3).value = round(row['Total_Hours'], 2)
            ws.cell(row=i, column=4).value = round(row['Weekly_Total'], 2)
            ws.cell(row=i, column=5).value = row['Rounded_Weekly']
    else:
        # Generic format - create a standard report

        # Add header using helper function
        excel_set_header(ws['A1'], "Payroll Report", size=14)

        # Add column headers using helper function
        headers = ["ID", "Name", "Total Hours", "Pay Rate", "Total Pay"]
        excel_set_column_headers(ws, headers, row=2, start_col=1)

        # For demo data, just put it in the report
        row_num = 3
        for i, row in df.iterrows():
            if 'ID' in df.columns and 'Name' in df.columns and 'Hours' in df.columns and 'Rate' in df.columns:
                # Use actual column data if available
                ws.cell(row=row_num, column=1).value = row['ID']
                ws.cell(row=row_num, column=2).value = row['Name']
                ws.cell(row=row_num, column=3).value = row['Hours']
                ws.cell(row=row_num, column=4).value = row['Rate']
                ws.cell(row=row_num, column=5).value = round(row['Hours'] * row['Rate'], 2)
            else:
                # Otherwise use row number as a placeholder
                ws.cell(row=row_num, column=1).value = i + 1
                ws.cell(row=row_num, column=2).value = f"Row {i + 1}"
                for col_idx, col_name in enumerate(df.columns, 3):
                    if col_idx <= 5:  # Only show up to 3 columns of data
                        ws.cell(row=row_num, column=col_idx).value = row[col_name]
            row_num += 1

    # Format cells
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5).number_format = '"$"#,##0'

    # Set column widths
    column_widths = {'A': 15, 'B': 25, 'C': 15, 'D': 15, 'E': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save the workbook
    report_path = os.path.join(REPORT_FOLDER, filename)
    wb.save(report_path)
    return report_path

def convert_excel_to_pdf(excel_path):
    """
    Convert an Excel admin report to PDF format with BOTH summary and detailed sections.
    Returns the PDF as BytesIO object for sending as response.
    """
    try:
        from openpyxl import load_workbook
        import pandas as pd
        
        # Load the Excel file
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, 
                              rightMargin=0.4*inch, leftMargin=0.4*inch,
                              topMargin=0.75*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=6,
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        section_header_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=10,
            spaceBefore=15,
            alignment=TA_CENTER
        )
        
        # Extract title from A1
        title_text = str(ws['A1'].value) if ws['A1'].value else "Payroll Report"
        title = Paragraph(title_text, title_style)
        elements.append(title)
        
        # Extract creator info
        creator_text = ""
        if ws['A2'].value and 'Processed by' in str(ws['A2'].value):
            creator_text = str(ws['A2'].value)
        elif ws['AA1'].value:
            creator_text = f"Processed by: {ws['AA1'].value}"
        
        if creator_text:
            creator = Paragraph(creator_text, subtitle_style)
            elements.append(creator)
        
        # PART 1: Extract summary table (starts at column H=8, row 3)
        summary_col_start = 8
        summary_data = []
        
        # Headers are in row 3
        header_row = []
        for col in range(summary_col_start, summary_col_start + 5):
            val = ws.cell(row=3, column=col).value
            header_row.append(str(val) if val else "")
        summary_data.append(header_row)
        
        # Data rows start at row 4, continue until we hit "GRAND TOTAL"
        grand_total_row_num = None
        for row in range(4, max_row + 1):
            first_cell = ws.cell(row=row, column=summary_col_start + 1).value
            if first_cell and 'GRAND TOTAL' in str(first_cell).upper():
                grand_total_row_num = row
                break
                
            # Check if row has data
            person_id = ws.cell(row=row, column=summary_col_start).value
            if person_id:
                row_data = []
                for col in range(summary_col_start, summary_col_start + 5):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        # Format appropriately
                        col_idx = col - summary_col_start
                        if col_idx in [3, 4]:  # Pay columns
                            if isinstance(cell.value, (int, float)):
                                row_data.append(f"${cell.value:,.2f}")
                            else:
                                row_data.append(str(cell.value))
                        elif col_idx == 2:  # Hours
                            if isinstance(cell.value, (int, float)):
                                row_data.append(f"{cell.value:.2f}")
                            else:
                                row_data.append(str(cell.value))
                        else:
                            row_data.append(str(cell.value))
                    else:
                        row_data.append("")
                summary_data.append(row_data)
        
        # Add GRAND TOTAL row
        if grand_total_row_num:
            grand_row = []
            for col in range(summary_col_start, summary_col_start + 5):
                cell = ws.cell(row=grand_total_row_num, column=col)
                if cell.value is not None:
                    col_idx = col - summary_col_start
                    if col_idx in [3, 4]:  # Pay columns
                        if isinstance(cell.value, (int, float)):
                            grand_row.append(f"${cell.value:,.2f}")
                        else:
                            grand_row.append(str(cell.value))
                    elif col_idx == 2:  # Hours
                        if isinstance(cell.value, (int, float)):
                            grand_row.append(f"{cell.value:.2f}")
                        else:
                            grand_row.append(str(cell.value))
                    else:
                        grand_row.append(str(cell.value))
                else:
                    grand_row.append("")
            summary_data.append(grand_row)
        
        app.logger.info(f"PDF: Summary table has {len(summary_data)} rows (including header)")
        
        # Create summary table
        col_widths = [0.8*inch, 2.0*inch, 1.0*inch, 1.0*inch, 1.0*inch]
        summary_table = Table(summary_data, colWidths=col_widths)
        
        # Style summary table
        table_styles = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1e40af')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ]
        
        # Bold the last row (GRAND TOTAL)
        if len(summary_data) > 1:
            last_row_idx = len(summary_data) - 1
            table_styles.extend([
                ('FONTNAME', (0, last_row_idx), (-1, last_row_idx), 'Helvetica-Bold'),
                ('BACKGROUND', (0, last_row_idx), (-1, last_row_idx), colors.HexColor('#e6e6fa')),
            ])
        
        summary_table.setStyle(TableStyle(table_styles))
        elements.append(summary_table)
        
        # PART 2: Extract "Detailed Breakdown by Employee" section
        # Get list of valid employee names from summary table (skip GRAND TOTAL)
        valid_employee_names = []
        for row_data in summary_data[1:]:  # Skip header
            if row_data[1] and 'GRAND TOTAL' not in row_data[1].upper():
                valid_employee_names.append(row_data[1].strip())
        
        app.logger.info(f"PDF: Valid employee names: {valid_employee_names}")
        
        # Find where detailed section starts
        detailed_start_row = None
        for row in range((grand_total_row_num or 10), min(max_row + 1, (grand_total_row_num or 10) + 10)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and 'Detailed Breakdown' in str(cell_val):
                detailed_start_row = row + 1
                break
        
        if detailed_start_row and detailed_start_row < max_row and valid_employee_names:
            elements.append(Spacer(1, 10))
            section_header = Paragraph("Detailed Breakdown by Employee", section_header_style)
            elements.append(section_header)
            elements.append(Spacer(1, 6))
            
            # The detailed section has 3 columns: starting at col 1, 8, and 15
            col_starts = [1, 8, 15]
            
            # Extract employee cards - only for valid employee names
            employee_data_list = []
            
            for current_row in range(detailed_start_row, max_row + 1):
                for col_start in col_starts:
                    cell = ws.cell(row=current_row, column=col_start)
                    cell_val = str(cell.value).strip() if cell.value else ""
                    
                    # Check if this cell contains a valid employee name
                    if cell_val in valid_employee_names:
                        emp_dict = {'name': cell_val, 'details': [], 'id_rate': ''}
                        
                        # Next row has ID and Rate
                        id_row = current_row + 1
                        id_cell = ws.cell(row=id_row, column=col_start).value
                        if id_cell and 'ID:' in str(id_cell):
                            emp_dict['id_rate'] = str(id_cell)
                        
                        # Skip header row, start at data
                        data_start = id_row + 2
                        for data_row in range(data_start, min(data_start + 10, max_row + 1)):
                            date_val = ws.cell(row=data_row, column=col_start).value
                            
                            # Stop if we hit Total: or empty or next employee
                            if not date_val or 'Total:' in str(date_val) or 'Rounded' in str(date_val) or 'Signature' in str(date_val):
                                break
                            
                            # Check if it's a date (contains /)
                            if '/' not in str(date_val):
                                continue
                                
                            # Extract row data
                            in_val = ws.cell(row=data_row, column=col_start + 1).value
                            out_val = ws.cell(row=data_row, column=col_start + 2).value
                            hours_val = ws.cell(row=data_row, column=col_start + 3).value
                            pay_val = ws.cell(row=data_row, column=col_start + 4).value
                            
                            # Only add if we have meaningful data
                            if hours_val and pay_val:
                                emp_dict['details'].append({
                                    'date': str(date_val) if date_val else '',
                                    'in': str(in_val)[:5] if in_val else '',
                                    'out': str(out_val)[:5] if out_val else '',
                                    'hours': f"{hours_val:.2f}" if isinstance(hours_val, (int, float)) else str(hours_val) if hours_val else '',
                                    'pay': f"${pay_val:.2f}" if isinstance(pay_val, (int, float)) else str(pay_val) if pay_val else ''
                                })
                        
                        if emp_dict['details']:  # Only add if we found data
                            employee_data_list.append(emp_dict)
                            app.logger.info(f"PDF: Found employee card for {cell_val} with {len(emp_dict['details'])} days")
            
            app.logger.info(f"PDF: Extracted {len(employee_data_list)} valid employee detail cards")
            
            # Render employee cards (2 per row for proper spacing - NO OVERLAP)
            for i in range(0, len(employee_data_list), 2):
                batch = employee_data_list[i:i+2]
                
                # Create side-by-side tables
                batch_tables = []
                for emp_dict in batch:
                    # Build mini table for this employee
                    emp_table_data = []
                    # Name row - merge across all columns to prevent wrapping
                    emp_table_data.append([{'content': emp_dict['name'], 'colspan': 5}])
                    # ID and Rate row - also merge
                    if emp_dict['id_rate']:
                        emp_table_data.append([{'content': emp_dict['id_rate'], 'colspan': 5}])
                    # Header row
                    emp_table_data.append(['Date', 'In', 'Out', 'Hrs', 'Pay'])
                    
                    # Data rows
                    for detail in emp_dict['details']:
                        emp_table_data.append([
                            detail['date'][:5],  # Just MM/DD
                            detail['in'],
                            detail['out'],
                            detail['hours'],
                            detail['pay']
                        ])
                    
                    # Flatten the table data (handle colspan)
                    flattened_data = []
                    for row in emp_table_data:
                        if isinstance(row[0], dict) and 'content' in row[0]:
                            # This is a merged cell row
                            flattened_data.append([row[0]['content']])
                        else:
                            flattened_data.append(row)
                    
                    # Create table with wider columns
                    emp_table = Table(flattened_data, colWidths=[0.65*inch, 0.65*inch, 0.65*inch, 0.5*inch, 0.75*inch])
                    
                    # Count header row position (after name and id/rate rows)
                    header_row_idx = 2 if emp_dict['id_rate'] else 1
                    
                    table_styles = [
                        ('FONTSIZE', (0, 0), (-1, -1), 7),
                        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),  # Name row
                        ('SPAN', (0, 0), (-1, 0)),  # Merge name row
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6e6e6')),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ]
                    
                    if emp_dict['id_rate']:
                        table_styles.extend([
                            ('SPAN', (0, 1), (-1, 1)),  # Merge ID/Rate row
                            ('FONTSIZE', (0, 1), (0, 1), 6),
                        ])
                    
                    # Header row styling
                    table_styles.extend([
                        ('FONTNAME', (0, header_row_idx), (-1, header_row_idx), 'Helvetica-Bold'),
                        ('BACKGROUND', (0, header_row_idx), (-1, header_row_idx), colors.HexColor('#1e40af')),
                        ('TEXTCOLOR', (0, header_row_idx), (-1, header_row_idx), colors.white),
                        ('GRID', (0, header_row_idx), (-1, -1), 0.5, colors.grey),
                    ])
                    
                    emp_table.setStyle(TableStyle(table_styles))
                    batch_tables.append(emp_table)
                
                # Create container with proper spacing
                if len(batch_tables) == 2:
                    container_data = [[batch_tables[0], '', batch_tables[1]]]
                    container_table = Table(container_data, colWidths=[3.5*inch, 0.25*inch, 3.5*inch])
                else:
                    container_table = Table([[batch_tables[0]]], colWidths=[3.5*inch])
                    
                container_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(container_table)
                elements.append(Spacer(1, 8))
        
        # Add footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer_text = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        app.logger.error(f"Error converting Excel to PDF: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        raise

def create_payslips(df, filename, creator=None):
    """Create individual payslips in Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Payslips"

    # Get the creator (username) - use parameter or default to Unknown
    if not creator:
        creator = "Unknown"

    # Disable grid lines through sheet properties to prevent Numbers from showing them
    ws.sheet_properties.showGridLines = False

    # Get week range for header
    try:
        start_date = pd.to_datetime(df['Date']).min().strftime('%Y-%m-%d')
        end_date = pd.to_datetime(df['Date']).max().strftime('%Y-%m-%d')
        date_range = f"{start_date} to {end_date}"
    except:
        date_range = "Current Period"

    # Add header
    ws['A1'] = f"Employee Payslips - {date_range}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add processor information
    ws['A2'] = f"Processed by: {creator}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:E2')

    # Store creator in hidden cell for reporting
    ws['AA1'] = creator

    # Load pay rates
    pay_rates = load_pay_rates()

    # Process data
    df['Daily Hours'] = df.apply(compute_daily_hours, axis=1)
    df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)
    df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)

    # Calculate totals per employee
    totals = df.groupby('Person ID').agg(
        Hours=('Daily Hours', 'sum'),
        Pay=('Daily Pay', 'sum'),
        First=('First Name', 'first'),
        Last=('Last Name', 'first'),
        Rate=('Hourly Rate', 'first')
    ).reset_index()

    # Add payslips for each employee
    row = 3
    for _, emp in totals.iterrows():
        # Employee header
        ws[f'A{row}'] = f"Employee: {emp['First']} {emp['Last']}"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        # Employee details
        ws[f'A{row}'] = "ID:"
        ws[f'B{row}'] = emp['Person ID']
        ws[f'C{row}'] = "Rate:"
        ws[f'D{row}'] = f"${emp['Rate']:.2f}/hour"
        row += 1

        # Add table headers
        headers = ["Date", "Clock In", "Clock Out", "Hours", "Pay"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
        row += 1

        # Add daily entries
        emp_df = df[df['Person ID'] == emp['Person ID']].sort_values('Date')
        for _, day in emp_df.iterrows():
            date_str = pd.to_datetime(day['Date']).strftime('%m/%d/%Y')
            ws.cell(row=row, column=1).value = date_str
            ws.cell(row=row, column=2).value = day['Clock In']
            ws.cell(row=row, column=3).value = day['Clock Out']
            ws.cell(row=row, column=4).value = day['Daily Hours']
            ws.cell(row=row, column=5).value = day['Daily Pay']
            row += 1

        # Add totals
        ws.cell(row=row, column=3).value = "Total:"
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = emp['Hours']
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5).value = emp['Pay']
        ws.cell(row=row, column=5).font = Font(bold=True)
        row += 1

        # Add rounded total
        ws.cell(row=row, column=3).value = "Rounded Pay:"
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=5).value = round(emp['Pay'])
        ws.cell(row=row, column=5).font = Font(bold=True)
        row += 2

        # Add signature line
        ws.cell(row=row, column=1).value = "Signature: _________________________"
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
        row += 3  # Space between employees

    # Format monetary values
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=5)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '"$"#,##0.00'

    # Set column widths
    column_widths = {'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save the workbook
    report_path = os.path.join(REPORT_FOLDER, filename)
    wb.save(report_path)
    return report_path

def create_combined_report(df, filename):
    """Create a combined report with summary and payslips with signatures"""
    wb = Workbook()

    # Create summary sheet
    ws_summary = wb.active
    ws_summary.title = "Payroll Summary"

    # Get week range for header
    try:
        start_date = pd.to_datetime(df['Date']).min().strftime('%Y-%m-%d')
        end_date = pd.to_datetime(df['Date']).max().strftime('%Y-%m-%d')
        date_range = f"{start_date} to {end_date}"
    except:
        date_range = "Current Period"

    # Add header
    ws_summary['A1'] = f"Payroll Summary - {date_range}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:E1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    # Load pay rates
    pay_rates = load_pay_rates()

    # Process data
    df['Daily Hours'] = df['Total Work Time(h)'].apply(parse_work_hours)
    df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)
    df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)

    # Calculate weekly totals per employee
    weekly_totals = df.groupby('Person ID').agg(
        Total_Hours=('Daily Hours', 'sum'),
        Weekly_Total=('Daily Pay', 'sum'),
        First_Name=('First Name', 'first'),
        Last_Name=('Last Name', 'first'),
        Rate=('Hourly Rate', 'first')
    ).reset_index()

    # Apply rounding
    weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
    weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
    weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)

    # Calculate grand totals
    grand_total_hours = weekly_totals['Total_Hours'].sum().round(2)
    grand_total_pay = weekly_totals['Weekly_Total'].sum().round(2)
    grand_total_rounded = weekly_totals['Rounded_Weekly'].sum()

    # Add column headers in row 3
    headers = ["Person ID", "Employee Name", "Total Hours", "Total Pay", "Rounded Pay"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    # Add summary data rows
    for i, (_, row) in enumerate(weekly_totals.iterrows(), 4):
        ws_summary.cell(row=i, column=1).value = row['Person ID']
        ws_summary.cell(row=i, column=2).value = f"{row['First_Name']} {row['Last_Name']}"
        ws_summary.cell(row=i, column=3).value = round(row['Total_Hours'], 2)
        ws_summary.cell(row=i, column=4).value = round(row['Weekly_Total'], 2)
        ws_summary.cell(row=i, column=5).value = row['Rounded_Weekly']

    # Add grand total row after employee rows
    grand_total_row = ws_summary.max_row + 1
    ws_summary.cell(row=grand_total_row, column=1).value = ""
    ws_summary.cell(row=grand_total_row, column=2).value = "GRAND TOTAL"
    ws_summary.cell(row=grand_total_row, column=2).font = Font(bold=True)
    ws_summary.cell(row=grand_total_row, column=3).value = grand_total_hours
    ws_summary.cell(row=grand_total_row, column=3).font = Font(bold=True)
    ws_summary.cell(row=grand_total_row, column=4).value = grand_total_pay
    ws_summary.cell(row=grand_total_row, column=4).font = Font(bold=True)
    ws_summary.cell(row=grand_total_row, column=5).value = grand_total_rounded
    ws_summary.cell(row=grand_total_row, column=5).font = Font(bold=True)
    ws_summary.cell(row=grand_total_row, column=3).number_format = '#,##0.00'
    ws_summary.cell(row=grand_total_row, column=4).number_format = '"$"#,##0.00'
    ws_summary.cell(row=grand_total_row, column=5).number_format = '"$"#,##0'

    # Add a header for detailed section
    current_row = grand_total_row + 2
    ws_summary.cell(row=current_row, column=1).value = "Detailed Breakdown by Employee"
    ws_summary.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws_summary.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2

    # Define border styles - only outer borders
    outer_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # For header cells only - bottom border
    header_border = Border(
        bottom=Side(style='thin', color='000000')
    )

    # For total rows only - top border
    total_border = Border(
        top=Side(style='thin', color='000000')
    )

    # Add detailed timesheet data for each employee
    for _, emp_data in weekly_totals.iterrows():
        emp_id = emp_data['Person ID']
        emp_name = f"{emp_data['First_Name']} {emp_data['Last_Name']}"

        # Employee header
        ws_summary.cell(row=current_row, column=1).value = f"Employee: {emp_name} (ID: {emp_id})"
        ws_summary.cell(row=current_row, column=1).font = Font(bold=True)
        ws_summary.merge_cells(f'A{current_row}:E{current_row}')
        ws_summary.cell(row=current_row, column=1).fill = PatternFill(start_color="E6E6E6", fill_type="solid")
        current_row += 1

        # Hourly rate
        ws_summary.cell(row=current_row, column=1).value = f"Hourly Rate: ${emp_data['Rate']:.2f}"
        ws_summary.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        # Mark the start of the employee section for border
        emp_section_start = current_row

        # Add table headers for daily entries
        detailed_headers = ["Date", "Clock In", "Clock Out", "Hours", "Pay"]
        for col, header in enumerate(detailed_headers, 1):
            cell = ws_summary.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")
            cell.border = header_border  # Only bottom border for headers
        current_row += 1

        # Add daily entries
        emp_df = df[df['Person ID'] == emp_id].sort_values('Date')
        for _, day in emp_df.iterrows():
            date_str = pd.to_datetime(day['Date']).strftime('%m/%d/%Y')
            ws_summary.cell(row=current_row, column=1).value = date_str
            ws_summary.cell(row=current_row, column=2).value = day['Clock In']
            ws_summary.cell(row=current_row, column=3).value = day['Clock Out']
            ws_summary.cell(row=current_row, column=4).value = day['Daily Hours']
            ws_summary.cell(row=current_row, column=5).value = day['Daily Pay']
            current_row += 1

        # Mark totals row start
        totals_row_start = current_row

        # Add totals with top border
        ws_summary.cell(row=current_row, column=3).value = "Total:"
        ws_summary.cell(row=current_row, column=3).font = Font(bold=True)
        ws_summary.cell(row=current_row, column=3).border = total_border
        ws_summary.cell(row=current_row, column=4).value = emp_data['Total_Hours']
        ws_summary.cell(row=current_row, column=4).font = Font(bold=True)
        ws_summary.cell(row=current_row, column=4).border = total_border
        ws_summary.cell(row=current_row, column=5).value = emp_data['Weekly_Total']
        ws_summary.cell(row=current_row, column=5).font = Font(bold=True)
        ws_summary.cell(row=current_row, column=5).border = total_border
        current_row += 1

        # Add rounded total
        ws_summary.cell(row=current_row, column=3).value = "Rounded Pay:"
        ws_summary.cell(row=current_row, column=3).font = Font(bold=True)
        ws_summary.cell(row=current_row, column=5).value = emp_data['Rounded_Weekly']
        ws_summary.cell(row=current_row, column=5).font = Font(bold=True)
        current_row += 1

        # Add outer borders to the entire employee section
        for row in range(emp_section_start, current_row):
            # Left border for first column
            ws_summary.cell(row=row, column=1).border = Border(
                left=Side(style='thin', color='000000')
            )
            # Right border for last column
            ws_summary.cell(row=row, column=5).border = Border(
                right=Side(style='thin', color='000000')
            )

        # Add top and bottom borders
        for col in range(1, 6):
            # Top border for first row
            border = ws_summary.cell(row=emp_section_start, column=col).border
            ws_summary.cell(row=emp_section_start, column=col).border = Border(
                left=border.left,
                right=border.right,
                top=Side(style='thin', color='000000'),
                bottom=border.bottom
            )

            # Bottom border for last row
            border = ws_summary.cell(row=current_row-1, column=col).border
            ws_summary.cell(row=current_row-1, column=col).border = Border(
                left=border.left,
                right=border.right,
                top=border.top,
                bottom=Side(style='thin', color='000000')
            )

        # Add signature line
        current_row += 2  # Add space before signature
        ws_summary.cell(row=current_row, column=1).value = "Signature: _________________________"
        ws_summary.merge_cells(f'A{current_row}:C{current_row}')
        ws_summary.cell(row=current_row, column=4).value = "Date: _____________"
        ws_summary.merge_cells(f'D{current_row}:E{current_row}')
        current_row += 2  # Add space between employees

    # Format monetary values in summary section
    for r in range(3, grand_total_row + 1):
        # Format pay columns
        ws_summary.cell(row=r, column=4).number_format = '"$"#,##0.00'
        ws_summary.cell(row=r, column=5).number_format = '"$"#,##0'

        # Format hours column
        if r > 3:  # Skip header row
            ws_summary.cell(row=r, column=3).number_format = '#,##0.00'

    # Add outer borders to the summary table
    for r in range(3, grand_total_row + 1):
        for c in range(1, 6):
            # Determine which borders to show
            left = Side(style='thin') if c == 1 else None
            right = Side(style='thin') if c == 5 else None
            top = Side(style='thin') if r == 3 else None
            bottom = Side(style='thin') if r == grand_total_row else None

            # Add any required borders
            if left or right or top or bottom:
                ws_summary.cell(row=r, column=c).border = Border(
                    left=left,
                    right=right,
                    top=top,
                    bottom=bottom
                )

    # Add horizontal borders for header row and grand total row
    for c in range(1, 6):
        # Border between header and data
        border = ws_summary.cell(row=3, column=c).border
        ws_summary.cell(row=3, column=c).border = Border(
            left=border.left,
            right=border.right,
            top=border.top,
            bottom=Side(style='thin', color='000000')
        )

        # Border above grand total
        border = ws_summary.cell(row=grand_total_row, column=c).border
        ws_summary.cell(row=grand_total_row, column=c).border = Border(
            left=border.left,
            right=border.right,
            top=Side(style='thin', color='000000'),
            bottom=border.bottom
        )

    # Set column widths
    column_widths = {'A': 15, 'B': 25, 'C': 15, 'D': 15, 'E': 15}
    for col, width in column_widths.items():
        ws_summary.column_dimensions[col].width = width

    # Save the workbook
    report_path = os.path.join(REPORT_FOLDER, filename)
    wb.save(report_path)
    return report_path

def create_consolidated_admin_report(df, filename, creator=None):
    """Create a single-sheet admin report with all employee data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Summary"

    # Disable grid lines through sheet properties to prevent Numbers from showing them
    ws.sheet_properties.showGridLines = False

    # Get week range for header
    try:
        start_date = pd.to_datetime(df['Date']).min().strftime('%Y-%m-%d')
        end_date = pd.to_datetime(df['Date']).max().strftime('%Y-%m-%d')
        date_range = f"{start_date} to {end_date}"
    except:
        date_range = "Current Period"

    # Add header
    ws['A1'] = f"Payroll Summary - {date_range}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:Z1')  # Merge across all columns used by the report
    ws['A1'].alignment = Alignment(horizontal='center')

    # Get the creator - use the provided parameter, or fall back to session username
    if not creator and 'username' in session:
        creator = session.get('username')
    elif not creator:
        creator = "Unknown"

    # Add processor information using helper (with manual merge for this report)
    ws['A2'].value = f"Processed by: {creator}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:Z2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Store the creator in hidden cell
    ws['AA1'] = creator

    # Continue with the existing function
    # Load pay rates
    pay_rates = load_pay_rates()

    # Process data
    df['Daily Hours'] = df.apply(compute_daily_hours, axis=1)
    df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)
    df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)

    # Calculate weekly totals per employee
    weekly_totals = df.groupby('Person ID').agg(
        Total_Hours=('Daily Hours', 'sum'),
        Weekly_Total=('Daily Pay', 'sum'),
        First_Name=('First Name', 'first'),
        Last_Name=('Last Name', 'first'),
        Rate=('Hourly Rate', 'first')
    ).reset_index()

    # Apply rounding
    weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
    weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
    weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)

    # Calculate grand totals
    grand_total_hours = weekly_totals['Total_Hours'].sum().round(2)
    grand_total_pay = weekly_totals['Weekly_Total'].sum().round(2)
    grand_total_rounded = weekly_totals['Rounded_Weekly'].sum()

    # Define minimal borders
    header_border = Border(bottom=Side(style='thin', color='000000'))
    top_border = Border(top=Side(style='hair', color='D3D3D3'))

    # Center the summary in the middle of the page
    summary_col_start = 8  # Center it better by moving to column H

    # Add column headers in row 3
    headers = ["Person ID", "Employee Name", "Total Hours", "Total Pay", "Rounded Pay"]
    for col, header in enumerate(headers):
        cell = ws.cell(row=3, column=summary_col_start + col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.border = header_border  # Only bottom border
        cell.alignment = Alignment(horizontal='center')

    # Add summary data rows
    for i, (_, row) in enumerate(weekly_totals.iterrows(), 4):
        ws.cell(row=i, column=summary_col_start).value = row['Person ID']
        ws.cell(row=i, column=summary_col_start+1).value = f"{row['First_Name']} {row['Last_Name']}"
        ws.cell(row=i, column=summary_col_start+2).value = round(row['Total_Hours'], 2)
        ws.cell(row=i, column=summary_col_start+3).value = round(row['Weekly_Total'], 2)
        ws.cell(row=i, column=summary_col_start+4).value = row['Rounded_Weekly']

    # Add grand total row after employee rows with a light top border
    grand_total_row = ws.max_row + 1
    for col in range(summary_col_start, summary_col_start+5):
        ws.cell(row=grand_total_row, column=col).border = top_border

    ws.cell(row=grand_total_row, column=summary_col_start).value = ""
    ws.cell(row=grand_total_row, column=summary_col_start+1).value = "GRAND TOTAL"
    ws.cell(row=grand_total_row, column=summary_col_start+1).font = Font(bold=True)
    ws.cell(row=grand_total_row, column=summary_col_start+2).value = grand_total_hours
    ws.cell(row=grand_total_row, column=summary_col_start+2).font = Font(bold=True)
    ws.cell(row=grand_total_row, column=summary_col_start+3).value = grand_total_pay
    ws.cell(row=grand_total_row, column=summary_col_start+3).font = Font(bold=True)
    ws.cell(row=grand_total_row, column=summary_col_start+4).value = grand_total_rounded
    ws.cell(row=grand_total_row, column=summary_col_start+4).font = Font(bold=True)

    # Add a header for detailed section with less spacing
    current_row = grand_total_row + 2
    ws.cell(row=current_row, column=1).value = "Detailed Breakdown by Employee"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:Z{current_row}')
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    current_row += 1

    # Define the three columns for employee data
    col1_start = 1
    col2_start = 8
    col3_start = 15
    col_width = 6

    # Process employees in batches of 3 across the page
    for batch_idx in range(0, len(weekly_totals), 3):
        # Get up to 3 employees for this row
        batch = weekly_totals.iloc[batch_idx:min(batch_idx+3, len(weekly_totals))]

        # Start a new row for this batch
        row_start = current_row
        max_rows_in_batch = 0

        # Process each employee (up to 3) in this batch
        for i, (_, emp_data) in enumerate(batch.iterrows()):
            # Set column start based on position (left, middle, right)
            if i == 0:
                col_start = col1_start
            elif i == 1:
                col_start = col2_start
            else:
                col_start = col3_start

            emp_id = emp_data['Person ID']
            emp_name = f"{emp_data['First_Name']} {emp_data['Last_Name']}"
            rate = emp_data['Rate']

            # Current row for this employee section
            emp_row = row_start

            # Employee header with shaded background
            ws.cell(row=emp_row, column=col_start).value = emp_name
            ws.cell(row=emp_row, column=col_start).font = Font(bold=True)
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+col_width-1)}{emp_row}')
            ws.cell(row=emp_row, column=col_start).fill = PatternFill(start_color="E6E6E6", fill_type="solid")
            emp_row += 1

            # ID and rate info
            ws.cell(row=emp_row, column=col_start).value = f"ID: {emp_id} | Rate: ${rate:.2f}"
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+col_width-1)}{emp_row}')
            emp_row += 1

            # Add table headers
            headers = ["Date", "In", "Out", "Hours", "Pay"]
            for j, header in enumerate(headers):
                if j < col_width:  # Stay within allocated width
                    cell = ws.cell(row=emp_row, column=col_start + j)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.border = header_border
            emp_row += 1

            # Add daily entries
            emp_df = df[df['Person ID'] == emp_id].sort_values('Date')
            for _, day in emp_df.iterrows():
                date_val = pd.to_datetime(day['Date']).strftime('%m/%d/%Y')
                ws.cell(row=emp_row, column=col_start).value = date_val
                ws.cell(row=emp_row, column=col_start+1).value = day['Clock In']
                ws.cell(row=emp_row, column=col_start+2).value = day['Clock Out']
                ws.cell(row=emp_row, column=col_start+3).value = day['Daily Hours']
                ws.cell(row=emp_row, column=col_start+4).value = day['Daily Pay']
                emp_row += 1

            # Add light top border for totals
            for j in range(col_start+3, col_start+6):
                if j < col_start+col_width:
                    ws.cell(row=emp_row, column=j).border = top_border

            # Add total hours and pay
            ws.cell(row=emp_row, column=col_start).value = "Total:"
            ws.cell(row=emp_row, column=col_start).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+3).value = emp_data['Total_Hours']
            ws.cell(row=emp_row, column=col_start+3).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+4).value = emp_data['Weekly_Total']
            ws.cell(row=emp_row, column=col_start+4).font = Font(bold=True)
            emp_row += 1

            # Add rounded pay
            ws.cell(row=emp_row, column=col_start).value = "Rounded Pay:"
            ws.cell(row=emp_row, column=col_start).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+4).value = emp_data['Rounded_Weekly']
            ws.cell(row=emp_row, column=col_start+4).font = Font(bold=True)
            emp_row += 1

            # Add signature line
            ws.cell(row=emp_row, column=col_start).value = "Signature: _______________"
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+3)}{emp_row}')
            emp_row += 1

            # Add date line
            ws.cell(row=emp_row, column=col_start).value = "Date: _________"
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+3)}{emp_row}')
            emp_row += 1

            # Format monetary values
            for r in range(row_start, emp_row):
                cell = ws.cell(row=r, column=col_start+4)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

            # Track max height
            rows_used = emp_row - row_start
            if rows_used > max_rows_in_batch:
                max_rows_in_batch = rows_used

        # Move to the next row after this batch, add some spacing
        current_row = row_start + max_rows_in_batch + 1

    # Format monetary values in the summary section
    for r in range(4, grand_total_row + 1):
        # Format pay columns
        ws.cell(row=r, column=summary_col_start+3).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=summary_col_start+4).number_format = '"$"#,##0'
        # Format hours column
        ws.cell(row=r, column=summary_col_start+2).number_format = '#,##0.00'

    # Format grand total row
    ws.cell(row=grand_total_row, column=summary_col_start+2).number_format = '#,##0.00'
    ws.cell(row=grand_total_row, column=summary_col_start+3).number_format = '"$"#,##0.00'
    ws.cell(row=grand_total_row, column=summary_col_start+4).number_format = '"$"#,##0'

    # Set optimized column widths
    ws.column_dimensions['A'].width = 10  # First column

    # Adjust width of each section (3 sections across the page)
    for i in range(col1_start, col1_start+col_width):
        ws.column_dimensions[get_column_letter(i)].width = 10
    for i in range(col2_start, col2_start+col_width):
        ws.column_dimensions[get_column_letter(i)].width = 10
    for i in range(col3_start, col3_start+col_width):
        ws.column_dimensions[get_column_letter(i)].width = 10

    # Adjust specific columns in each section
    for col_set in [col1_start, col2_start, col3_start]:
        ws.column_dimensions[get_column_letter(col_set)].width = 10      # Date
        ws.column_dimensions[get_column_letter(col_set+1)].width = 8     # In
        ws.column_dimensions[get_column_letter(col_set+2)].width = 8     # Out
        ws.column_dimensions[get_column_letter(col_set+3)].width = 6     # Hours
        ws.column_dimensions[get_column_letter(col_set+4)].width = 9     # Pay

    # Add space between columns for visual separation
    ws.column_dimensions[get_column_letter(col1_start+col_width)].width = 2
    ws.column_dimensions[get_column_letter(col2_start+col_width)].width = 2

    # Save the workbook
    report_path = os.path.join(REPORT_FOLDER, filename)
    wb.save(report_path)
    return report_path

def create_consolidated_payslips(df, filename, creator=None):
    """Create a single sheet with all employee payslips in a horizontal side-by-side layout"""
    # Import get_column_letter here to ensure it's available
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Payslips"

    # Get the creator (username) - use parameter or default to Unknown
    if not creator:
        creator = "Unknown"

    # Disable grid lines through sheet properties to prevent Numbers from showing them
    ws.sheet_properties.showGridLines = False

    # Get week range for header
    try:
        start_date = pd.to_datetime(df['Date']).min().strftime('%Y-%m-%d')
        end_date = pd.to_datetime(df['Date']).max().strftime('%Y-%m-%d')
        date_range = f"{start_date} to {end_date}"
    except:
        date_range = "Current Period"

    # Add title
    ws['A1'] = f"Employee Payslips - {date_range}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:Z1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add processor information
    ws['A2'] = f"Processed by: {creator}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:Z2')

    # Store creator in hidden cell for reporting
    ws['AA1'] = creator

    # Load pay rates
    pay_rates = load_pay_rates()

    # Process data
    df['Daily Hours'] = df.apply(compute_daily_hours, axis=1)
    df['Hourly Rate'] = df['Person ID'].astype(str).map(pay_rates).fillna(15.0)
    df['Daily Pay'] = (df['Daily Hours'] * df['Hourly Rate']).round(2)

    # Calculate weekly totals per employee
    weekly_totals = df.groupby('Person ID').agg(
        Total_Hours=('Daily Hours', 'sum'),
        Weekly_Total=('Daily Pay', 'sum'),
        First_Name=('First Name', 'first'),
        Last_Name=('Last Name', 'first'),
        Rate=('Hourly Rate', 'first')
    ).reset_index()

    # Apply rounding
    weekly_totals['Total_Hours'] = weekly_totals['Total_Hours'].round(2)
    weekly_totals['Weekly_Total'] = weekly_totals['Weekly_Total'].round(2)
    weekly_totals['Rounded_Weekly'] = weekly_totals['Weekly_Total'].round(0).astype(int)

    # IMPORTANT: Filter out employees with zero hours
    weekly_totals = weekly_totals[weekly_totals['Total_Hours'] > 0]

    # If no employees have hours, return an empty report
    if len(weekly_totals) == 0:
        ws['A3'] = "No employees with hours worked in this period"
        ws['A3'].font = Font(bold=True)
        report_path = os.path.join(REPORT_FOLDER, filename)
        wb.save(report_path)
        return report_path

    # Define a minimal header border - just a bottom line
    header_border = Border(
        bottom=Side(style='thin', color='000000')
    )

    # Each payslip takes 5 columns
    payslip_width = 5

    # Add a spacer column for easier cutting
    spacer_width = 1
    total_width_per_payslip = payslip_width + spacer_width

    # Start at row 3 (after title)
    start_row = 3

    # Calculate how many employees can fit per row
    max_payslips_per_row = 3  # Maximum 3 payslips side by side (18 columns total with spacers)

    # Process employees in batches for each row
    for batch_idx in range(0, len(weekly_totals), max_payslips_per_row):
        current_row = start_row
        batch_employees = weekly_totals.iloc[batch_idx:batch_idx+max_payslips_per_row]

        # Track the maximum height of any payslip in this row
        max_height = 0

        # Process each employee in this batch (for this row)
        for i, (_, emp_data) in enumerate(batch_employees.iterrows()):
            emp_id = emp_data['Person ID']
            emp_name = f"{emp_data['First_Name']} {emp_data['Last_Name']}"

            # Calculate starting column for this payslip
            col_start = 1 + (i * total_width_per_payslip)

            # Add dotted line in spacer column for cutting guide
            if i > 0:
                # Add vertical dotted line in the spacer column
                cut_col = col_start - 1

                # Use a light gray border for the cutting guide instead of values
                side = Side(style='dashDot', color='DDDDDD')
                border = Border(right=side)

                # Apply the border to a range of cells in the spacer column
                for r in range(current_row, current_row + 30):  # Reasonable height estimate
                    # Only set the border, not the value
                    cell = ws.cell(row=r, column=cut_col)
                    cell.border = border

            # Reset row counter for this employee
            emp_row = current_row

            # Employee header
            ws.cell(row=emp_row, column=col_start).value = f"Employee: {emp_name}"
            ws.cell(row=emp_row, column=col_start).font = Font(bold=True)
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+payslip_width-1)}{emp_row}')
            emp_row += 1

            # Pay period and employee ID in one row
            ws.cell(row=emp_row, column=col_start).value = f"Pay Period: {date_range}"
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+2)}{emp_row}')
            ws.cell(row=emp_row, column=col_start+3).value = f"ID: {emp_id}"
            ws.merge_cells(f'{get_column_letter(col_start+3)}{emp_row}:{get_column_letter(col_start+payslip_width-1)}{emp_row}')
            emp_row += 1

            # Hourly rate
            ws.cell(row=emp_row, column=col_start).value = f"Hourly Rate: ${emp_data['Rate']:.2f}"
            ws.merge_cells(f'{get_column_letter(col_start)}{emp_row}:{get_column_letter(col_start+payslip_width-1)}{emp_row}')
            emp_row += 1

            # Add table headers for daily entries
            detailed_headers = ["Date", "In", "Out", "Hours", "Pay"]
            for col, header in enumerate(detailed_headers, 0):
                cell = ws.cell(row=emp_row, column=col_start + col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")
                cell.border = header_border  # Only bottom border for headers
            emp_row += 1

            # Add daily entries
            emp_df = df[df['Person ID'] == emp_id].sort_values('Date')
            for _, day in emp_df.iterrows():
                date_str = pd.to_datetime(day['Date']).strftime('%m/%d/%Y')
                ws.cell(row=emp_row, column=col_start).value = date_str
                ws.cell(row=emp_row, column=col_start+1).value = day['Clock In']
                ws.cell(row=emp_row, column=col_start+2).value = day['Clock Out']
                ws.cell(row=emp_row, column=col_start+3).value = day['Daily Hours']
                ws.cell(row=emp_row, column=col_start+4).value = day['Daily Pay']
                emp_row += 1

            # Add space between days and totals
            emp_row += 1

            # Add totals with a light top border
            top_border = Border(top=Side(style='hair', color='D3D3D3'))

            # Total Hours
            ws.cell(row=emp_row, column=col_start+2).value = "Total Hours:"
            ws.cell(row=emp_row, column=col_start+2).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+2).border = top_border
            ws.cell(row=emp_row, column=col_start+3).value = emp_data['Total_Hours']
            ws.cell(row=emp_row, column=col_start+3).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+3).border = top_border
            emp_row += 1

            # Total Pay
            ws.cell(row=emp_row, column=col_start+2).value = "Total Pay:"
            ws.cell(row=emp_row, column=col_start+2).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+4).value = emp_data['Weekly_Total']
            ws.cell(row=emp_row, column=col_start+4).font = Font(bold=True)
            emp_row += 1

            # Rounded Pay
            ws.cell(row=emp_row, column=col_start+2).value = "Rounded Pay:"
            ws.cell(row=emp_row, column=col_start+2).font = Font(bold=True)
            ws.cell(row=emp_row, column=col_start+4).value = emp_data['Rounded_Weekly']
            ws.cell(row=emp_row, column=col_start+4).font = Font(bold=True)
            emp_row += 1

            # Format monetary values
            for r in range(current_row, emp_row):
                cell = ws.cell(row=r, column=col_start+4)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

            # Track the maximum height used by any payslip in this row
            if emp_row > current_row + max_height:
                max_height = emp_row - current_row

        # After processing all employees in this batch (row),
        # add a cut line below this row of payslips
        cut_line_row = current_row + max_height + 1

        # Prepare a cut line with scissors symbols at appropriate positions
        cut_line = ""
        for i in range(80):  # Sufficient length
            # Add scissors at approximate positions where cutting should occur
            if i > 0 and i % 25 == 0:
                cut_line += "✂️"
            else:
                cut_line += "-"

        # Add the horizontal cut line (BEFORE merging cells)
        ws.cell(row=cut_line_row, column=1).value = cut_line
        ws.cell(row=cut_line_row, column=1).font = Font(color="999999")
        ws.cell(row=cut_line_row, column=1).alignment = Alignment(horizontal='center')

        # Now merge the cells for the cut line
        ws.merge_cells(f'A{cut_line_row}:Z{cut_line_row}')

        # Set the starting row for the next batch (row) of employees
        start_row = cut_line_row + 2

    # Set column widths
    for col in range(1, 21):
        # Determine if this is a spacer column
        if (col - 1) % total_width_per_payslip == payslip_width:
            # This is a spacer column - make it narrow
            ws.column_dimensions[get_column_letter(col)].width = 2
        else:
            # Determine position within a payslip
            position = ((col - 1) % total_width_per_payslip)
            if position == 0:  # Date column
                width = 12
            elif position in [1, 2]:  # In/Out columns
                width = 8
            elif position == 3:  # Hours column
                width = 7
            else:  # Pay column
                width = 10
            ws.column_dimensions[get_column_letter(col)].width = width

    # Save the workbook
    report_path = os.path.join(REPORT_FOLDER, filename)
    wb.save(report_path)
    return report_path

def validate_timesheet(df):
    """Validate timesheet data and identify records with missing clock in/out times"""
    # Work with a copy to avoid modifying the original dataframe
    df_check = df.copy()
    
    # Create validation columns - True means there's an issue
    df_check['Missing_Clock_In'] = df_check['Clock In'].isna() | (df_check['Clock In'] == '')
    df_check['Missing_Clock_Out'] = df_check['Clock Out'].isna() | (df_check['Clock Out'] == '')
    df_check['Has_Issue'] = df_check['Missing_Clock_In'] | df_check['Missing_Clock_Out']

    # Return only records with issues (without the validation columns)
    issues = df_check[df_check['Has_Issue']].copy()
    # Drop the temporary validation columns
    issues = issues.drop(columns=['Missing_Clock_In', 'Missing_Clock_Out', 'Has_Issue'], errors='ignore')
    return issues

def get_unique_employees_from_df(df):
    """Extract unique employees from dataframe"""
    try:
        if 'Person ID' in df.columns and 'First Name' in df.columns and 'Last Name' in df.columns:
            employees = df[['Person ID', 'First Name', 'Last Name']].drop_duplicates()
            return employees.to_dict('records')
        return []
    except Exception:
        return []

@app.route('/validate', methods=['POST'])
def validate():
    """Validate uploaded CSV and show issues if any"""
    try:
        if 'file' not in request.files:
            return "No file part", 400

        file = request.files['file']
        if file.filename == '':
            return "No selected file", 400

        # Save uploaded file
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        # Store file path in session for later processing
        session['uploaded_file'] = file_path

        # For non-CSV files, redirect to regular processing
        if not file.filename.endswith('.csv'):
            return redirect(url_for('process'))

        # Read CSV
        df = pd.read_csv(file_path)

        # Check if this looks like a timesheet
        is_timesheet = all(col in df.columns for col in
                          ['Person ID', 'First Name', 'Last Name', 'Date', 'Clock In', 'Clock Out'])

        if not is_timesheet:
            # Not a timesheet, proceed with normal processing
            return redirect(url_for('process'))

        # SAFE VALIDATION - Check for missing times EARLY
        # Use the fixed validate_timesheet that works on a copy
        issues = validate_timesheet(df)
        
        if len(issues) > 0:
            # Found missing times - let user fix them BEFORE employee selection
            missing_records = []
            for idx, row in issues.iterrows():
                missing_records.append({
                    'index': idx,
                    'person_id': row['Person ID'],
                    'name': f"{row['First Name']} {row['Last Name']}",
                    'date': row['Date'],
                    'clock_in': row['Clock In'] if pd.notna(row['Clock In']) and row['Clock In'] != '' else '',
                    'clock_out': row['Clock Out'] if pd.notna(row['Clock Out']) and row['Clock Out'] != '' else ''
                })
            
            session['file_path'] = file_path
            session['missing_records'] = missing_records
            return redirect(url_for('fix_missing_times'))
        
        # No issues - go directly to employee confirmation
        return redirect(url_for('confirm_employees'))

        # Legacy code below (kept for reference but unreachable)

        # Calculate suggested times based on other employees
        def get_suggested_time(target_date, time_type, current_df):
            """Get suggested time based on other employees' times for the same date"""
            same_date_entries = current_df[current_df['Date'] == target_date]

            if time_type == 'Clock In':
                valid_times = same_date_entries[same_date_entries['Clock In'].notna()]['Clock In']
            else:
                valid_times = same_date_entries[same_date_entries['Clock Out'].notna()]['Clock Out']

            if len(valid_times) > 0:
                # Convert times to datetime objects for averaging
                times_list = []
                for time_str in valid_times:
                    if time_str and str(time_str).strip():
                        try:
                            time_obj = datetime.strptime(str(time_str).strip(), '%H:%M:%S')
                            times_list.append(time_obj)
                        except:
                            pass

                if times_list:
                    # Find the most common time range (within 30 minutes)
                    from collections import Counter
                    rounded_times = []
                    for t in times_list:
                        # Round to nearest 15 minutes
                        minutes = t.minute
                        rounded_min = round(minutes / 15) * 15
                        if rounded_min == 60:
                            rounded_time = t.replace(hour=(t.hour + 1) % 24, minute=0, second=0)
                        else:
                            rounded_time = t.replace(minute=rounded_min, second=0)
                        rounded_times.append(rounded_time.strftime('%H:%M:%S'))

                    # Get most common time
                    most_common = Counter(rounded_times).most_common(1)
                    if most_common:
                        return most_common[0][0]

            # Default suggestions if no data
            return '09:00:00' if time_type == 'Clock In' else '17:00:00'

        # Generate HTML table with issues
        issue_rows = ""
        for _, row in issues.iterrows():
            person_id = row['Person ID']
            name = f"{row['First Name']} {row['Last Name']}"
            date = row['Date']
            clock_in = row['Clock In'] if not row['Missing_Clock_In'] else ''
            clock_out = row['Clock Out'] if not row['Missing_Clock_Out'] else ''

            # Determine row class based on missing data
            row_class = ""
            if row['Missing_Clock_In'] and row['Missing_Clock_Out']:
                row_class = "both-missing"
            elif row['Missing_Clock_In'] or row['Missing_Clock_Out']:
                row_class = "one-missing"

            # Get suggestions for missing times
            clock_in_suggestion = ""
            clock_out_suggestion = ""
            if row['Missing_Clock_In']:
                suggested_in = get_suggested_time(date, 'Clock In', df)
                clock_in_suggestion = f'<br><span class="suggested">Suggested: {suggested_in}</span>'
            if row['Missing_Clock_Out']:
                suggested_out = get_suggested_time(date, 'Clock Out', df)
                clock_out_suggestion = f'<br><span class="suggested">Suggested: {suggested_out}</span>'

            issue_rows += f"""
            <tr class="{row_class}">
                <td>{person_id}</td>
                <td>{name}</td>
                <td>{date}</td>
                <td><input type="text" name="clock_in_{person_id}_{date}" value="{clock_in}"
                    placeholder="HH:MM:SS">{clock_in_suggestion}</td>
                <td><input type="text" name="clock_out_{person_id}_{date}" value="{clock_out}"
                    placeholder="HH:MM:SS">{clock_out_suggestion}</td>
            </tr>
            """

        # Generate HTML for validation form
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Fix Missing Clock Times</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }}
                h1, h2 {{ color: #333; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 8px; text-align: left; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; }}
                input[type="text"] {{ padding: 5px; width: 100px; }}
                .missing {{ background-color: #ffeeee; }}
                .both-missing {{ background-color: #ffcccc; }} /* Light red for both missing */
                .one-missing {{ background-color: #ffffcc; }} /* Light yellow for one missing */
                .suggested {{ color: #666; font-style: italic; font-size: 0.9em; }}
                .button {{
                    display: inline-block;
                    padding: 10px 15px;
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    cursor: pointer;
                    margin-right: 20px;
                    margin-bottom: 20px;
                }}
                .options {{
                    margin: 20px 0;
                    padding: 20px;
                    background-color: #f9f9f9;
                    border-radius: 5px;
                }}
                .btn-container {{
                    display: flex;
                    gap: 20px;
                    margin-top: 20px;
                }}
            </style>
        </head>
        <body>
            <h1>Fix Missing Clock Times</h1>

            <div class="options">
                <h2>The following timesheet entries have missing clock in or clock out times:</h2>
                <p>Fill in only the values you want to fix and leave the rest empty, or choose an option below.</p>
                <p><strong>Note:</strong> You can leave fields empty if you don't want to fix them - only fill in the times you need to correct.</p>
                <p><strong>Important:</strong> Both Clock In and Clock Out times are required to calculate hours. Entries with only one time will not appear in the payroll report.</p>
                <p><strong>Days Off:</strong> If an employee had the day off, leave both fields empty. They won't appear in the payroll report and won't be paid for that day.</p>

                <div style="margin: 15px 0; padding: 10px; border: 1px solid #ddd; border-radius: 5px;">
                    <strong>Color Legend:</strong>
                    <span style="display: inline-block; width: 20px; height: 15px; background-color: #ffcccc; border: 1px solid #ccc; vertical-align: middle;"></span> Both times missing
                    <span style="margin-left: 20px; display: inline-block; width: 20px; height: 15px; background-color: #ffffcc; border: 1px solid #ccc; vertical-align: middle;"></span> One time missing
                </div>

                <form action="/fix_times" method="post" id="fix-form">
                    <table>
                        <tr>
                            <th>ID</th>
                            <th>Name</th>
                            <th>Date</th>
                            <th>Clock In</th>
                            <th>Clock Out</th>
                        </tr>
                        {issue_rows}
                    </table>

                    <div class="btn-container">
                        <button type="submit" class="button">Process with Fixed Times</button>

                        <a href="/process_ignore" class="button" style="background-color: #f44336; text-decoration: none;">
                            Ignore Issues and Process Anyway
                        </a>
                    </div>
                </form>
            </div>

            <p><a href="/" style="text-decoration:none; color:#0275d8;">← Back to Home</a></p>
        </body>
        </html>
        """

        return html

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"Error validating file: {str(e)}<br><pre>{error_details}</pre>", 500

@app.route('/process_ignore')
def process_ignore():
    """Process the original file ignoring any issues"""
    try:
        # Get the file path from session
        file_path = session.get('uploaded_file')
        if not file_path:
            return "No file found in session. Please upload again.", 400

        # Read the CSV file
        df = pd.read_csv(file_path)

        # Get current username for report creation
        username = session.get('username', 'Unknown')

        # Continue with normal processing
        try:
            # Check if this looks like a timesheet
            is_timesheet = all(col in df.columns for col in
                             ['Person ID', 'First Name', 'Last Name', 'Date'])

            if is_timesheet:
                try:
                    df['Date'] = pd.to_datetime(df['Date'])
                    week_str = df['Date'].min().strftime('%Y-%m-%d')
                except:
                    week_str = datetime.now().strftime('%Y-%m-%d')
            else:
                week_str = datetime.now().strftime('%Y-%m-%d')

            # Generate reports
            reports = {}

            # Main payroll report
            summary_filename = f"payroll_summary_{week_str}.xlsx"
            summary_path = create_excel_report(df, summary_filename, username)
            reports['summary'] = summary_filename

            # Individual payslips
            if is_timesheet:
                payslips_filename = f"employee_payslips_{week_str}.xlsx"
                payslips_path = create_payslips(df, payslips_filename, username)
                reports['payslips'] = payslips_filename

                # NEW CONSOLIDATED SINGLE-SHEET REPORTS
                admin_filename = f"admin_report_{week_str}.xlsx"
                admin_path = create_consolidated_admin_report(df, admin_filename, username)
                reports['admin'] = admin_filename

                payslip_filename = f"payslips_for_cutting_{week_str}.xlsx"
                payslip_path = create_consolidated_payslips(df, payslip_filename, username)
                reports['payslips_sheet'] = payslip_filename

            # Store the reports in session
            session['reports'] = reports
            session['week'] = week_str

            return redirect(url_for('success'))

        except Exception as e:
            # If processing fails, create an error report
            import traceback
            txt_filename = "error_report.txt"
            report_path = os.path.join(REPORT_FOLDER, txt_filename)
            with open(report_path, 'w') as f:
                f.write(f"Error processing file: {str(e)}\n")
                f.write(traceback.format_exc())
            session['reports'] = {'error': txt_filename}

            return redirect(url_for('success'))

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"Error ignoring issues: {str(e)}<br><pre>{error_details}</pre>", 500

@app.route('/fix_times', methods=['POST'])
def fix_times():
    """Process the form with fixed clock times"""
    try:
        # Get file path from session - using uploaded_file instead of original_file
        file_path = session.get('uploaded_file')
        if not file_path:
            return "File not found in session. Please upload again.", 400

        # Extract the filename from the path
        filename = os.path.basename(file_path)

        # Read the CSV
        df = pd.read_csv(file_path)

        # Log form data for debugging
        app.logger.info(f"fix_times route called - CSV file: {file_path}")
        app.logger.debug(f"Form data received: {len(request.form)} fields")


        # Extract clock time fixes from form
        updates_made = []
        for key, value in request.form.items():
            if key.startswith('clock_in_') or key.startswith('clock_out_'):
                parts = key.split('_', 2)  # split into ['clock', 'in/out', 'person_id_date']
                clock_type = parts[0].capitalize() + ' ' + parts[1].capitalize()  # 'Clock In' or 'Clock Out'
                person_id_date = parts[2]  # 'person_id_date'

                # Further split person_id_date
                id_date_parts = person_id_date.split('_', 1)
                person_id = id_date_parts[0]
                date = id_date_parts[1]

                # Only update if value is not empty
                if value and value.strip():
                    app.logger.debug(f"Looking for Person ID: {person_id}, Date: {date}")

                    # Try to convert date format if needed
                    # The form might send YYYY-MM-DD but CSV might have MM/DD/YYYY or other format
                    matching_rows = []
                    for idx, row in df.iterrows():
                        row_person_id = str(row['Person ID'])
                        row_date = str(row['Date'])

                        # Try to match dates in different formats
                        if row_person_id == str(person_id):
                            # Direct string match
                            if row_date == date:
                                matching_rows.append(idx)
                            else:
                                # Try parsing both dates and comparing
                                try:
                                    # Convert different date formats
                                    # Form sends: 2025-06-12
                                    # CSV has: 06/12/2025 or similar

                                    # Method 1: Try parsing with pandas
                                    row_date_parsed = pd.to_datetime(row_date).date()
                                    form_date_parsed = pd.to_datetime(date).date()

                                    if row_date_parsed == form_date_parsed:
                                        matching_rows.append(idx)
                                        app.logger.debug(f"Date match found: {row_date} == {date}")
                                except Exception as e:
                                    # Method 2: Manual parsing for common formats
                                    try:
                                        # Check if form date is YYYY-MM-DD
                                        if '-' in date and len(date.split('-')[0]) == 4:
                                            year, month, day = date.split('-')
                                            # Check if CSV date is MM/DD/YYYY
                                            if '/' in row_date:
                                                parts = row_date.split('/')
                                                if len(parts) == 3:
                                                    csv_month, csv_day, csv_year = parts
                                                    if (csv_year == year and
                                                        csv_month.zfill(2) == month.zfill(2) and
                                                        csv_day.zfill(2) == day.zfill(2)):
                                                        matching_rows.append(idx)
                                                        app.logger.debug(f"Manual date match: {row_date} == {date}")
                                    except:
                                        pass

                    if matching_rows:
                        for idx in matching_rows:
                            before_val = df.at[idx, clock_type] if pd.notna(df.at[idx, clock_type]) else 'Empty'
                            df.at[idx, clock_type] = value.strip()
                            after_val = df.at[idx, clock_type]
                            updates_made.append(f"{clock_type} for Person ID {person_id} on {date}: {before_val} -> {after_val}")
                            app.logger.info(f"Updated row {idx}: {clock_type} for Person ID {person_id} on {date} from '{before_val}' to '{value.strip()}'")
                    else:
                        app.logger.warning(f"No matching row found for Person ID {person_id} on {date}")

        app.logger.info(f"Total updates made: {len(updates_made)}")
        for update in updates_made:
            app.logger.debug(f"  - {update}")

        # Save the updated CSV
        fixed_file_path = os.path.join(UPLOAD_FOLDER, f"fixed_{filename}")
        df.to_csv(fixed_file_path, index=False)

        app.logger.info(f"Saved fixed file to: {fixed_file_path}")
        app.logger.debug(f"First few rows after fix:\n{df.head(20)}")

        # Update the session with the fixed file path
        session['uploaded_file'] = fixed_file_path

        # Redirect to process_ignore to handle the fixed file
        return redirect(url_for('process_ignore'))

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"Error fixing times: {str(e)}<br><pre>{error_details}</pre>", 500

@app.route('/process', methods=['GET', 'POST'])
@login_required
def process():
    """Process the uploaded file"""
    try:
        # Check if this is a redirect from validation (file already in session)
        if request.method == 'GET' and 'uploaded_file' in session:
            file_path = session.get('uploaded_file')
            if not file_path or not os.path.exists(file_path):
                return "No file found in session. Please upload again.", 400
            filename = os.path.basename(file_path)
        else:
            # Original POST handling with file upload
            if 'file' not in request.files:
                return "No file part", 400

            file = request.files['file']
            if file.filename == '':
                return "No selected file", 400

            # Save the file
            file_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(file_path)
            filename = file.filename

        # Get current username for report creation
        username = session.get('username', 'Unknown')

        # For CSV files, create reports
        if filename.endswith('.csv'):
            try:
                # Read the CSV file
                df = pd.read_csv(file_path)

                # Check if this looks like a timesheet (has Person ID, Date, etc.)
                is_timesheet = all(col in df.columns for col in
                                  ['Person ID', 'First Name', 'Last Name', 'Date'])

                if is_timesheet:
                    # Check for missing clock in/out values
                    missing_data = False
                    missing_records = []

                    for idx, row in df.iterrows():
                        # Check for empty/null values in clock in/out
                        if pd.isna(row['Clock In']) or str(row['Clock In']).strip() == '' or \
                           pd.isna(row['Clock Out']) or str(row['Clock Out']).strip() == '':
                            missing_data = True
                            missing_records.append({
                                'index': idx,
                                'person_id': row['Person ID'],
                                'name': f"{row['First Name']} {row['Last Name']}",
                                'date': row['Date'],
                                'clock_in': '' if pd.isna(row['Clock In']) else row['Clock In'],
                                'clock_out': '' if pd.isna(row['Clock Out']) else row['Clock Out']
                            })

                    # If missing data, store in session and redirect to correction page
                    if missing_data:
                        session['file_path'] = file_path
                        session['missing_records'] = missing_records
                        return redirect(url_for('fix_missing_times'))

                    # Try to parse dates
                    try:
                        df['Date'] = pd.to_datetime(df['Date'])
                        week_str = df['Date'].min().strftime('%Y-%m-%d')
                    except:
                        week_str = datetime.now().strftime('%Y-%m-%d')
                else:
                    week_str = datetime.now().strftime('%Y-%m-%d')

                # Generate reports
                reports = {}

                # Get current username for report creation
                username = session.get('username', 'Unknown')

                # Main payroll report
                summary_filename = f"payroll_summary_{week_str}.xlsx"
                summary_path = create_excel_report(df, summary_filename, username)
                reports['summary'] = summary_filename

                # Individual payslips
                if is_timesheet:
                    payslips_filename = f"employee_payslips_{week_str}.xlsx"
                    payslips_path = create_payslips(df, payslips_filename, username)
                    reports['payslips'] = payslips_filename

                    # NEW CONSOLIDATED SINGLE-SHEET REPORTS
                    admin_filename = f"admin_report_{week_str}.xlsx"
                    admin_path = create_consolidated_admin_report(df, admin_filename, username)
                    reports['admin'] = admin_filename

                    payslip_filename = f"payslips_for_cutting_{week_str}.xlsx"
                    payslip_path = create_consolidated_payslips(df, payslip_filename, username)
                    reports['payslips_sheet'] = payslip_filename

                # Store the reports in session
                session['reports'] = reports
                session['week'] = week_str

                return redirect(url_for('success'))

            except Exception as e:
                # If processing fails, create an error report
                import traceback
                txt_filename = "error_report.txt"
                report_path = os.path.join(REPORT_FOLDER, txt_filename)
                with open(report_path, 'w') as f:
                    f.write(f"Error processing {filename}: {str(e)}\n")
                    f.write(traceback.format_exc())
                session['reports'] = {'error': txt_filename}

                return redirect(url_for('success'))
        else:
            # For non-CSV files, create a simple text report
            txt_filename = "file_report.txt"
            report_path = os.path.join(REPORT_FOLDER, txt_filename)
            with open(report_path, 'w') as f:
                f.write(f"Report for {filename}\n")
                f.write(f"File size: {os.path.getsize(file_path)} bytes\n")
            session['reports'] = {'file': txt_filename}

            return redirect(url_for('success'))

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"Error: {str(e)}<br><pre>{error_details}</pre>", 500

@app.route('/fix_missing_times', methods=['GET', 'POST'])
@login_required
def fix_missing_times():
    """Page to fix missing clock in/out times"""
    if request.method == 'POST':
        # Get the original file path
        file_path = session.get('file_path')
        if not file_path:
            return "File not found", 404

        # Load the original dataframe
        df = pd.read_csv(file_path)

        # Get all the fixes from the form
        for key, value in request.form.items():
            if key.startswith('fix_'):
                _, action, idx = key.split('_')
                idx = int(idx)

                if action == 'clockin':
                    df.at[idx, 'Clock In'] = value if value.strip() else None
                elif action == 'clockout':
                    df.at[idx, 'Clock Out'] = value if value.strip() else None
            elif key == 'action':
                action_type = value  # 'fix' or 'ignore'

        # Save the updated dataframe
        df.to_csv(file_path, index=False)
        
        # Update session with fixed file
        session['uploaded_file'] = file_path

        # After fixing times, go to employee confirmation
        # RESTORED ORIGINAL WORKING ORDER
        return redirect(url_for('confirm_employees'))

    # GET request - show the form
    try:
        missing_records = session.get('missing_records', [])
        
        # If no missing records, redirect to process
        if not missing_records:
            return redirect(url_for('process_confirmed'))

        # Get the original dataframe to calculate suggestions
        file_path = session.get('file_path')
        df = None
        if file_path and os.path.exists(file_path):
            df = pd.read_csv(file_path)

        # Function to calculate suggested times
        def get_suggested_time(target_date, time_type, current_df):
            """Get suggested time based on other employees' times for the same date"""
            if current_df is None:
                return '09:00:00' if time_type == 'Clock In' else '17:00:00'

            same_date_entries = current_df[current_df['Date'] == target_date]

            if time_type == 'Clock In':
                valid_times = same_date_entries[same_date_entries['Clock In'].notna()]['Clock In']
            else:
                valid_times = same_date_entries[same_date_entries['Clock Out'].notna()]['Clock Out']

            if len(valid_times) > 0:
                # Convert times to datetime objects for averaging
                times_list = []
                for time_str in valid_times:
                    if time_str and str(time_str).strip():
                        try:
                            time_obj = datetime.strptime(str(time_str).strip(), '%H:%M:%S')
                            times_list.append(time_obj)
                        except:
                            pass

                if times_list:
                    # Find the most common time range (within 30 minutes)
                    from collections import Counter
                    rounded_times = []
                    for t in times_list:
                        # Round to nearest 15 minutes
                        minutes = t.minute
                        rounded_min = round(minutes / 15) * 15
                        if rounded_min == 60:
                            rounded_time = t.replace(hour=(t.hour + 1) % 24, minute=0, second=0)
                        else:
                            rounded_time = t.replace(minute=rounded_min, second=0)
                        rounded_times.append(rounded_time.strftime('%H:%M:%S'))

                    # Get most common time
                    most_common = Counter(rounded_times).most_common(1)
                    if most_common:
                        return most_common[0][0]

            # Default suggestions if no data
            return '09:00:00' if time_type == 'Clock In' else '17:00:00'

        username = session.get('username', 'Unknown')
        menu_html = get_menu_html(username)
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Fix Missing Time Entries - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .fix-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
        .both-missing {{ background-color: var(--color-danger-light) !important; }}
        .one-missing {{ background-color: var(--color-warning-light) !important; }}
        .suggested {{
            color: var(--color-gray-600);
            font-style: italic;
            font-size: var(--font-size-xs);
            margin-top: var(--spacing-1);
            display: block;
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="fix-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-1);font-size:var(--font-size-2xl)">Fix Missing Time Entries</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-sm);margin:0">Fill in missing Clock In/Out times or ignore entries</p>
        </div>
    </div>
    
    <div class="container">
        <div class="alert alert-info">
            <svg style="width:20px;height:20px;flex-shrink:0" fill="currentColor" viewBox="0 0 20 20">
                <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7-4a1 1 0 11-2 0 1 1 0 012 0zM9 9a1 1 0 000 2v3a1 1 0 001 1h1a1 1 0 100-2v-3a1 1 0 00-1-1H9z" clip-rule="evenodd"/>
            </svg>
            <div>
                <p style="margin-bottom:var(--spacing-2)">Some entries have missing Clock In or Clock Out values. Fill in only the values you want to fix and leave the rest empty, or choose to ignore these entries.</p>
                <p style="margin-bottom:var(--spacing-2)"><strong>Time format:</strong> HH:MM:SS (e.g., 09:00:00)</p>
                <p style="margin-bottom:var(--spacing-2)"><strong>Note:</strong> You can leave fields empty if you don't want to fix them - only fill in the times you need to correct.</p>
                <p style="margin-bottom:0"><strong>Important:</strong> Both Clock In and Clock Out times are required to calculate hours. Entries with only one time will not appear in the payroll report.</p>
            </div>
        </div>
        
        <div class="card" style="margin-bottom:var(--spacing-3)">
            <div style="display:flex;gap:var(--spacing-4);align-items:center;flex-wrap:wrap">
                <strong style="color:var(--color-gray-900)">Color Legend:</strong>
                <div style="display:flex;align-items:center;gap:var(--spacing-2)">
                    <span style="display:inline-block;width:20px;height:15px;background-color:var(--color-danger-light);border:1px solid var(--color-danger);border-radius:2px"></span>
                    <span style="font-size:var(--font-size-sm);color:var(--color-gray-700)">Both times missing</span>
                </div>
                <div style="display:flex;align-items:center;gap:var(--spacing-2)">
                    <span style="display:inline-block;width:20px;height:15px;background-color:var(--color-warning-light);border:1px solid var(--color-warning);border-radius:2px"></span>
                    <span style="font-size:var(--font-size-sm);color:var(--color-gray-700)">One time missing</span>
                </div>
            </div>
        </div>

        <form action="/fix_missing_times" method="post">
            <div class="card">
                <div class="table-wrapper">
                    <table class="table">
                        <thead>
                            <tr>
                                <th>Employee</th>
                                <th>Date</th>
                                <th>Clock In</th>
                                <th>Clock Out</th>
                            </tr>
                        </thead>
                        <tbody>
    """

        for record in missing_records:
            # Determine row class based on missing data
            row_class = ""
            if not record['clock_in'] and not record['clock_out']:
                row_class = "both-missing"
            elif not record['clock_in'] or not record['clock_out']:
                row_class = "one-missing"

            # Get suggestions for missing times
            clock_in_suggestion = ""
            clock_out_suggestion = ""
            if not record['clock_in'] and df is not None:
                suggested_in = get_suggested_time(record['date'], 'Clock In', df)
                clock_in_suggestion = f'<span class="suggested">Suggested: {suggested_in}</span>'
            if not record['clock_out'] and df is not None:
                suggested_out = get_suggested_time(record['date'], 'Clock Out', df)
                clock_out_suggestion = f'<span class="suggested">Suggested: {suggested_out}</span>'

            html += f"""
                            <tr class="{row_class}">
                                <td><strong>{escape(record['name'])}</strong><br><span style="font-size:var(--font-size-xs);color:var(--color-gray-600)">ID: {escape(str(record['person_id']))}</span></td>
                                <td>{escape(str(record['date']))}</td>
                                <td>
                                    <input type="text" name="fix_clockin_{record['index']}" value="{escape(str(record['clock_in'])) if record['clock_in'] else ''}" class="form-input" placeholder="e.g., 09:00:00" style="width:100%;margin-bottom:var(--spacing-1)">
                                    {clock_in_suggestion}
                                </td>
                                <td>
                                    <input type="text" name="fix_clockout_{record['index']}" value="{escape(str(record['clock_out'])) if record['clock_out'] else ''}" class="form-input" placeholder="e.g., 17:00:00" style="width:100%;margin-bottom:var(--spacing-1)">
                                    {clock_out_suggestion}
                                </td>
                            </tr>
            """

        html += """
                        </tbody>
                    </table>
                </div>
                
                <div style="margin-top:var(--spacing-4);display:flex;gap:var(--spacing-3);justify-content:flex-end">
                    <button type="submit" name="action" value="ignore" class="btn btn-secondary">Ignore Missing Values</button>
                    <button type="submit" name="action" value="fix" class="btn btn-success">
                        <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd"/>
                        </svg>
                        Fix and Continue
                    </button>
                </div>
            </form>
        </div>
    </div>
</body>
</html>
        """

        # Return HTML directly (already fully built with f-strings)
        return html
    
    except Exception as e:
        import traceback
        error_msg = f"Error in fix_missing_times: {str(e)}\n\n{traceback.format_exc()}"
        return f"""
        <html><head><title>Error</title></head><body>
        <h1>Error Loading Fix Times Page</h1>
        <pre style="background:#f8d7da; padding:20px; border-radius:8px; color:#721c24;">{error_msg}</pre>
        <br><a href="/" style="padding:10px 20px; background:#007bff; color:white; text-decoration:none; border-radius:5px;">Go Home</a>
        </body></html>
        """, 500

@app.route('/success')
@login_required
def success():
    """Success page with download links"""
    reports = session.get('reports', {})
    week = session.get('week', datetime.now().strftime('%Y-%m-%d'))
    username = session.get('username', 'Unknown')
    menu_html = get_menu_html(username)

    # Clear the report cache
    clear_report_cache()

    # Start HTML with design system - COMPACT LAYOUT
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Payroll Complete - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .success-header {{
            background: linear-gradient(135deg, #059669 0%, #10b981 100%);
            color: white;
            padding: var(--spacing-3) 0;
            margin-bottom: var(--spacing-3);
        }}
        .success-banner {{
            background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
            border: 2px solid var(--color-success);
            border-radius: var(--radius-lg);
            padding: var(--spacing-3);
            text-align: center;
            margin-bottom: var(--spacing-3);
        }}
        .success-icon {{
            font-size: 32px;
            color: var(--color-success);
            margin-bottom: var(--spacing-1);
        }}
        .success-layout {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: var(--spacing-3);
            align-items: start;
        }}
        @media (max-width: 1024px) {{
            .success-layout {{
                grid-template-columns: 1fr;
            }}
        }}
        .report-item {{
            padding: var(--spacing-2);
            margin-bottom: var(--spacing-2);
            border-bottom: 1px solid var(--color-gray-200);
        }}
        .report-item:last-child {{
            border-bottom: none;
            margin-bottom: 0;
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="success-header">
        <div class="container">
            <div style="display:flex;justify-content:space-between;align-items:center">
                <div>
                    <h1 style="color:white;margin-bottom:0;font-size:var(--font-size-xl)">Payroll Complete</h1>
                    <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-sm);margin:0">Week: {week}</p>
                </div>
                <div style="font-size:48px;color:rgba(255,255,255,0.9)">✓</div>
            </div>
        </div>
    </div>
    
    <div class="container">
        <div class="success-layout">
    """

    if 'admin' in reports and 'payslips_sheet' in reports:
        html += f"""
            <!-- Reports Card -->
            <div class="card" style="margin-bottom:0">
                <div class="card-header" style="padding-bottom:var(--spacing-2);margin-bottom:var(--spacing-2)">
                    <h2 class="card-title" style="font-size:var(--font-size-base)">
                        <svg style="width:20px;height:20px;display:inline;margin-right:6px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M4 4a2 2 0 012-2h4.586A2 2 0 0112 2.586L15.414 6A2 2 0 0116 7.414V16a2 2 0 01-2 2H6a2 2 0 01-2-2V4zm2 6a1 1 0 011-1h6a1 1 0 110 2H7a1 1 0 01-1-1zm1 3a1 1 0 100 2h6a1 1 0 100-2H7z" clip-rule="evenodd"/>
                        </svg>
                        Recommended Reports
                    </h2>
                </div>
                
                <!-- Admin Report -->
                <div class="report-item">
                    <div style="display:flex;justify-content:space-between;align-items:start;margin-bottom:var(--spacing-2)">
                        <div style="flex:1">
                            <h3 style="font-size:var(--font-size-sm);font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:var(--spacing-1)">Admin Report</h3>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">All employee data with signature lines</p>
                        </div>
                        <div style="display:flex;gap:var(--spacing-2)">
                            <a href="/download/admin" class="btn btn-primary btn-sm">
                                <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd"/>
                                </svg>
                                Download
                            </a>
                            <a href="/print/admin" target="_blank" class="btn btn-secondary btn-sm">
                                <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M5 4v3H4a2 2 0 00-2 2v3a2 2 0 002 2h1v2a2 2 0 002 2h6a2 2 0 002-2v-2h1a2 2 0 002-2V9a2 2 0 00-2-2h-1V4a2 2 0 00-2-2H7a2 2 0 00-2 2zm8 0H7v3h6V4zm0 8H7v4h6v-4z" clip-rule="evenodd"/>
                                </svg>
                                Print
                            </a>
                        </div>
                    </div>
                    <p style="font-size:var(--font-size-xs);color:var(--color-gray-500);margin:0"><a href="/static/reports/{reports['admin']}" style="color:var(--color-primary);text-decoration:underline">{reports['admin']}</a></p>
                </div>
                
                <!-- Payslips Report -->
                <div class="report-item">
                    <div style="display:flex;justify-content:space-between;align-items:start;margin-bottom:var(--spacing-2)">
                        <div style="flex:1">
                            <h3 style="font-size:var(--font-size-sm);font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:var(--spacing-1)">Cuttable Payslips</h3>
                            <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">All payslips with cut lines for distribution</p>
                        </div>
                        <div style="display:flex;gap:var(--spacing-2)">
                            <a href="/download/payslips_sheet" class="btn btn-primary btn-sm">
                                <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd"/>
                                </svg>
                                Download
                            </a>
                            <a href="/print/payslips" target="_blank" class="btn btn-secondary btn-sm">
                                <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M5 4v3H4a2 2 0 00-2 2v3a2 2 0 002 2h1v2a2 2 0 002 2h6a2 2 0 002-2v-2h1a2 2 0 002-2V9a2 2 0 00-2-2h-1V4a2 2 0 00-2-2H7a2 2 0 00-2 2zm8 0H7v3h6V4zm0 8H7v4h6v-4z" clip-rule="evenodd"/>
                                </svg>
                                Print
                            </a>
                        </div>
                    </div>
                    <p style="font-size:var(--font-size-xs);color:var(--color-gray-500);margin:0"><a href="/static/reports/{reports['payslips_sheet']}" style="color:var(--color-primary);text-decoration:underline">{reports['payslips_sheet']}</a></p>
                </div>
            </div>
            
            <!-- Zoho Books Integration -->
            <div class="card" style="margin-bottom:0;background:linear-gradient(135deg, #dbeafe 0%, #e0e7ff 100%);border-color:var(--color-primary)">
                <div class="card-header" style="padding-bottom:var(--spacing-2);margin-bottom:var(--spacing-2)">
                    <h2 class="card-title" style="font-size:var(--font-size-base)">
                        <svg style="width:20px;height:20px;display:inline;margin-right:6px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M4 4a2 2 0 012-2h4.586A2 2 0 0112 2.586L15.414 6A2 2 0 0116 7.414V16a2 2 0 01-2 2H6a2 2 0 01-2-2V4z" clip-rule="evenodd"/>
                        </svg>
                        Zoho Books Integration
                    </h2>
                </div>
                
                <p style="font-size:var(--font-size-xs);color:var(--color-gray-700);margin-bottom:var(--spacing-3)">Automatically create an expense and attach the admin report</p>
                
                <form id="zoho-expense-form" action="/zoho/create_expense" method="post">
                    <div class="form-group" style="margin-bottom:var(--spacing-2)">
                        <label for="company" class="form-label" style="font-size:var(--font-size-xs)">Company</label>
                        <select id="company" name="company" class="form-select" style="padding:var(--spacing-2)">
                            <option value="haute">Haute Brands</option>
                            <option value="boomin">Boomin Brands</option>
                        </select>
                    </div>
                    <input type="hidden" name="week" value="{week}">

                    <div class="form-group" style="margin-bottom:var(--spacing-3)">
                        <label for="custom_desc" class="form-label" style="font-size:var(--font-size-xs)">Notes (optional)</label>
                        <input type="text" id="custom_desc" name="custom_desc" class="form-input" style="padding:var(--spacing-2)" placeholder="Optional notes...">
                    </div>
                    
                    <button type="submit" class="btn btn-success btn-sm" style="width:100%">
                        <svg style="width:18px;height:18px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/>
                        </svg>
                        Push to Zoho Books
                    </button>
                </form>
                <script>
                (function(){{
                    const form = document.getElementById('zoho-expense-form');
                    if(!form) return;
                    form.addEventListener('submit', async function(ev){{
                        ev.preventDefault();
                        const data = new FormData(form);
                        data.append('ajax','1');
                        const btn = form.querySelector('button[type="submit"]');
                        if(btn){{ btn.disabled = true; btn.innerHTML = '<svg style="width:18px;height:18px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/></svg> Pushing...'; }}
                        try {{
                            const res = await fetch(form.action, {{ method: 'POST', body: data, headers: {{'X-Requested-With':'XMLHttpRequest'}} }});
                            let payload = null;
                            let textBody = '';
                            try {{ payload = await res.clone().json(); }} catch(e) {{ payload = null; }}
                            try {{ textBody = await res.text(); }} catch(e) {{ textBody = ''; }}
                            if (payload && payload.status === 'ok') {{
                                const msg = payload.duplicate ? ('Expense already exists. ID: ' + payload.expense_id)
                                                              : ('Expense created successfully. ID: ' + payload.expense_id);
                                alert(msg);
                            }} else {{
                                const fallback = textBody || (payload ? JSON.stringify(payload) : (res.status + ' ' + res.statusText));
                                alert('Expense push response: ' + fallback);
                            }}
                        }} catch (err) {{
                            alert('Error creating expense: ' + err);
                        }} finally {{
                            if(btn){{ btn.disabled = false; btn.innerHTML = '<svg style="width:18px;height:18px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/></svg> Push to Zoho Books'; }}
                        }}
                    }});
                }})();
                </script>
            </div>
        """

    if 'combined' in reports and 'combined_no_sig' in reports:
        html += f"""
            <div class="card" style="grid-column:1/-1;margin-bottom:0">
                <div class="card-header" style="padding-bottom:var(--spacing-2);margin-bottom:var(--spacing-2)">
                    <h2 class="card-title" style="font-size:var(--font-size-base)">Multi-Tab Combined Reports</h2>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:var(--spacing-3)">
                    <div>
                        <a href="/download/combined" class="btn btn-primary btn-sm" style="width:100%;margin-bottom:var(--spacing-1)">Download (With Signatures)</a>
                        <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Summary + individual sheets with signature lines</p>
                    </div>
                    <div>
                        <a href="/download/combined_no_sig" class="btn btn-primary btn-sm" style="width:100%;margin-bottom:var(--spacing-1)">Download (Without Signatures)</a>
                        <p style="font-size:var(--font-size-xs);color:var(--color-gray-600);margin:0">Summary + individual sheets without signatures</p>
                    </div>
                </div>
            </div>
        """

    if 'error' in reports:
        html += f"""
            <div class="alert alert-danger" style="grid-column:1/-1;padding:var(--spacing-3);margin-bottom:0">
                <svg style="width:18px;height:18px;flex-shrink:0" fill="currentColor" viewBox="0 0 20 20">
                    <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zM8.707 7.293a1 1 0 00-1.414 1.414L8.586 10l-1.293 1.293a1 1 0 101.414 1.414L10 11.414l1.293 1.293a1 1 0 001.414-1.414L11.414 10l1.293-1.293a1 1 0 00-1.414-1.414L10 8.586 8.707 7.293z" clip-rule="evenodd"/>
                </svg>
                <div style="flex:1">
                    <strong style="font-size:var(--font-size-sm)">Error Report</strong>
                    <p style="font-size:var(--font-size-xs);margin-top:var(--spacing-1);margin-bottom:var(--spacing-2)">There was an error processing your file. Check details below.</p>
                    <a href="/static/reports/{reports['error']}" class="btn btn-danger btn-sm">View Error Report</a>
                </div>
            </div>
        """

    html += """
        </div>
        
        <div style="text-align:center;margin-top:var(--spacing-3);padding-top:var(--spacing-3);border-top:1px solid var(--color-gray-200)">
            <a href="/" class="btn btn-primary">
                <svg style="width:18px;height:18px" fill="currentColor" viewBox="0 0 20 20">
                    <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/>
                </svg>
                Process Another File
            </a>
        </div>
    </div>
</body>
</html>
"""
    return html

@app.route('/print/<report_type>')
def print_friendly(report_type):
    """Generate a print-friendly version of a report"""
    try:
        # Import datetime at the beginning of the function
        from datetime import datetime, timedelta

        if report_type == 'admin':
            filename = session.get('reports', {}).get('admin', '')
            if not filename:
                return "No admin report found", 404

            # Get week range from session
            week = session.get('week', datetime.now().strftime('%Y-%m-%d'))

            # Get the file path to read the Excel data
            file_path = os.path.join(REPORT_FOLDER, filename)
            if not os.path.exists(file_path):
                return f"File not found: {filename}", 404

            # Extract data from Excel file
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # Prefer date range from the workbook header if present
            date_range_display = session.get('week', datetime.now().strftime('%Y-%m-%d'))
            try:
                header_text = str(ws['A1'].value or '')
                if ' - ' in header_text:
                    # Extract portion after the first ' - '
                    date_range_display = header_text.split(' - ', 1)[1].strip()
            except Exception:
                pass

            # Find the summary section (starting from row 3)
            summary_data = []
            summary_headers = []
            summary_start_col = None
            grand_total_row = None

            # Find where the summary table starts
            for row in range(3, 10):  # Check first few rows
                for col in range(1, 20):  # Check several columns
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value == "Person ID":
                        summary_start_col = col
                        # Extract headers
                        for h_col in range(col, col+5):
                            summary_headers.append(ws.cell(row=row, column=h_col).value)
                        break
                if summary_start_col:
                    break

            # Extract summary data - this contains all employees
            employee_data = {}  # Track all employees by ID
            if summary_start_col:
                current_row = 4
                while True:
                    cell_value = ws.cell(row=current_row, column=summary_start_col).value
                    if cell_value is None or ws.cell(row=current_row, column=summary_start_col+1).value == "GRAND TOTAL":
                        break

                    # Collect data from summary row
                    person_id = cell_value
                    name = ws.cell(row=current_row, column=summary_start_col+1).value
                    hours = ws.cell(row=current_row, column=summary_start_col+2).value
                    pay = ws.cell(row=current_row, column=summary_start_col+3).value
                    rounded_pay = ws.cell(row=current_row, column=summary_start_col+4).value

                    # Store the employee summary data for later use
                    employee_data[str(person_id)] = {
                        'id': person_id,
                        'name': name,
                        'total_hours': hours,
                        'total_pay': pay,
                        'rounded_pay': rounded_pay,
                        'info': {},
                        'days': []
                    }

                    # Add to the summary data list as well
                    row_data = []
                    for col in range(summary_start_col, summary_start_col+5):
                        row_data.append(ws.cell(row=current_row, column=col).value)

                    summary_data.append(row_data)
                    current_row += 1

                # Check for grand total row
                if ws.cell(row=current_row, column=summary_start_col+1).value == "GRAND TOTAL":
                    grand_total_row = []
                    for col in range(summary_start_col, summary_start_col+5):
                        grand_total_row.append(ws.cell(row=current_row, column=col).value)

            # Find all employee breakdown sections by looking for employee names and collecting their details
            # Look through the entire Excel file for employee details

            for row in range(1, ws.max_row):
                # Look for patterns that indicate employee data
                for col in range(1, min(ws.max_column, 20)):  # Check a reasonable number of columns
                    cell_value = ws.cell(row=row, column=col).value

                    # Look for employee ID row indicators like "ID: X | Rate: $Y.ZZ"
                    if cell_value and isinstance(cell_value, str) and "ID:" in cell_value and "Rate:" in cell_value:
                        # Extract the employee ID
                        try:
                            id_part = cell_value.split("|")[0].strip()
                            id_value = id_part.replace("ID:", "").strip()

                            # Look for the employee name in the row above
                            employee_name = ws.cell(row=row-1, column=col).value

                            # If we have this employee in our summary data, add the details
                            if id_value in employee_data:
                                employee_data[id_value]['info']['details'] = cell_value

                                # Look for the data table in the rows below
                                date_row = row + 1

                                # Check if this row contains the date header
                                if ws.cell(row=date_row, column=col).value == "Date":
                                    # Found the date header, start collecting daily entries from the next row
                                    day_row = date_row + 1

                                    # Collect employee daily records
                                    while day_row < min(day_row + 20, ws.max_row):  # Reasonable limit
                                        date_val = ws.cell(row=day_row, column=col).value

                                        # Stop if we hit an empty row or "Total:"
                                        if not date_val:
                                            break

                                        if isinstance(date_val, str) and date_val.startswith("Total:"):
                                            # Found totals row
                                            employee_data[id_value]['totals'] = {
                                                'hours': ws.cell(row=day_row, column=col+3).value,
                                                'pay': ws.cell(row=day_row, column=col+4).value
                                            }
                                            break

                                        # Add day entry
                                        try:
                                            day_data = {
                                                'date': date_val,
                                                'in': ws.cell(row=day_row, column=col+1).value,
                                                'out': ws.cell(row=day_row, column=col+2).value,
                                                'hours': ws.cell(row=day_row, column=col+3).value,
                                                'pay': ws.cell(row=day_row, column=col+4).value
                                            }
                                            employee_data[id_value]['days'].append(day_data)
                                        except Exception as e:
                                            app.logger.error(f"Error processing day row: {e}")

                                        day_row += 1
                        except Exception as e:
                            app.logger.error(f"Error processing employee info: {e}")

            # Generate print-friendly HTML by constructing from the collected employee_data
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Admin Report - {date_range_display}</title>
                <style>
                    @page {{ size: landscape; margin: 0.5cm; }}

                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 5px; font-size: 9pt; line-height: 1.1; }}
                    h1 {{ color: #333; text-align: center; margin: 5px 0; font-size: 12pt; }}
                    h2 {{ color: #333; text-align: center; margin: 5px 0; font-size: 11pt; }}
                    p {{ margin: 2px 0; }}

                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 5px; }}
                    th {{ background-color: #f2f2f2; text-align: left; border-bottom: 1px solid #ddd; padding: 2px; font-size: 9pt; }}
                    td {{ padding: 1px; font-size: 8pt; }}

                    .summary-table {{ width: 100%; margin-bottom: 10px; }}
                    .summary-table th {{ background-color: #f2f2f2; }}

                    .employee-section {{ width: 32%; margin-right: 1.5%; margin-bottom: 10px; float: left; }}
                    .employee-section:nth-child(3n) {{ margin-right: 0; }}
                    .employee-header {{ font-size: 10pt; font-weight: bold; margin-bottom: 3px; }}
                    .employee-info {{ font-size: 8pt; margin-bottom: 3px; }}

                    .employee-table {{ width: 100%; }}
                    .employee-table th {{ font-size: 8pt; padding: 1px; }}
                    .employee-table td {{ font-size: 8pt; padding: 1px; }}

                    .text-right {{ text-align: right; }}
                    .total-row {{ border-top: 1px solid #ddd; font-weight: bold; }}

                    .signature-line {{ margin-top: 5px; padding: 1px 0; }}

                    .print-button {{ position: fixed; top: 10px; right: 10px; padding: 8px 15px; background-color: #2196F3;
                                   color: white; border: none; border-radius: 4px; cursor: pointer; z-index: 999; }}
                    .print-button:hover {{ background-color: #0b7dda; }}

                    .clearfix::after {{
                        content: "";
                        clear: both;
                        display: table;
                    }}

                    /* For print view */
                    @media print {{
                        .print-button {{ display: none; }}
                        .page-break {{ page-break-before: always; }}
                        body {{ zoom: 100%; }}
                    }}
                </style>
                <script>
                    function printReport() {{
                        window.print();
                    }}

                    // Auto print when the page loads
                    window.onload = function() {{
                        // Wait a moment to ensure everything has rendered
                        setTimeout(function() {{
                            window.print();
                        }}, 500);
                    }};
                </script>
            </head>
            <body>
                <button onclick="printReport()" class="print-button">Print Report</button>

                <h1>Payroll Summary - {date_range_display}</h1>

                <!-- Summary Table -->
                <table class="summary-table">
                    <tr>
            """

            # Add summary headers
            for header in summary_headers:
                html += f"<th>{header}</th>"
            html += "</tr>"

            # Add summary data rows
            for row in summary_data:
                html += "<tr>"
                for i, cell in enumerate(row):
                    # Format monetary values
                    if i >= 3 and isinstance(cell, (int, float)):
                        html += f'<td class="text-right">${cell:.2f}</td>'
                    else:
                        html += f"<td>{cell}</td>"
                html += "</tr>"

            # Add grand total row if found
            if grand_total_row:
                html += '<tr class="total-row">'
                for i, cell in enumerate(grand_total_row):
                    if i == 0:
                        html += "<td></td>"
                    elif i == 1:
                        html += f"<td><strong>{cell}</strong></td>"
                    # Format monetary values
                    elif i >= 3 and isinstance(cell, (int, float)):
                        html += f'<td class="text-right"><strong>${cell:.2f}</strong></td>'
                    else:
                        html += f"<td><strong>{cell}</strong></td>"
                html += "</tr>"

            html += """
                </table>

                <h2>Detailed Breakdown by Employee</h2>

                <div class="clearfix">
            """

            # Convert our employee data dictionary to a sorted list to display
            employee_list = list(employee_data.values())

            # Sort employees by ID
            employee_list.sort(key=lambda x: str(x['id']))

            # Check if we have any breakdown sections
            if not employee_list:
                html += "<p>No detailed breakdown data found.</p>"
            else:
                # Number of employees per page - 9 on first page (3 rows of 3), rest on second page
                employees_per_page = 9

                # Add employee sections, 3 per row
                for i, emp in enumerate(employee_list):
                    # Add page break after the first 9 employees
                    if i == employees_per_page:
                        html += '</div><div class="page-break"></div><div class="clearfix">'

                    # Add a new row every 3 employees
                    if i > 0 and i % 3 == 0:
                        html += '</div><div class="clearfix">'

                    # Format total values
                    total_hours = emp.get('total_hours', 0)
                    total_pay = emp.get('total_pay', 0)
                    rounded_pay = emp.get('rounded_pay', 0)

                    # Make sure total hours and pay are numeric
                    if not isinstance(total_hours, (int, float)):
                        try:
                            total_hours = float(total_hours)
                        except:
                            total_hours = 0

                    if not isinstance(total_pay, (int, float)):
                        try:
                            total_pay = float(total_pay)
                        except:
                            total_pay = 0

                    if not isinstance(rounded_pay, (int, float)):
                        try:
                            rounded_pay = float(rounded_pay)
                        except:
                            rounded_pay = 0

                    html += f"""
                    <div class="employee-section">
                        <div class="employee-header">{emp['name']}</div>
                        <div class="employee-info">{emp['info'].get('details', f"ID: {emp['id']}")}</div>

                        <table class="employee-table">
                            <tr>
                                <th>Date</th>
                                <th>In</th>
                                <th>Out</th>
                                <th>Hours</th>
                                <th>Pay</th>
                            </tr>
                    """

                    # Add daily entries
                    for day in emp.get('days', []):
                        date_str = day['date']
                        pay_val = day['pay']
                        pay_str = f"${pay_val:.2f}" if isinstance(pay_val, (int, float)) else pay_val

                        html += f"""
                            <tr>
                                <td>{date_str}</td>
                                <td>{day['in'] or ''}</td>
                                <td>{day['out'] or ''}</td>
                                <td>{day['hours'] or ''}</td>
                                <td class="text-right">{pay_str}</td>
                            </tr>
                        """

                    # If no days data but we have total hours/pay, show empty rows with zeros
                    if not emp.get('days') and total_hours > 0:
                        # Format for a typical work week (Mon-Fri)
                        for day_num in range(5):
                            html += f"""
                                <tr>
                                    <td>{'4/' + str(26 + day_num) + '/2025'}</td>
                                    <td>None</td>
                                    <td>None</td>
                                    <td>0</td>
                                    <td class="text-right">$0.00</td>
                                </tr>
                            """

                    # Add total row
                    html += f"""
                        <tr>
                            <td colspan="4">Total:</td>
                            <td class="text-right">${total_pay:.2f}</td>
                        </tr>
                    """

                    # Add rounded pay
                    html += f"""
                        <tr>
                            <td colspan="4">Rounded Pay:</td>
                            <td class="text-right">${rounded_pay:.2f}</td>
                        </tr>
                    </table>

                    <div class="signature-line">
                        <span>Signature:_________</span>
                        <span style="float:right">Date:_________</span>
                    </div>
                    </div>
                    """

            html += """
                </div>
            </body>
            </html>
            
                </div>
            </main>
        </div>
    </div>
</body>
</html>
    """

            return html

        elif report_type == 'payslips':
            filename = session.get('reports', {}).get('payslips_sheet', '')
            if not filename:
                return "No payslips report found", 404

            # Get week range from session
            week = session.get('week', datetime.now().strftime('%Y-%m-%d'))

            # Get the file path to read the Excel data
            file_path = os.path.join(REPORT_FOLDER, filename)
            if not os.path.exists(file_path):
                return f"File not found: {filename}", 404

            # Extract data from Excel file
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # Prefer date range from the workbook header if present
            date_range_display = session.get('week', datetime.now().strftime('%Y-%m-%d'))
            try:
                header_text = str(ws['A1'].value or '')
                if ' - ' in header_text:
                    date_range_display = header_text.split(' - ', 1)[1].strip()
            except Exception:
                pass

            # Find all payslip sections - add debugging
            payslips = []
            found_rows = []

            # Add debug log to help diagnose the issue
            debug_info = f"Excel file: {file_path}, Worksheet dimensions: {ws.max_row}x{ws.max_column}\n"

            # First, scan entire sheet for "Employee:" cells to locate all employee sections
            for row in range(1, ws.max_row):
                for col in range(1, min(ws.max_column + 1, 30)):
                    cell_value = ws.cell(row=row, column=col).value
                    if isinstance(cell_value, str) and "Employee:" in cell_value:
                        found_rows.append((row, col))
                        debug_info += f"Found employee at row {row}, col {col}: {cell_value}\n"

            debug_info += f"Total employee sections found: {len(found_rows)}\n"

            # Now process each found employee section
            for row, col in found_rows:
                cell_value = ws.cell(row=row, column=col).value

                # Found a payslip
                payslip = {
                    'name': cell_value.replace("Employee:", "").strip(),
                    'info': {},
                    'days': []
                }

                # Find the employee ID
                id_found = False
                # Look for ID in the next row or nearby cells
                for r in range(row, min(row+3, ws.max_row)):
                    for c in range(max(1, col-2), min(col+6, ws.max_column + 1)):
                        cell_text = ws.cell(row=r, column=c).value
                        if cell_text and isinstance(cell_text, str) and "ID:" in cell_text:
                            payslip['info']['id'] = cell_text.replace("ID:", "").strip()
                            id_found = True
                            break
                    if id_found:
                        break

                # Find pay period
                period_found = False
                for r in range(row, min(row+3, ws.max_row)):
                    for c in range(max(1, col-2), min(col+6, ws.max_column + 1)):
                        cell_text = ws.cell(row=r, column=c).value
                        if cell_text and isinstance(cell_text, str) and "Pay Period:" in cell_text:
                            payslip['period'] = cell_text.replace("Pay Period:", "").strip()
                            period_found = True
                            break
                    if period_found:
                        break

                # Find hourly rate
                rate_found = False
                for r in range(row, min(row+3, ws.max_row)):
                    for c in range(max(1, col-2), min(col+6, ws.max_column + 1)):
                        cell_text = ws.cell(row=r, column=c).value
                        if cell_text and isinstance(cell_text, str) and "Rate:" in cell_text:
                            payslip['info']['rate'] = cell_text
                            rate_found = True
                            break
                    if rate_found:
                        break

                # Find the "Date" header to locate the timesheet area
                date_header_row = None
                date_header_col = None
                for r in range(row, min(row+5, ws.max_row)):
                    for c in range(max(1, col-2), min(col+6, ws.max_column + 1)):
                        header_val = ws.cell(row=r, column=c).value
                        if header_val == "Date":
                            date_header_row = r
                            date_header_col = c
                            break
                    if date_header_row:
                        break

                if not date_header_row:
                    debug_info += f"Could not find Date header for employee at row {row}, col {col}\n"
                    continue  # Skip if we can't find the date header

                debug_info += f"Found Date header at row {date_header_row}, col {date_header_col} for employee at row {row}\n"

                # Find the column headers to map the structure
                headers = []
                header_cols = {}

                for c in range(max(1, date_header_col-1), min(date_header_col+6, ws.max_column + 1)):
                    header_text = ws.cell(row=date_header_row, column=c).value
                    if header_text:
                        headers.append(header_text)
                        header_cols[header_text] = c
                        debug_info += f"Found header '{header_text}' at column {c}\n"

                # Get daily entries - start from the row after headers
                day_row = date_header_row + 1
                total_hours = None
                total_pay = None
                rounded_pay = None

                days_found = 0
                while day_row < min(day_row + 20, ws.max_row):  # Increased limit
                    date_val = ws.cell(row=day_row, column=date_header_col).value

                    # Exit if empty row and we've processed some data already
                    if not date_val and days_found > 0 and day_row > date_header_row + 6:
                        break

                    # Check for totals row
                    if isinstance(date_val, str) and ("Total" in date_val or date_val.strip() == "Total:"):
                        # Found hours or pay totals
                        debug_info += f"Found totals row at {day_row}: {date_val}\n"
                        if "Hours" in header_cols:
                            total_hours = ws.cell(row=day_row, column=header_cols["Hours"]).value
                        if "Pay" in header_cols:
                            total_pay = ws.cell(row=day_row, column=header_cols["Pay"]).value
                        day_row += 1
                        continue

                    # Check for Rounded Pay row
                    if isinstance(date_val, str) and "Rounded" in date_val:
                        debug_info += f"Found rounded pay at {day_row}: {date_val}\n"
                        if "Pay" in header_cols:
                            rounded_pay = ws.cell(row=day_row, column=header_cols["Pay"]).value
                        day_row += 1
                        continue

                    # Only process rows with actual dates (not empty or text)
                    if date_val and not isinstance(date_val, str):
                        try:
                            # Convert Excel serial date if needed
                            if isinstance(date_val, (int, float)):
                                # Use the already imported datetime and timedelta from function scope
                                date_val = (datetime(1899, 12, 30) + timedelta(days=date_val)).strftime('%m/%d/%Y')
                        except:
                            pass

                    try:
                        # Only add if it has a date value and hours > 0
                        hours_val = None
                        if "Hours" in header_cols:
                            hours_val = ws.cell(row=day_row, column=header_cols["Hours"]).value

                        pay_val = None
                        if "Pay" in header_cols:
                            pay_val = ws.cell(row=day_row, column=header_cols["Pay"]).value

                        # Add valid entries with dates and hours
                        if date_val and hours_val and float(hours_val) > 0:
                            in_val = None
                            out_val = None

                            if "In" in header_cols:
                                in_val = ws.cell(row=day_row, column=header_cols["In"]).value

                            if "Out" in header_cols:
                                out_val = ws.cell(row=day_row, column=header_cols["Out"]).value

                            day_data = {
                                'date': date_val,
                                'in': in_val,
                                'out': out_val,
                                'hours': hours_val,
                                'pay': pay_val
                            }
                            payslip['days'].append(day_data)
                            days_found += 1
                            debug_info += f"Added day entry at row {day_row}: date={date_val}, hours={hours_val}\n"
                    except Exception as e:
                        debug_info += f"Error processing row {day_row}: {str(e)}\n"

                    day_row += 1

                # Add totals to the payslip object
                if total_hours is not None:
                    payslip['total_hours'] = total_hours
                elif payslip['days']:
                    # Calculate from individual days if not found
                    total_hours = sum(float(day['hours']) for day in payslip['days'] if day['hours'])
                    payslip['total_hours'] = round(total_hours, 2)

                if total_pay is not None:
                    payslip['total_pay'] = total_pay
                elif payslip['days']:
                    # Calculate from individual days if not found
                    total_pay = sum(float(day['pay']) for day in payslip['days'] if day['pay'])
                    payslip['total_pay'] = round(total_pay, 2)

                if rounded_pay is not None:
                    payslip['rounded_pay'] = rounded_pay
                elif total_pay is not None:
                    # Round to nearest dollar if not found
                    payslip['rounded_pay'] = round(float(total_pay))

                # Only add payslips with actual day entries
                if len(payslip['days']) > 0:
                    payslips.append(payslip)
                    debug_info += f"Added payslip for {payslip['name']} with {len(payslip['days'])} days\n"
                else:
                    debug_info += f"Skipped payslip for {payslip['name']} because it has no days\n"

            # If no payslips found, return debug info
            if not payslips:
                return f"No payslips found in the Excel file. Debug info:<br><pre>{debug_info}</pre>", 400

            # Generate print-friendly HTML
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Employee Payslips - {date_range_display}</title>
                <style>
                    @page {{ size: portrait; margin: 0.5cm; scale: 78%; }}

                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 0; font-size: 8pt; line-height: 1.2; }}
                    h1 {{ color: #333; text-align: center; margin: 5px 0; font-size: 12pt; }}
                    p {{ margin: 2px 0; }}

                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 5px; }}
                    th {{ background-color: #f2f2f2; text-align: left; border-bottom: 1px solid #ddd; padding: 2px; font-size: 8pt; }}
                    td {{ padding: 2px; font-size: 8pt; }}

                    .payslip {{ break-inside: avoid; page-break-inside: avoid; border: 1px solid #ddd;
                               border-radius: 3px; padding: 8px; margin-bottom: 10px; width: 100%; box-sizing: border-box; }}
                    .payslip-header {{ border-bottom: 1px solid #eee; padding-bottom: 4px; margin-bottom: 4px; }}
                    .employee-name {{ font-size: 10pt; font-weight: bold; }}
                    .employee-id {{ text-align: right; font-weight: normal; font-size: 8pt; }}

                    .info-row {{ display: flex; justify-content: space-between; margin: 2px 0; }}
                    .period {{ font-style: italic; color: #666; font-size: 7pt; }}
                    .total-row {{ border-top: 1px solid #ddd; font-weight: bold; }}
                    .text-right {{ text-align: right; }}

                    .signature-line {{ margin-top: 5px; padding: 2px 0; display: flex; justify-content: space-between; font-size: 7pt; }}

                    .print-button {{ position: fixed; top: 10px; right: 10px; padding: 8px 15px; background-color: #2196F3;
                                   color: white; border: none; border-radius: 4px; cursor: pointer; z-index: 999; }}
                    .print-button:hover {{ background-color: #0b7dda; }}

                    .debug-button {{ position: fixed; top: 10px; left: 10px; padding: 8px 15px; background-color: #FF5722;
                                   color: white; border: none; border-radius: 4px; cursor: pointer; z-index: 999; }}
                    .debug-info {{ display: none; background: #f5f5f5; border: 1px solid #ddd; padding: 15px;
                                 margin: 20px 0; font-family: monospace; font-size: 12px; white-space: pre-wrap; }}

                    .payslips-container {{ display: flex; flex-wrap: wrap; justify-content: space-between; align-items: flex-start; }}

                    /* For print view, optimize to fit multiple per page */
                    @media print {{
                        .print-button, .debug-button, .debug-info {{ display: none; }}
                        .page-break {{ page-break-before: always; }}

                        /* Compact layout for print */
                        body {{ font-size: 7pt; line-height: 1.1; }}

                        /* Two columns of payslips */
                        .payslip {{
                            width: 48%;
                            display: inline-block;
                            vertical-align: top;
                            margin-bottom: 0.2cm;
                            padding: 0.4cm;
                            border: 1px solid #ccc;
                        }}

                        table {{ margin-bottom: 2px; }}
                        td, th {{ padding: 1px; font-size: 7pt; }}
                        .signature-line {{ margin-top: 5px; font-size: 6pt; }}
                        .employee-name {{ font-size: 9pt; }}
                    }}
                </style>
                <script>
                    function printReport() {{
                        window.print();
                    }}

                    function toggleDebug() {{
                        var debugInfo = document.getElementById('debug-info');
                        if (debugInfo.style.display === 'none' || !debugInfo.style.display) {{
                            debugInfo.style.display = 'block';
                        }} else {{
                            debugInfo.style.display = 'none';
                        }}
                    }}

                    // Set print scale to 78%
                    function setPrintScale() {{
                        // For Chrome and Safari
                        if (window.matchMedia) {{
                            const mediaQueryList = window.matchMedia('print');
                            mediaQueryList.addListener(function(mql) {{
                                if (mql.matches) {{
                                    document.body.style.zoom = "78%";
                                }}
                            }});
                        }}
                    }}

                    // Auto print when the page loads
                    window.onload = function() {{
                        // Set print scale
                        setPrintScale();

                        // Wait a moment to ensure everything has rendered
                        setTimeout(function() {{
                            window.print();
                        }}, 500);
                    }};
                </script>
            </head>
            <body>
                <button onclick="printReport()" class="print-button">Print Payslips</button>
                <button onclick="toggleDebug()" class="debug-button">Toggle Debug Info</button>

                <div id="debug-info" class="debug-info">
                    <h3>Debug Information</h3>
                    <p>Found {len(payslips)} payslips</p>
                    <pre>{debug_info}</pre>
                </div>

                <h1>Employee Payslips - {date_range_display}</h1>
                <p style="text-align: center; margin: 0 0 5px 0;">Found {len(payslips)} employee payslips</p>


                <div class="payslips-container">
            """

            # Add each payslip
            for i, payslip in enumerate(payslips):
                emp_id = payslip['info'].get('id', '')

                # Remove page breaks to fit all on one page
                # if i > 0 and i % 3 == 0:
                #     html += '<div class="page-break"></div>'

                html += f"""
                <div class="payslip">
                    <div class="payslip-header">
                        <div class="employee-name">
                            {payslip['name']}
                            <span class="employee-id">ID: {emp_id}</span>
                        </div>
                        <div class="info-row">
                            <div class="period">Pay Period: {payslip.get('period', week)}</div>
                            <div>{payslip['info'].get('rate', '')}</div>
                        </div>
                    </div>

                    <table>
                        <tr>
                            <th>Date</th>
                            <th>In</th>
                            <th>Out</th>
                            <th>Hours</th>
                            <th>Pay</th>
                        </tr>
                """

                # Add daily entries
                for day in payslip['days']:
                    date_str = day['date']
                    pay_val = day['pay']

                    # Format pay as currency
                    pay_str = ""
                    if pay_val is not None:
                        if isinstance(pay_val, (int, float)):
                            pay_str = f"${pay_val:.2f}"
                        else:
                            pay_str = str(pay_val)

                    # Format hours to show 2 decimal places
                    hours_val = day['hours']
                    hours_str = ""
                    if hours_val is not None:
                        if isinstance(hours_val, (int, float)):
                            hours_str = f"{hours_val:.2f}"
                        else:
                            hours_str = str(hours_val)

                    html += f"""
                        <tr>
                            <td>{date_str}</td>
                            <td>{day['in'] or ''}</td>
                            <td>{day['out'] or ''}</td>
                            <td>{hours_str}</td>
                            <td class="text-right">{pay_str}</td>
                        </tr>
                    """

                # Format totals with proper currency and decimals
                total_hours = payslip.get('total_hours', '')
                if isinstance(total_hours, (int, float)):
                    total_hours = f"{total_hours:.2f}"

                total_pay = payslip.get('total_pay', '')
                total_pay_str = ""
                if isinstance(total_pay, (int, float)):
                    total_pay_str = f"${total_pay:.2f}"
                else:
                    total_pay_str = f"${total_pay}" if total_pay else "$0.00"

                rounded_pay = payslip.get('rounded_pay', '')
                rounded_pay_str = ""
                if isinstance(rounded_pay, (int, float)):
                    rounded_pay_str = f"${rounded_pay:.2f}"
                else:
                    rounded_pay_str = f"${rounded_pay}" if rounded_pay else "$0.00"

                # Add totals
                html += f"""
                    <tr class="total-row">
                        <td colspan="3">Total Hours:</td>
                        <td>{total_hours}</td>
                        <td class="text-right">{total_pay_str}</td>
                    </tr>
                    <tr>
                        <td colspan="3">Rounded Pay:</td>
                        <td></td>
                        <td class="text-right">{rounded_pay_str}</td>
                    </tr>
                    </table>

                    <!-- Removed signature line -->
                </div>
                """

            html += """
            </div>
            </body>
            </html>
            """

            return html
        else:
            return "Invalid report type for printing", 400
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"Error generating print-friendly version: {str(e)}<br><pre>{error_details}</pre>", 500

@app.route('/download/<report_type>')
def download(report_type):
    """Download a report file"""
    try:
        if report_type == 'summary':
            filename = session.get('reports', {}).get('summary', '')
        elif report_type == 'payslips':
            filename = session.get('reports', {}).get('payslips', '')
        elif report_type == 'combined':
            filename = session.get('reports', {}).get('combined', '')
        elif report_type == 'combined_no_sig':
            filename = session.get('reports', {}).get('combined_no_sig', '')
        elif report_type == 'admin':
            filename = session.get('reports', {}).get('admin', '')
        elif report_type == 'payslips_sheet':
            filename = session.get('reports', {}).get('payslips_sheet', '')
        else:
            return "Invalid report type", 400

        if not filename:
            return "No report found", 404

        file_path = os.path.join(REPORT_FOLDER, filename)
        if not os.path.exists(file_path):
            return f"File not found: {filename}", 404

        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return f"Error downloading file: {str(e)}", 500

@app.route('/download_pdf/<filename>')
def download_pdf(filename):
    """Download an Excel report as PDF"""
    try:
        # Ensure filename ends with .xlsx
        if not filename.endswith('.xlsx'):
            return "Invalid file type", 400
        
        # Get the Excel file path
        file_path = os.path.join(REPORT_FOLDER, filename)
        if not os.path.exists(file_path):
            return f"File not found: {filename}", 404
        
        # Convert to PDF
        pdf_buffer = convert_excel_to_pdf(file_path)
        
        # Generate PDF filename
        pdf_filename = filename.replace('.xlsx', '.pdf')
        
        # Send PDF file
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=pdf_filename,
            mimetype='application/pdf'
        )
    except Exception as e:
        app.logger.error(f"Error generating PDF: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return f"Error generating PDF: {str(e)}", 500


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTS & DOWNLOADS
# ═══════════════════════════════════════════════════════════════════════════════
# Report listing, viewing, and download functionality

@app.route('/reports')
@login_required
def reports():
    """Display all generated reports grouped by week"""
    # Get username for menu display
    username = session.get('username', 'Unknown')
    menu_html = get_menu_html(username)

    # Check if we have cached report data that's still valid (less than 5 minutes old)
    current_time = datetime.now()
    cache_key = 'all_reports'
    if cache_key in report_cache and report_cache_expiry.get(cache_key, datetime.min) > current_time:
        # Use cached data
        sorted_weeks = report_cache[cache_key]['sorted_weeks']
        reports_by_week = report_cache[cache_key]['reports_by_week']
    else:
        # Existing code for reports preparation - but optimized
        report_files = []

        # Files to exclude from reports display
        excluded_files = [
            'error_report.txt',
            'pandas_report.csv',
            'file_report.txt'  # Also exclude this system file
        ]

        # Get all files in the reports directory - newest first and limited
        try:
            entries = []
            for f in os.listdir(REPORT_FOLDER):
                if not (f.startswith('admin_report_') and f.endswith('.xlsx')):
                    continue
                fp = os.path.join(REPORT_FOLDER, f)
                if not os.path.isfile(fp) or f in excluded_files:
                    continue
                entries.append((f, os.path.getmtime(fp)))
            entries.sort(key=lambda x: x[1], reverse=True)
            all_files = [f for f, _ in entries[:REPORTS_LIST_LIMIT]]
        except Exception:
            all_files = []

        for filename in all_files:
            # Extract week from filename
            week_match = re.search(r'_(\d{4}-\d{2}-\d{2})\.', filename)
            week = week_match.group(1) if week_match else "Unknown"

            # Get file path
            file_path = os.path.join(REPORT_FOLDER, filename)

            # Determine report type and extract title from file if possible (use metadata cache for heavy fields)
            report_type = "Admin Report"
            report_title = None
            total_amount = None
            creator = None

            # Try to extract title and total amount from Excel files - now cached
            if filename.endswith('.xlsx'):
                try:
                    meta = _load_reports_metadata()
                    rec = _ensure_report_metadata(file_path, filename, meta)
                    _save_reports_metadata(meta)
                    creator = rec.get('creator') or 'Unknown'
                    total_amount = rec.get('total_amount')
                    date_range = rec.get('date_range')

                    # Use the cached date range if available (this is the actual payroll period)
                    if date_range:
                        week = date_range


                    # Title is cheap to read, so we still grab A1
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    if ws['A1'].value:
                        report_title = ws['A1'].value
                except Exception as e:
                    # If we can't read the Excel file, just continue without detailed info
                    pass

            # Get file creation time
            creation_time = datetime.fromtimestamp(os.path.getctime(file_path))
            
            # Extract end date from date range for proper sorting
            sort_date = creation_time  # Default to creation time
            try:
                if " to " in week:
                    # Extract end date from range like "2025-01-04 to 2025-01-10"
                    date_parts = week.split(" to ")
                    if len(date_parts) == 2:
                        sort_date = datetime.strptime(date_parts[1].strip(), "%Y-%m-%d")
                elif week != "Unknown":
                    # Single date format
                    try:
                        sort_date = datetime.strptime(week, "%Y-%m-%d")
                    except:
                        pass
            except Exception as e:
                app.logger.warning(f"Could not parse date from week '{week}': {e}")
                # Fall back to creation_time

            report_files.append({
                'filename': filename,
                'week': week,
                'type': report_type,
                'title': report_title,
                'total_amount': total_amount,
                'creator': creator,
                'created': creation_time,
                'sort_date': sort_date,  # Use this for sorting
                'size': os.path.getsize(file_path)
            })

        # Sort by actual payroll period end date (newest first)
        report_files.sort(key=lambda x: x['sort_date'], reverse=True)

        # Group by week
        reports_by_week = {}
        for report in report_files:
            if report['week'] not in reports_by_week:
                reports_by_week[report['week']] = []
            reports_by_week[report['week']].append(report)

        # Sort weeks chronologically by extracting end date (newest first)
        def extract_sort_date_from_week(week_str):
            """Extract end date from week string for sorting"""
            try:
                if " to " in week_str:
                    # Extract end date from range
                    date_parts = week_str.split(" to ")
                    if len(date_parts) == 2:
                        return datetime.strptime(date_parts[1].strip(), "%Y-%m-%d")
                elif week_str != "Unknown":
                    # Single date
                    try:
                        return datetime.strptime(week_str, "%Y-%m-%d")
                    except:
                        pass
            except:
                pass
            # Return minimum date for unknown/unparseable weeks (sorts to bottom)
            return datetime.min
        
        sorted_weeks = sorted(reports_by_week.keys(), 
                            key=extract_sort_date_from_week, 
                            reverse=True)

        # Cache the results for 5 minutes
        report_cache[cache_key] = {
            'sorted_weeks': sorted_weeks,
            'reports_by_week': reports_by_week
        }
        report_cache_expiry[cache_key] = current_time + timedelta(minutes=5)

    # Prepare flash messages
    flashes = get_flashed_messages(with_categories=True)
    flash_html = ''
    if flashes:
        for category, message in flashes:
            alert_class = 'alert-info'
            if category == 'success':
                alert_class = 'alert-success'
            elif category == 'error':
                alert_class = 'alert-danger'
            elif category == 'warning':
                alert_class = 'alert-warning'
            flash_html += f'<div class="alert {alert_class}">{escape(message)}</div>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Reports - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .reports-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
        .empty-state {{
            text-align: center;
            padding: var(--spacing-4);
        }}
        .empty-state-icon {{
            width: 48px;
            height: 48px;
            margin: 0 auto var(--spacing-2);
            color: var(--color-gray-400);
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="reports-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-2)">Payroll Reports</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-lg);margin:0">View and download all generated payroll reports</p>
        </div>
    </div>
    
    <div class="container">
        {flash_html}
        
    """

    if not reports_by_week:
        html += """
        <div class="card">
            <div class="empty-state">
                <svg class="empty-state-icon" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z" />
                </svg>
                <h3 style="font-size:var(--font-size-xl);font-weight:var(--font-weight-semibold);color:var(--color-gray-900);margin-bottom:var(--spacing-2)">No Reports Found</h3>
                <p style="color:var(--color-gray-600);margin-bottom:var(--spacing-3)">No payroll reports have been generated yet. Process a timesheet to create reports.</p>
                <a href="/" class="btn btn-primary">Process Payroll</a>
            </div>
        </div>
        """
    else:
        # Render a professional table with all reports
        html += """
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">
                    <svg style="width:24px;height:24px;display:inline;margin-right:8px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M4 4a2 2 0 012-2h4.586A2 2 0 0112 2.586L15.414 6A2 2 0 0116 7.414V16a2 2 0 01-2 2H6a2 2 0 01-2-2V4zm2 6a1 1 0 011-1h6a1 1 0 110 2H7a1 1 0 01-1-1zm1 3a1 1 0 100 2h6a1 1 0 100-2H7z" clip-rule="evenodd"/>
                    </svg>
                    Generated Reports
                </h2>
            </div>
            
            <div class="table-wrapper">
                <table class="table">
                    <thead>
                        <tr>
                            <th>Week</th>
                            <th class="text-right">Amount</th>
                            <th>Created By</th>
                            <th>Posting Date</th>
                            <th class="text-right">Actions</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        for week in sorted_weeks:
            # Format the week date for display as a range and compute posting date (end-of-week + 1)
            try:
                # Check if week is a date range (e.g., "2025-01-04 to 2025-01-10")
                if " to " in week:
                    date_range_match = re.search(r'(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})', week)
                    if date_range_match:
                        start_date = datetime.strptime(date_range_match.group(1), "%Y-%m-%d")
                        end_date = datetime.strptime(date_range_match.group(2), "%Y-%m-%d")
                        week_display = f"{start_date.strftime('%b %d')} – {end_date.strftime('%b %d, %Y')}"
                        posting_date_display = (end_date + timedelta(days=1)).strftime('%b %d, %Y')
                    else:
                        week_display = week
                        posting_date_display = ''
                else:
                    # Fallback: single date format
                    week_date = datetime.strptime(week, "%Y-%m-%d")
                    week_end = week_date + timedelta(days=6)
                    week_display = f"{week_date.strftime('%b %d')} – {week_end.strftime('%b %d, %Y')}"
                    posting_date_display = (week_end + timedelta(days=1)).strftime('%b %d, %Y')
            except:
                week_display = week
                posting_date_display = ''
            # Pick one entry per week (prefer Admin Report) for concise display
            entries = reports_by_week.get(week, [])
            admin_entry = None
            for e in entries:
                if (e.get('type') or '').lower().startswith('admin'):
                    admin_entry = e
                    break
            if not admin_entry and entries:
                admin_entry = entries[0]

            amount_str = 'N/A'
            creator_str = 'Unknown'
            download_filename = ''
            if admin_entry:
                if admin_entry.get('total_amount') is not None:
                    amount_str = f"${admin_entry['total_amount']:.2f}"
                creator_str = admin_entry.get('creator') or 'Unknown'
                download_filename = admin_entry.get('filename') or ''

            html += f"""
                        <tr>
                            <td><strong>{week_display}</strong></td>
                            <td class="text-right"><span style="color:var(--color-success);font-weight:var(--font-weight-semibold)">{amount_str}</span></td>
                            <td><span class="badge badge-primary">{escape(creator_str)}</span></td>
                            <td>{posting_date_display}</td>
                            <td class="text-right">
            """
            
            if download_filename:
                html += f'<a href="/download_pdf/{download_filename}" class="btn btn-primary btn-sm"><svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M3 17a1 1 0 011-1h12a1 1 0 110 2H4a1 1 0 01-1-1zm3.293-7.707a1 1 0 011.414 0L9 10.586V3a1 1 0 112 0v7.586l1.293-1.293a1 1 0 111.414 1.414l-3 3a1 1 0 01-1.414 0l-3-3a1 1 0 010-1.414z" clip-rule="evenodd"/></svg> Download PDF</a>'
            else:
                html += '<span style="color:var(--color-gray-500);font-size:var(--font-size-sm)">N/A</span>'
            
            html += """
                            </td>
                        </tr>
            """

        html += """
                    </tbody>
                </table>
            </div>
        </div>
        """

    html += """
    </div>
</body>
</html>
    """

    return html

# Add a cache clearing function for when new reports are generated
def clear_report_cache():
    """Clear the report cache to ensure fresh data is shown after new reports are created"""
    global report_cache, report_cache_expiry
    report_cache = {}
    report_cache_expiry = {}

# ========== ZOHO BOOKS EXPENSE ACTION ==========
def _auto_push_expense_if_configured(week):
    """Optionally auto-create expense in Zoho Books if env is configured."""
    try:
        auto_flag = os.getenv('ZB_AUTO_PUSH_EXPENSE', 'false').lower() in ('1', 'true', 'yes')
        default_company = os.getenv('ZB_DEFAULT_COMPANY', '').strip()
        if not (auto_flag and default_company):
            return
        reports = session.get('reports', {})
        if 'admin' not in reports:
            return
        admin_file = os.path.join(REPORT_FOLDER, reports['admin'])
        # Determine amount from CSV
        uploaded_file = session.get('uploaded_file')
        amount = None
        try:
            from openpyxl import load_workbook
            if os.path.exists(admin_file):
                wb = load_workbook(admin_file, data_only=True, read_only=True)
                ws = wb.active
                max_rows = min(ws.max_row, 40)
                for r in range(3, max_rows + 1):
                    row_text = ''.join([str(ws.cell(row=r, column=c).value or '') for c in range(1, min(ws.max_column, 20))])
                    if 'GRAND TOTAL' in row_text.upper():
                        for c in range(min(ws.max_column, 20), 1, -1):
                            cell_val = ws.cell(row=r, column=c).value
                            if isinstance(cell_val, (int, float)) and cell_val > 0:
                                amount = float(cell_val)
                                break
                        if amount is not None:
                            break
        except Exception:
            amount = None
        if amount is None and uploaded_file and os.path.exists(uploaded_file):
            df = pd.read_csv(uploaded_file)
            if all(col in df.columns for col in ['Person ID', 'First Name', 'Last Name', 'Date']):
                _, total_pay, _ = compute_grand_totals_for_expense(df)
                amount = round(total_pay, 2)
        if amount is None:
            return
        # Prevent duplicate creation for the same company+week
        # First check session cache for quick lookup
        existing = _get_existing_expense(default_company, week)
        if existing:
            # If previously stored, ensure it still exists; if not, allow recreation
            if zoho_get_expense(default_company, existing):
                app.logger.info(f"Duplicate prevented (session cache): expense {existing} already exists")
                return
            _clear_existing_expense(default_company, week)
        
        # Second check: Search Zoho directly by reference number for robust duplicate prevention
        start_str, end_str = compute_week_range_strings(week)
        reference_number = f"PAYROLL-{start_str}_to_{end_str}"
        
        # Check if expense with this reference number already exists in Zoho
        existing_by_ref = zoho_find_expense_by_reference(default_company, reference_number)
        if existing_by_ref:
            app.logger.warning(f"Duplicate prevented (Zoho search): expense {existing_by_ref} with reference {reference_number} already exists")
            # Store in session to avoid future API calls
            _set_existing_expense(default_company, week, existing_by_ref)
            return
        post_date = compute_expense_date_from_data(week)
        # Auto-notes from summary
        csv_path = session.get('filtered_file') or session.get('uploaded_file')
        auto_notes = build_admin_summary_text_from_csv(csv_path, start_str, end_str)
        base_desc = f"Weekly payroll expense for {start_str} to {end_str} created by {session.get('username', 'Unknown')}"
        description = _compose_zoho_description(base_desc, auto_notes, '')
        expense_id = zoho_create_expense(
            default_company,
            date_str=post_date,
            amount=amount,
            description=description,
            reference_number=reference_number
        )
        _set_existing_expense(default_company, week, expense_id)
        if os.path.exists(admin_file):
            try:
                zoho_attach_receipt(default_company, expense_id, admin_file)
            except Exception:
                pass
    except Exception:
        # Silent fail; visible UX still shows downloads
        pass
@app.route('/zoho/create_expense', methods=['POST'])
@login_required
def zoho_create_expense_route():
    """Create a Zoho Books expense for the last run and attach the Admin report."""
    try:
        company = request.form.get('company', 'haute')
        week = request.form.get('week', datetime.now().strftime('%Y-%m-%d'))

        # Ensure reports exist in session
        reports = session.get('reports', {})
        if 'admin' not in reports:
            return "Admin report not found in session. Please process payroll first.", 400

        # Prevent duplicate creation for same company+week
        # First check session cache
        existing = _get_existing_expense(company, week)
        if existing:
            # Validate that the expense still exists in Zoho; if deleted, clear cache and proceed
            exists_remote = zoho_get_expense(company, existing)
            if not exists_remote:
                _clear_existing_expense(company, week)
            else:
                app.logger.info(f"Duplicate prevented (session cache): expense {existing} already exists")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('ajax') == '1':
                    return jsonify({'status': 'ok', 'expense_id': existing, 'duplicate': True})
                return f"<script>alert('Expense already exists. ID: {existing}'); history.back();</script>", 200
        
        # Second check: Search Zoho directly by reference number for robust duplicate prevention
        start_str, end_str = compute_week_range_strings(week)
        reference_number = f"PAYROLL-{start_str}_to_{end_str}"
        
        # Check if expense with this reference number already exists in Zoho
        existing_by_ref = zoho_find_expense_by_reference(company, reference_number)
        if existing_by_ref:
            app.logger.warning(f"Duplicate prevented (Zoho search): expense {existing_by_ref} with reference {reference_number} already exists")
            # Store in session to avoid future API calls
            _set_existing_expense(company, week, existing_by_ref)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('ajax') == '1':
                return jsonify({'status': 'ok', 'expense_id': existing_by_ref, 'duplicate': True})
            return f"<script>alert('Expense already exists in Zoho. ID: {existing_by_ref}\\nReference: {reference_number}'); history.back();</script>", 200

        # Compute total amount from dataframe stored during processing is not persisted.
        # Re-open the admin report to extract the grand total if possible; otherwise recompute from uploaded CSV.
        amount = None
        try:
            from openpyxl import load_workbook
            admin_path = os.path.join(REPORT_FOLDER, reports['admin'])
            if os.path.exists(admin_path):
                wb = load_workbook(admin_path, data_only=True, read_only=True)
                ws = wb.active
                # Try to locate "GRAND TOTAL" row
                found_amount = None
                max_rows = min(ws.max_row, 40)
                for r in range(3, max_rows + 1):
                    val = ws.cell(row=r, column=9).value  # Column I often has labels in our layout, but we search broadly below
                    # Instead of relying on fixed col, scan row text
                    row_text = ''.join([str(ws.cell(row=r, column=c).value or '') for c in range(1, min(ws.max_column, 20))])
                    if 'GRAND TOTAL' in row_text.upper():
                        # Find rightmost numeric on this row
                        for c in range(min(ws.max_column, 20), 1, -1):
                            cell_val = ws.cell(row=r, column=c).value
                            if isinstance(cell_val, (int, float)) and cell_val > 0:
                                found_amount = float(cell_val)
                                break
                        if found_amount is not None:
                            break
                if found_amount is not None:
                    amount = round(found_amount, 2)
        except Exception:
            amount = None

        # If we couldn't extract, derive from the CSV used
        if amount is None:
            uploaded_file = session.get('uploaded_file')
            if not uploaded_file or not os.path.exists(uploaded_file):
                return "Could not determine amount; CSV missing. Re-run processing.", 400
            df = pd.read_csv(uploaded_file)
            # If not a timesheet, abort
            if not all(col in df.columns for col in ['Person ID', 'First Name', 'Last Name', 'Date']):
                return "Uploaded file is not a timesheet. Cannot compute payroll total.", 400
            _, total_pay, _ = compute_grand_totals_for_expense(df)
            amount = round(total_pay, 2)

        # Create expense with posting date = end-of-week + 1
        # Note: start_str, end_str, and reference_number already defined in duplicate check above
        post_date = compute_expense_date_from_data(week)
        # Compose description with automated admin summary + optional notes
        base_desc = f"Weekly payroll expense for {start_str} to {end_str} created by {session.get('username', 'Unknown')}"
        extra_desc = request.form.get('custom_desc', '').strip()
        csv_path = session.get('filtered_file') or session.get('uploaded_file')
        auto_notes = build_admin_summary_text_from_csv(csv_path, start_str, end_str)
        # Build within Zoho's 500-char limit
        final_desc = _compose_zoho_description(base_desc, auto_notes, extra_desc)

        expense_id = zoho_create_expense(
            company,
            date_str=post_date,
            amount=amount,
            description=final_desc,
            reference_number=f"PAYROLL-{start_str}_to_{end_str}",
            paid_through_account_name=None
        )
        _set_existing_expense(company, week, expense_id)

        # Attach admin report as receipt
        admin_file = os.path.join(REPORT_FOLDER, reports['admin'])
        if os.path.exists(admin_file):
            try:
                zoho_attach_receipt(company, expense_id, admin_file)
            except Exception as e:
                # Don't fail entire request if attachment fails; report message instead
                flash(f"Expense created but failed to attach report: {str(e)}", 'warning')

        # Return a small HTML/JS snippet that shows a confirmation dialog and closes the tab/window
        # If request is AJAX, return JSON so the page stays; otherwise show a minimal page with JS alert
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('ajax') == '1':
            return jsonify({'status': 'ok', 'expense_id': expense_id})
        return f"<script>alert('Expense created successfully. ID: {expense_id}'); history.back();</script>", 200
    except Exception as e:
        import traceback
        return f"Error creating Zoho expense: {str(e)}<br><pre>{traceback.format_exc()}</pre>", 500

# Add user management feature
@app.route('/manage_users')
@login_required
def manage_users():
    """Manage system users"""
    username = session.get('username', 'Unknown')
    if username != 'admin':
        return redirect(url_for('index'))
    
    menu_html = get_menu_html(username)
    users = load_users()

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Manage Users - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .users-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="users-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-2)">Manage Users</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-lg);margin:0">Add and remove system users</p>
        </div>
    </div>
    
    <div class="container">
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">
                    <svg style="width:24px;height:24px;display:inline;margin-right:8px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                        <path d="M9 6a3 3 0 11-6 0 3 3 0 016 0zM17 6a3 3 0 11-6 0 3 3 0 016 0zM12.93 17c.046-.327.07-.66.07-1a6.97 6.97 0 00-1.5-4.33A5 5 0 0119 16v1h-6.07zM6 11a5 5 0 015 5v1H1v-1a5 5 0 015-5z"/>
                    </svg>
                    Current Users
                </h2>
            </div>
            
            <div class="table-wrapper">
                <table class="table">
                    <thead>
                        <tr>
                            <th>Username</th>
                            <th class="text-right">Actions</th>
                        </tr>
                    </thead>
                    <tbody>
"""
    
    for user in users.keys():
        is_admin_user = user == 'admin'
        if is_admin_user:
            html += f"""
                        <tr>
                            <td>
                                <strong>{escape(user)}</strong>
                                <span class="badge badge-primary" style="margin-left:var(--spacing-2)">Admin</span>
                            </td>
                            <td class="text-right">
                                <span style="color:var(--color-gray-500);font-size:var(--font-size-sm);font-style:italic">Cannot delete admin</span>
                            </td>
                        </tr>
"""
        else:
            html += f"""
                        <tr>
                            <td><strong>{escape(user)}</strong></td>
                            <td class="text-right">
                                <form method="post" action="/delete_user/{escape(user)}" style="display:inline;" onsubmit="return confirm('Delete user {escape(user)}?');">
                                    <button type="submit" class="btn btn-danger btn-sm">
                                        <svg style="width:16px;height:16px" fill="currentColor" viewBox="0 0 20 20">
                                            <path fill-rule="evenodd" d="M9 2a1 1 0 00-.894.553L7.382 4H4a1 1 0 000 2v10a2 2 0 002 2h8a2 2 0 002-2V6a1 1 0 100-2h-3.382l-.724-1.447A1 1 0 0011 2H9zM7 8a1 1 0 012 0v6a1 1 0 11-2 0V8zm5-1a1 1 0 00-1 1v6a1 1 0 102 0V8a1 1 0 00-1-1z" clip-rule="evenodd"/>
                                        </svg>
                                        Delete
                                    </button>
                                </form>
                            </td>
                        </tr>
"""
    
    html += """
                    </tbody>
                </table>
            </div>
        </div>
        
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">
                    <svg style="width:24px;height:24px;display:inline;margin-right:8px;vertical-align:middle" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M10 3a1 1 0 011 1v5h5a1 1 0 110 2h-5v5a1 1 0 11-2 0v-5H4a1 1 0 110-2h5V4a1 1 0 011-1z" clip-rule="evenodd"/>
                    </svg>
                    Add New User
                </h2>
            </div>
            
            <form method="post" action="/add_user">
                <div class="form-group">
                    <label for="username" class="form-label form-label-required">Username</label>
                    <input type="text" id="username" name="username" class="form-input" placeholder="Enter username" required>
                    <span class="form-help">3-50 characters, alphanumeric only</span>
                </div>
                
                <div class="form-group">
                    <label for="password" class="form-label form-label-required">Password</label>
                    <input type="password" id="password" name="password" class="form-input" placeholder="Enter password" required>
                    <span class="form-help">At least 8 characters, must include letters and numbers</span>
                </div>
                
                <div style="margin-top:var(--spacing-3);text-align:right">
                    <button type="submit" class="btn btn-success">
                        <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M10 3a1 1 0 011 1v5h5a1 1 0 110 2h-5v5a1 1 0 11-2 0v-5H4a1 1 0 110-2h5V4a1 1 0 011-1z" clip-rule="evenodd"/>
                        </svg>
                        Add User
                    </button>
                </div>
            </form>
        </div>
    </div>
</body>
</html>"""
    
    return html

@app.route('/add_user', methods=['POST'])
@login_required
def add_user():
    """Add a new user with input validation"""
    if session.get('username') != 'admin':
        return "Only admin can add users", 403

    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')

    # Validate username
    valid, error = validate_username(username)
    if not valid:
        return error, 400

    # Validate password
    valid, error = validate_password(password)
    if not valid:
        return error, 400

    users = load_users()

    if username in users:
        return "Username already exists", 400

    users[username] = hash_password(password)
    save_users(users)
    app.logger.info(f"New user added: {username}")

    return redirect(url_for('manage_users'))

@app.route('/delete_user', methods=['POST'])
@login_required
def delete_user():
    """Delete a user"""
    if session.get('username') != 'admin':
        return "Only admin can delete users", 403

    username = request.form.get('username')

    if not username:
        return "Username is required", 400

    # Can't delete admin
    if username == 'admin':
        return "Cannot delete admin user", 403

    users = load_users()

    if username not in users:
        return "User not found", 404

    del users[username]
    save_users(users)

    return redirect(url_for('manage_users'))


@app.route('/fetch_timecard', methods=['GET', 'POST'])
@login_required
def fetch_timecard():
    """Fetch timecard data from the NGTeco system"""
    username = session.get('username', 'Unknown')
    menu_html = get_menu_html(username)
    
    if request.method == 'GET':
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Fetch Timecard - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .fetch-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="fetch-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-1);font-size:var(--font-size-2xl)">Fetch Timecard Data from NGTeco</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-sm);margin:0">Automatically download timecard data and convert to CSV</p>
        </div>
    </div>
    
    <div class="container container-narrow">
        <div class="alert alert-info">
            <svg style="width:20px;height:20px;flex-shrink:0" fill="currentColor" viewBox="0 0 20 20">
                <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7-4a1 1 0 11-2 0 1 1 0 012 0zM9 9a1 1 0 000 2v3a1 1 0 001 1h1a1 1 0 100-2v-3a1 1 0 00-1-1H9z" clip-rule="evenodd"/>
            </svg>
            <div>
                <strong>Automatic Processing:</strong> This will log into NGTeco and download timecard data for the specified date range. The data will be converted to CSV format and processed automatically.
            </div>
        </div>
        
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">Choose Your Method</h2>
            </div>
            
            <form method="post">
                <div class="form-group">
                    <label for="method" class="form-label">Method</label>
                    <select id="method" name="method" class="form-select" onchange="toggleMethod()">
                        <option value="auto">Direct Login (Recommended)</option>
                        <option value="paste">Copy & Paste Table</option>
                    </select>
                </div>

                <div id="auto-section">
                    <div class="alert alert-success" style="margin-bottom:var(--spacing-3)">
                        <div>
                            <strong>Direct Login:</strong> This will log into NGTeco directly and fetch your timecard data automatically!
                            <p style="margin-top:var(--spacing-2);margin-bottom:0;color:var(--color-warning)"><strong>Note:</strong> For PythonAnywhere free accounts, you may need to request whitelisting for office.ngteco.com</p>
                        </div>
                    </div>

                    <div class="form-group">
                        <label for="username" class="form-label form-label-required">NGTeco Username</label>
                        <input type="text" id="username" name="username" class="form-input" required>
                    </div>

                    <div class="form-group">
                        <label for="password" class="form-label form-label-required">NGTeco Password</label>
                        <input type="password" id="password" name="password" class="form-input" required>
                    </div>

                    <div class="grid grid-cols-2" style="gap:var(--spacing-3)">
                        <div class="form-group">
                            <label for="start_date" class="form-label form-label-required">Start Date</label>
                            <input type="date" id="start_date" name="start_date" class="form-input" required>
                        </div>

                        <div class="form-group">
                            <label for="end_date" class="form-label form-label-required">End Date</label>
                            <input type="date" id="end_date" name="end_date" class="form-input" required>
                        </div>
                    </div>
                </div>

                <div id="paste-section" style="display: none;">
                    <div class="alert alert-info" style="margin-bottom:var(--spacing-3)">
                        <div>
                            <strong>Copy & Paste Method:</strong>
                            <ol style="margin-top:var(--spacing-2);margin-bottom:0;padding-left:var(--spacing-4)">
                                <li>Go to your NGTeco timecard page</li>
                                <li>Select your date range</li>
                                <li>Select the entire table</li>
                                <li>Copy (Ctrl+C) and paste below</li>
                            </ol>
                        </div>
                    </div>

                    <div class="form-group">
                        <label for="table_data" class="form-label form-label-required">Paste Table Data</label>
                        <textarea id="table_data" name="table_data" class="form-textarea" rows="15" placeholder="Copy the table from NGTeco and paste it here..."></textarea>
                    </div>
                </div>

                <div style="margin-top:var(--spacing-4);text-align:right">
                    <button type="submit" class="btn btn-success">
                        <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/>
                        </svg>
                        Process Timecard Data
                    </button>
                </div>
            </form>
        </div>
    </div>
    
    <script>
        function toggleMethod() {{
            var method = document.getElementById('method').value;
            var autoSection = document.getElementById('auto-section');
            var pasteSection = document.getElementById('paste-section');

            if (method === 'auto') {{
                autoSection.style.display = 'block';
                pasteSection.style.display = 'none';
                document.getElementById('username').required = true;
                document.getElementById('password').required = true;
                document.getElementById('start_date').required = true;
                document.getElementById('end_date').required = true;
                document.getElementById('table_data').required = false;
            }} else {{
                autoSection.style.display = 'none';
                pasteSection.style.display = 'block';
                document.getElementById('username').required = false;
                document.getElementById('password').required = false;
                document.getElementById('start_date').required = false;
                document.getElementById('end_date').required = false;
                document.getElementById('table_data').required = true;
            }}
        }}
    </script>
</body>
</html>
        """
        return html

    # POST - Process the request
    method = request.form.get('method', 'paste')

    try:
        if method == 'paste':
            # Parse pasted table data
            table_data = request.form.get('table_data', '')
            csv_data = parse_ngteco_table(table_data)
        else:
            # Automated fetch (requires Selenium)
            username = request.form.get('username')
            password = request.form.get('password')
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            csv_data = fetch_ngteco_automated(username, password, start_date, end_date)

        # Save to a temporary file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'timecard_auto_{timestamp}.csv'
        file_path = os.path.join(UPLOAD_FOLDER, filename)

        # Write CSV data
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            f.write(csv_data)

        # Store file path in session
        session['uploaded_file'] = file_path

        # Redirect to validation
        return redirect(url_for('validate'))

    except Exception as e:
        return f"Error processing timecard data: {str(e)}<br>Please check the format and try again.", 500


def parse_ngteco_table(table_data):
    """
    Parse copy-pasted table data from NGTeco into CSV format
    Expected columns: Person Name, Person ID, Date, Timesheet, Clock In, Clock Out, Clock Time(h), etc.
    """
    lines = table_data.strip().split('\n')
    csv_lines = []
    csv_lines.append('Person ID,First Name,Last Name,Date,Timesheet,Clock In,Clock Out,Clock Time(h),Total Break Time(h),Total Work Time(h)')

    for line in lines:
        # Skip empty lines or headers
        if not line.strip() or 'Person Name' in line or 'Clock In' in line:
            continue

        # Split by tabs or multiple spaces
        parts = re.split(r'\t+|\s{2,}', line.strip())

        if len(parts) >= 7:
            try:
                # Extract fields based on your screenshot
                person_name = parts[0]
                person_id = parts[1]
                date = parts[2]
                timesheet = parts[3] if len(parts) > 3 else 'Production TimeSheet'
                clock_in = parts[4] if len(parts) > 4 else ''
                clock_out = parts[5] if len(parts) > 5 else ''
                clock_time = parts[6] if len(parts) > 6 else ''
                break_time = parts[7] if len(parts) > 7 else ''
                work_time = parts[8] if len(parts) > 8 else clock_time

                # Split name into first and last
                name_parts = person_name.split(' ', 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ''

                                 # Convert date format to YYYY-MM-DD
                try:
                    # Handle DD-MM-YYYY format (common in NGTeco)
                    if '-' in date and len(date.split('-')[0]) == 2:
                        date_obj = datetime.strptime(date, '%d-%m-%Y')
                        date = date_obj.strftime('%Y-%m-%d')
                    # Handle YYYY-MM-DD format (already correct)
                    elif '-' in date and len(date.split('-')[0]) == 4:
                        pass  # Already in correct format
                    # Handle MM/DD/YYYY format
                    elif '/' in date:
                        date_obj = datetime.strptime(date, '%m/%d/%Y')
                        date = date_obj.strftime('%Y-%m-%d')
                except:
                    pass  # Keep original date format if parsing fails

                # Format the CSV line
                csv_line = f'{person_id},{first_name},{last_name},{date},{timesheet},{clock_in},{clock_out},{clock_time},{break_time},{work_time}'
                csv_lines.append(csv_line)

            except Exception as e:
                app.logger.error(f"Error parsing line: {line} - {str(e)}")
                continue

    return '\n'.join(csv_lines)


def fetch_ngteco_automated(username, password, start_date, end_date):
    """
    Direct fetch using requests library - works on PythonAnywhere!
    """
    import requests
    from bs4 import BeautifulSoup

    # Create a session to maintain cookies
    session = requests.Session()

    # Headers to appear like a real browser
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1'
    }
    session.headers.update(headers)

    try:
        # Step 1: Get login page to obtain any CSRF tokens
        login_url = 'https://office.ngteco.com/login'
        login_page = session.get(login_url)
        soup = BeautifulSoup(login_page.text, 'html.parser')

        # Look for CSRF token (adjust based on actual form)
        csrf_token = None
        csrf_input = soup.find('input', {'name': '_token'}) or soup.find('input', {'name': 'csrf_token'})
        if csrf_input:
            csrf_token = csrf_input.get('value')

        # Step 2: Login
        login_data = {
            'username': username,
            'password': password,
        }
        if csrf_token:
            login_data['_token'] = csrf_token

        # Post login
        login_response = session.post(login_url, data=login_data, allow_redirects=True)

        # Check if login was successful
        if 'dashboard' not in login_response.url and 'timecard' not in login_response.text:
            raise Exception("Login failed. Please check credentials.")

        # Step 3: Navigate to timecard page with date parameters
        timecard_url = f'https://office.ngteco.com/att/timecard/timecard?start_date={start_date}&end_date={end_date}'
        timecard_response = session.get(timecard_url)

        # Step 4: Parse the timecard data
        soup = BeautifulSoup(timecard_response.text, 'html.parser')

        # Try to find the table (adjust selectors based on actual HTML)
        table = soup.find('table', {'class': 'timecard-table'}) or \
                soup.find('table', {'id': 'timecard-table'}) or \
                soup.find('table')

        if not table:
            # If no table found, try to extract data from JSON or other format
            # Check if data is in a script tag
            scripts = soup.find_all('script')
            for script in scripts:
                if 'timecardData' in str(script) or 'tableData' in str(script):
                    # Extract JSON data if present
                    import json
                    import re
                    match = re.search(r'var\s+(?:timecardData|tableData)\s*=\s*(\[.*?\]);', str(script), re.DOTALL)
                    if match:
                        data = json.loads(match.group(1))
                        return convert_json_to_csv(data)

            raise Exception("Could not find timecard table in the response")

        # Parse HTML table
        csv_lines = ['Person ID,First Name,Last Name,Date,Timesheet,Clock In,Clock Out,Clock Time(h),Total Break Time(h),Total Work Time(h)']

        rows = table.find_all('tr')
        for row in rows[1:]:  # Skip header
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 7:
                # Extract text from each cell
                person_name = cells[0].get_text(strip=True)
                person_id = cells[1].get_text(strip=True)
                date = cells[2].get_text(strip=True)
                timesheet = cells[3].get_text(strip=True) if len(cells) > 3 else 'Production TimeSheet'
                clock_in = cells[4].get_text(strip=True) if len(cells) > 4 else ''
                clock_out = cells[5].get_text(strip=True) if len(cells) > 5 else ''
                clock_time = cells[6].get_text(strip=True) if len(cells) > 6 else ''
                break_time = cells[7].get_text(strip=True) if len(cells) > 7 else ''
                work_time = cells[8].get_text(strip=True) if len(cells) > 8 else clock_time

                # Split name
                name_parts = person_name.split(' ', 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ''

                # Convert date format
                try:
                    if '-' in date and len(date.split('-')[0]) == 2:
                        date_obj = datetime.strptime(date, '%d-%m-%Y')
                        date = date_obj.strftime('%Y-%m-%d')
                except:
                    pass

                csv_lines.append(f'{person_id},{first_name},{last_name},{date},{timesheet},{clock_in},{clock_out},{clock_time},{break_time},{work_time}')

        return '\n'.join(csv_lines)

    except requests.exceptions.RequestException as e:
        raise Exception(f"Network error: {str(e)}")
    except Exception as e:
        raise Exception(f"Error fetching data: {str(e)}")


def convert_json_to_csv(data):
    """Convert JSON data to CSV format"""
    csv_lines = ['Person ID,First Name,Last Name,Date,Timesheet,Clock In,Clock Out,Clock Time(h),Total Break Time(h),Total Work Time(h)']

    for record in data:
        # Adjust field names based on actual JSON structure
        person_name = record.get('employee_name', '')
        person_id = record.get('employee_id', '')
        date = record.get('date', '')
        clock_in = record.get('clock_in', '')
        clock_out = record.get('clock_out', '')
        work_time = record.get('work_hours', '')

        name_parts = person_name.split(' ', 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ''

        csv_lines.append(f'{person_id},{first_name},{last_name},{date},Production TimeSheet,{clock_in},{clock_out},,{work_time}')

    return '\n'.join(csv_lines)



@app.route('/confirm_employees')
@login_required
def confirm_employees():
    """Show employee confirmation page before processing"""
    try:
        username = session.get('username', 'Unknown')
        is_admin = username == 'admin'
        
        file_path = session.get('uploaded_file')
        if not file_path:
            return "No file found in session. Please upload again.", 400
        
        df = pd.read_csv(file_path)
        employees = get_unique_employees_from_df(df)
        employees_json = json.dumps(employees)
        
        menu_html = get_menu_html(username)
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Confirm Employees - Payroll Management</title>
    <link rel="icon" type="image/svg+xml" href="/static/favicon.svg">
    <link rel="stylesheet" href="/static/design-system.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        .confirm-header {{
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            color: white;
            padding: var(--spacing-4) 0;
            margin-bottom: var(--spacing-4);
        }}
        .employee-item {{
            padding: var(--spacing-3);
            margin-bottom: var(--spacing-2);
            background: var(--color-gray-50);
            border-radius: var(--radius-md);
            display: flex;
            align-items: center;
            gap: var(--spacing-3);
            transition: background var(--transition-fast);
        }}
        .employee-item:hover {{
            background: var(--color-gray-100);
        }}
        .employee-item input[type="checkbox"] {{
            width: 18px;
            height: 18px;
            cursor: pointer;
        }}
    </style>
</head>
<body>
    {menu_html}
    
    <div class="confirm-header">
        <div class="container">
            <h1 style="color:white;margin-bottom:var(--spacing-1);font-size:var(--font-size-2xl)">Confirm Employees for Payroll</h1>
            <p style="color:rgba(255,255,255,0.9);font-size:var(--font-size-sm);margin:0">Select employees to include in this payroll run</p>
        </div>
    </div>
    
    <div class="container container-narrow">
        <div class="card">
            <div class="card-header">
                <h2 class="card-title">Select Employees to Include</h2>
            </div>
            
            <div id="employee-list"></div>
            
            <div style="margin-top:var(--spacing-4);display:flex;gap:var(--spacing-3);justify-content:flex-end">
                <a href="/" class="btn btn-secondary">Cancel</a>
                <button id="process-btn" onclick="processPayroll(this)" class="btn btn-success">
                    <svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd"/>
                    </svg>
                    Confirm & Process
                </button>
            </div>
        </div>
    </div>
    
    <script>
        const employees = {employees_json};
        function populateEmployees() {{
            const list = document.getElementById('employee-list');
            employees.forEach(emp => {{
                const div = document.createElement('div');
                div.className = 'employee-item';
                div.innerHTML = '<input type="checkbox" value="' + emp['Person ID'] + '" checked> <strong>' + escape(emp['First Name'] + ' ' + emp['Last Name']) + '</strong> <span style="color:var(--color-gray-600);font-size:var(--font-size-sm)">(ID: ' + escape(emp['Person ID']) + ')</span>';
                list.appendChild(div);
            }});
        }}
        function escape(str) {{
            const div = document.createElement('div');
            div.textContent = str;
            return div.innerHTML;
        }}
        function processPayroll(button) {{
            const checkboxes = document.querySelectorAll('input[type="checkbox"]:checked');
            const selectedIds = Array.from(checkboxes).map(cb => cb.value);
            if (selectedIds.length === 0) {{
                alert('Select at least one employee.');
                return;
            }}
            
            const originalText = button.innerHTML;
            button.disabled = true;
            button.innerHTML = '<svg style="width:20px;height:20px" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-8.707l-3-3a1 1 0 00-1.414 0l-3 3a1 1 0 001.414 1.414L9 9.414V13a1 1 0 102 0V9.414l1.293 1.293a1 1 0 001.414-1.414z" clip-rule="evenodd"/></svg> Processing...';
            
            fetch('/process_confirmed', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{employee_ids: selectedIds}})
            }})
            .then(response => {{
                if (response.ok) {{
                    return response.json();
                }} else {{
                    return response.json().then(data => {{
                        throw new Error(data.error || 'Error processing payroll');
                    }}).catch(() => {{
                        throw new Error('Error processing payroll. Please try again.');
                    }});
                }}
            }})
            .then(data => {{
                if (data.redirect) {{
                    window.location.href = data.redirect;
                }} else {{
                    window.location.href = '/success';
                }}
            }})
            .catch(error => {{
                button.disabled = false;
                button.innerHTML = originalText;
                alert('Error: ' + (error.message || 'Error processing payroll. Please try again.'));
            }});
        }}
        populateEmployees();
    </script>
</body>
</html>"""
        
        return html
    except Exception as e:
        import traceback
        return f"Error: {str(e)}<br><pre>{traceback.format_exc()}</pre>", 500

@app.route('/confirm_and_process', methods=['POST'])
@login_required
def confirm_and_process():
    """Store confirmed employee IDs"""
    try:
        data = request.get_json()
        session['confirmed_employee_ids'] = data.get('employee_ids', [])
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/process_confirmed', methods=['GET', 'POST'])
@login_required
def process_confirmed():
    """Process payroll for confirmed employees only"""
    try:
        # Handle POST request from confirm_employees page
        if request.method == 'POST':
            try:
                data = request.get_json()
                employee_ids = data.get('employee_ids', [])
                if employee_ids:
                    session['confirmed_employee_ids'] = employee_ids
                else:
                    return jsonify({'error': 'No employees selected'}), 400
            except Exception as e:
                app.logger.error(f"Error parsing JSON in process_confirmed: {str(e)}")
                return jsonify({'error': 'Invalid request data'}), 400
        
        file_path = session.get('uploaded_file')
        confirmed_ids = session.get('confirmed_employee_ids', [])
        
        if not file_path:
            if request.method == 'POST':
                return jsonify({'error': 'No file found. Please upload again.'}), 400
            return "No file found. Please upload again.", 400
        
        df = pd.read_csv(file_path)
        
        # Filter to only selected employees
        if confirmed_ids:
            df = df[df['Person ID'].astype(str).isin(confirmed_ids)]
            # Save filtered CSV for Zoho summary
            filtered_path = file_path.replace('.csv', '_filtered.csv')
            df.to_csv(filtered_path, index=False)
            session['filtered_file'] = filtered_path
        else:
            session['filtered_file'] = file_path
        # NO VALIDATION HERE - just process directly
        username = session.get('username', 'Unknown')
        
        is_timesheet = all(col in df.columns for col in ['Person ID', 'First Name', 'Last Name', 'Date'])
        
        if is_timesheet:
            try:
                df['Date'] = pd.to_datetime(df['Date'])
                week_str = df['Date'].min().strftime('%Y-%m-%d')
            except:
                week_str = datetime.now().strftime('%Y-%m-%d')
        else:
            week_str = datetime.now().strftime('%Y-%m-%d')
        
        reports = {}
        
        summary_filename = f"payroll_summary_{week_str}.xlsx"
        summary_path = create_excel_report(df, summary_filename, username)
        reports['summary'] = summary_filename
        
        if is_timesheet:
            payslips_filename = f"employee_payslips_{week_str}.xlsx"
            payslips_path = create_payslips(df, payslips_filename, username)
            reports['payslips'] = payslips_filename
            
            admin_filename = f"admin_report_{week_str}.xlsx"
            admin_path = create_consolidated_admin_report(df, admin_filename, username)
            reports['admin'] = admin_filename
            
            payslip_filename = f"payslips_for_cutting_{week_str}.xlsx"
            payslip_path = create_consolidated_payslips(df, payslip_filename, username)
            reports['payslips_sheet'] = payslip_filename
        
        session['reports'] = reports
        session['week'] = week_str
        session.pop('confirmed_employee_ids', None)
        
        # For POST requests (from JavaScript), return JSON
        if request.method == 'POST':
            return jsonify({'status': 'ok', 'redirect': url_for('success')})
        
        return redirect(url_for('success'))
        
    except Exception as e:
        import traceback
        app.logger.error(f"Error in process_confirmed: {str(e)}\n{traceback.format_exc()}")
        
        # For POST requests, return JSON error
        if request.method == 'POST':
            return jsonify({'error': f'Error processing payroll: {str(e)}'}), 500
        
        txt_filename = "error_report.txt"
        report_path = os.path.join(REPORT_FOLDER, txt_filename)
        with open(report_path, 'w') as f:
            f.write(f"Error: {str(e)}\n")
            f.write(traceback.format_exc())
        session['reports'] = {'error': txt_filename}
        return redirect(url_for('success'))


# ═══════════════════════════════════════════════════════════════════════════════
# ERROR HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════
# Custom error pages with consistent UI design

@app.errorhandler(404)
def page_not_found(e):
    """Custom 404 error page - Page not found"""
    app.logger.warning(f"404 Error: {request.url}")
    return render_template_string('''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>404 - Page Not Found | Payroll System</title>
            <script src="https://cdn.tailwindcss.com"></script>
            <link rel="preconnect" href="https://fonts.googleapis.com">
            <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
            <script>
                tailwind.config = {
                    theme: {
                        extend: {
                            fontFamily: {
                                sans: ['Inter', 'sans-serif'],
                            },
                            colors: {
                                primary: '#1e40af',
                                secondary: '#64748b',
                                accent: '#3b82f6',
                                textDark: '#1e293b',
                            }
                        }
                    }
                }
            </script>
        </head>
        <body class="bg-gradient-to-br from-slate-50 to-blue-50 min-h-screen flex items-center justify-center font-sans">
            <div class="max-w-2xl mx-auto px-6 py-12 text-center">
                <div class="bg-white rounded-2xl shadow-2xl p-12">
                    <!-- 404 Icon -->
                    <div class="mb-8">
                        <svg class="w-32 h-32 mx-auto text-blue-100" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7 4a1 1 0 11-2 0 1 1 0 012 0zm-1-9a1 1 0 00-1 1v4a1 1 0 102 0V6a1 1 0 00-1-1z" clip-rule="evenodd" />
                        </svg>
                    </div>
                    
                    <!-- Error Message -->
                    <h1 class="text-6xl font-bold text-primary mb-4">404</h1>
                    <h2 class="text-2xl font-semibold text-textDark mb-4">Page Not Found</h2>
                    <p class="text-lg text-secondary mb-8">
                        The page you're looking for doesn't exist or has been moved.
                    </p>
                    
                    <!-- Action Buttons -->
                    <div class="flex flex-col sm:flex-row gap-4 justify-center">
                        <a href="/" class="px-8 py-3 bg-gradient-to-r from-primary to-blue-700 text-white font-semibold rounded-lg shadow-lg hover:shadow-xl hover:from-primary/90 hover:to-blue-600 transition-all transform hover:-translate-y-0.5">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6" />
                            </svg>
                            Go Home
                        </a>
                        <button onclick="window.history.back()" class="px-8 py-3 bg-white text-primary border-2 border-primary font-semibold rounded-lg hover:bg-primary hover:text-white transition-all">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10 19l-7-7m0 0l7-7m-7 7h18" />
                            </svg>
                            Go Back
                        </button>
                    </div>
                </div>
                
                <!-- Footer -->
                <div class="mt-8 text-sm text-secondary">
                    Payroll Management System {{ version }}
                </div>
            </div>
        </body>
        </html>
    ''', version=get_version_display()), 404

@app.errorhandler(500)
def internal_server_error(e):
    """Custom 500 error page - Internal server error"""
    app.logger.error(f"500 Error: {str(e)}", exc_info=True)
    return render_template_string('''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>500 - Server Error | Payroll System</title>
            <script src="https://cdn.tailwindcss.com"></script>
            <link rel="preconnect" href="https://fonts.googleapis.com">
            <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
            <script>
                tailwind.config = {
                    theme: {
                        extend: {
                            fontFamily: {
                                sans: ['Inter', 'sans-serif'],
                            },
                            colors: {
                                primary: '#1e40af',
                                secondary: '#64748b',
                                accent: '#3b82f6',
                                textDark: '#1e293b',
                            }
                        }
                    }
                }
            </script>
        </head>
        <body class="bg-gradient-to-br from-slate-50 to-blue-50 min-h-screen flex items-center justify-center font-sans">
            <div class="max-w-2xl mx-auto px-6 py-12 text-center">
                <div class="bg-white rounded-2xl shadow-2xl p-12">
                    <!-- Error Icon -->
                    <div class="mb-8">
                        <svg class="w-32 h-32 mx-auto text-red-100" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zM8.707 7.293a1 1 0 00-1.414 1.414L8.586 10l-1.293 1.293a1 1 0 101.414 1.414L10 11.414l1.293 1.293a1 1 0 001.414-1.414L11.414 10l1.293-1.293a1 1 0 00-1.414-1.414L10 8.586 8.707 7.293z" clip-rule="evenodd" />
                        </svg>
                    </div>
                    
                    <!-- Error Message -->
                    <h1 class="text-6xl font-bold text-red-600 mb-4">500</h1>
                    <h2 class="text-2xl font-semibold text-textDark mb-4">Internal Server Error</h2>
                    <p class="text-lg text-secondary mb-8">
                        Something went wrong on our end. We've logged the error and will look into it.
                    </p>
                    
                    <!-- Help Text -->
                    <div class="bg-blue-50 border-l-4 border-blue-400 p-4 mb-8 text-left">
                        <div class="flex">
                            <div class="flex-shrink-0">
                                <svg class="h-5 w-5 text-blue-400" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7-4a1 1 0 11-2 0 1 1 0 012 0zM9 9a1 1 0 000 2v3a1 1 0 001 1h1a1 1 0 100-2v-3a1 1 0 00-1-1H9z" clip-rule="evenodd" />
                                </svg>
                            </div>
                            <div class="ml-3">
                                <p class="text-sm text-blue-700">
                                    <strong>What you can do:</strong><br>
                                    • Try refreshing the page<br>
                                    • Go back and try again<br>
                                    • If the problem persists, contact your administrator
                                </p>
                            </div>
                        </div>
                    </div>
                    
                    <!-- Action Buttons -->
                    <div class="flex flex-col sm:flex-row gap-4 justify-center">
                        <a href="/" class="px-8 py-3 bg-gradient-to-r from-primary to-blue-700 text-white font-semibold rounded-lg shadow-lg hover:shadow-xl hover:from-primary/90 hover:to-blue-600 transition-all transform hover:-translate-y-0.5">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6" />
                            </svg>
                            Go Home
                        </a>
                        <button onclick="window.history.back()" class="px-8 py-3 bg-white text-primary border-2 border-primary font-semibold rounded-lg hover:bg-primary hover:text-white transition-all">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10 19l-7-7m0 0l7-7m-7 7h18" />
                            </svg>
                            Go Back
                        </button>
                    </div>
                </div>
                
                <!-- Footer -->
                <div class="mt-8 text-sm text-secondary">
                    Payroll Management System {{ version }}
                </div>
            </div>
        </body>
        </html>
    ''', version=get_version_display()), 500

@app.errorhandler(403)
def forbidden(e):
    """Custom 403 error page - Forbidden/Unauthorized access"""
    app.logger.warning(f"403 Error: {request.url} - User: {session.get('username', 'Anonymous')}")
    return render_template_string('''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>403 - Access Denied | Payroll System</title>
            <script src="https://cdn.tailwindcss.com"></script>
            <link rel="preconnect" href="https://fonts.googleapis.com">
            <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
            <script>
                tailwind.config = {
                    theme: {
                        extend: {
                            fontFamily: {
                                sans: ['Inter', 'sans-serif'],
                            },
                            colors: {
                                primary: '#1e40af',
                                secondary: '#64748b',
                                accent: '#3b82f6',
                                textDark: '#1e293b',
                            }
                        }
                    }
                }
            </script>
        </head>
        <body class="bg-gradient-to-br from-slate-50 to-blue-50 min-h-screen flex items-center justify-center font-sans">
            <div class="max-w-2xl mx-auto px-6 py-12 text-center">
                <div class="bg-white rounded-2xl shadow-2xl p-12">
                    <!-- Lock Icon -->
                    <div class="mb-8">
                        <svg class="w-32 h-32 mx-auto text-yellow-100" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M5 9V7a5 5 0 0110 0v2a2 2 0 012 2v5a2 2 0 01-2 2H5a2 2 0 01-2-2v-5a2 2 0 012-2zm8-2v2H7V7a3 3 0 016 0z" clip-rule="evenodd" />
                        </svg>
                    </div>
                    
                    <!-- Error Message -->
                    <h1 class="text-6xl font-bold text-yellow-600 mb-4">403</h1>
                    <h2 class="text-2xl font-semibold text-textDark mb-4">Access Denied</h2>
                    <p class="text-lg text-secondary mb-8">
                        You don't have permission to access this resource.
                    </p>
                    
                    <!-- Help Text -->
                    <div class="bg-yellow-50 border-l-4 border-yellow-400 p-4 mb-8 text-left">
                        <div class="flex">
                            <div class="flex-shrink-0">
                                <svg class="h-5 w-5 text-yellow-400" fill="currentColor" viewBox="0 0 20 20">
                                    <path fill-rule="evenodd" d="M8.257 3.099c.765-1.36 2.722-1.36 3.486 0l5.58 9.92c.75 1.334-.213 2.98-1.742 2.98H4.42c-1.53 0-2.493-1.646-1.743-2.98l5.58-9.92zM11 13a1 1 0 11-2 0 1 1 0 012 0zm-1-8a1 1 0 00-1 1v3a1 1 0 002 0V6a1 1 0 00-1-1z" clip-rule="evenodd" />
                                </svg>
                            </div>
                            <div class="ml-3">
                                <p class="text-sm text-yellow-700">
                                    <strong>Possible reasons:</strong><br>
                                    • You need to log in to access this page<br>
                                    • Your session may have expired<br>
                                    • You don't have the required permissions
                                </p>
                            </div>
                        </div>
                    </div>
                    
                    <!-- Action Buttons -->
                    <div class="flex flex-col sm:flex-row gap-4 justify-center">
                        <a href="/login" class="px-8 py-3 bg-gradient-to-r from-primary to-blue-700 text-white font-semibold rounded-lg shadow-lg hover:shadow-xl hover:from-primary/90 hover:to-blue-600 transition-all transform hover:-translate-y-0.5">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M11 16l-4-4m0 0l4-4m-4 4h14m-5 4v1a3 3 0 01-3 3H6a3 3 0 01-3-3V7a3 3 0 013-3h7a3 3 0 013 3v1" />
                            </svg>
                            Go to Login
                        </a>
                        <a href="/" class="px-8 py-3 bg-white text-primary border-2 border-primary font-semibold rounded-lg hover:bg-primary hover:text-white transition-all">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6" />
                            </svg>
                            Go Home
                        </a>
                    </div>
                </div>
                
                <!-- Footer -->
                <div class="mt-8 text-sm text-secondary">
                    Payroll Management System {{ version }}
                </div>
            </div>
        </body>
        </html>
    ''', version=get_version_display()), 403

@app.errorhandler(405)
def method_not_allowed(e):
    """Custom 405 error page - Method not allowed"""
    app.logger.warning(f"405 Error: {request.method} {request.url}")
    return render_template_string('''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>405 - Method Not Allowed | Payroll System</title>
            <script src="https://cdn.tailwindcss.com"></script>
            <link rel="preconnect" href="https://fonts.googleapis.com">
            <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
            <script>
                tailwind.config = {
                    theme: {
                        extend: {
                            fontFamily: {
                                sans: ['Inter', 'sans-serif'],
                            },
                            colors: {
                                primary: '#1e40af',
                                secondary: '#64748b',
                                accent: '#3b82f6',
                                textDark: '#1e293b',
                            }
                        }
                    }
                }
            </script>
        </head>
        <body class="bg-gradient-to-br from-slate-50 to-blue-50 min-h-screen flex items-center justify-center font-sans">
            <div class="max-w-2xl mx-auto px-6 py-12 text-center">
                <div class="bg-white rounded-2xl shadow-2xl p-12">
                    <!-- Error Icon -->
                    <div class="mb-8">
                        <svg class="w-32 h-32 mx-auto text-orange-100" fill="currentColor" viewBox="0 0 20 20">
                            <path fill-rule="evenodd" d="M13.477 14.89A6 6 0 015.11 6.524l8.367 8.368zm1.414-1.414L6.524 5.11a6 6 0 018.367 8.367zM18 10a8 8 0 11-16 0 8 8 0 0116 0z" clip-rule="evenodd" />
                        </svg>
                    </div>
                    
                    <!-- Error Message -->
                    <h1 class="text-6xl font-bold text-orange-600 mb-4">405</h1>
                    <h2 class="text-2xl font-semibold text-textDark mb-4">Method Not Allowed</h2>
                    <p class="text-lg text-secondary mb-8">
                        The request method is not supported for this resource.
                    </p>
                    
                    <!-- Action Buttons -->
                    <div class="flex flex-col sm:flex-row gap-4 justify-center">
                        <a href="/" class="px-8 py-3 bg-gradient-to-r from-primary to-blue-700 text-white font-semibold rounded-lg shadow-lg hover:shadow-xl hover:from-primary/90 hover:to-blue-600 transition-all transform hover:-translate-y-0.5">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6" />
                            </svg>
                            Go Home
                        </a>
                        <button onclick="window.history.back()" class="px-8 py-3 bg-white text-primary border-2 border-primary font-semibold rounded-lg hover:bg-primary hover:text-white transition-all">
                            <svg class="w-5 h-5 inline-block mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10 19l-7-7m0 0l7-7m-7 7h18" />
                            </svg>
                            Go Back
                        </button>
                    </div>
                </div>
                
                <!-- Footer -->
                <div class="mt-8 text-sm text-secondary">
                    Payroll Management System {{ version }}
                </div>
            </div>
        </body>
        </html>
    ''', version=get_version_display()), 405


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
